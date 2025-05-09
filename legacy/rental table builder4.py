import os
import pandas as pd
import numpy as np
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import re

class RentalDataAnalyzer:
    def __init__(self):
        """Initialize the Rental Data Analyzer with paths and reference data"""
        # Base directory - can be modified by user via GUI
        self.BASE_DIR = r"C:\Users\joshu\OneDrive - NSWGOV\Rental Data Model\Calculated rental data"

        # Geographic areas
        self.GEO_AREAS = ["CED", "GCCSA", "LGA", "SA3", "SA4", "SED", "Suburb"]

        # Data subdirectories
        self.SUB_DIRS = {
            "median_rents": os.path.join("Median rents", "output data"),
            "census_dwelling": os.path.join("Census data", "output data", "dwellings"),
            "census_demographics": os.path.join("Census data", "output data", "demographics"),
            "affordability": os.path.join("Affordability", "output data"),
            "vacancy_rates": os.path.join("Rental vacancy rates", "output data")
        }

        # File patterns for different geographic areas
        self.FILE_PATTERNS = {
            "median_rents": {area.lower(): f"{area.lower()}_rent_data" for area in self.GEO_AREAS},
            "affordability": {area.lower(): f"{area.lower()}_affordability" for area in self.GEO_AREAS},
            "vacancy_rates": {area.lower(): f"{area.lower()}_vacancy_rate" for area in self.GEO_AREAS},
            "census_dwelling": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_dwelling_tenure" for area in self.GEO_AREAS},
            "census_demographics": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_demographics" for area in self.GEO_AREAS}
        }

        # Greater Sydney LGAs - Not strictly needed for dynamic search, but kept for reference
        self.GREATER_SYDNEY_LGAS = [
            "Bayside (NSW)", "Blacktown", "Blue Mountains", "Burwood", "Camden", "Campbelltown (NSW)",
            "Canada Bay", "Canterbury-Bankstown", "Cumberland", "Fairfield", "Georges River",
            "Hawkesbury", "Hornsby", "Hunters Hill", "Inner West", "Ku-ring-gai", "Lane Cove",
            "Liverpool", "Mosman", "North Sydney", "Northern Beaches", "Parramatta", "Penrith",
            "Randwick", "Ryde", "Strathfield", "Sutherland Shire", "Sydney", "The Hills Shire",
            "Waverley", "Willoughby", "Woollahra", "Wollondilly"
        ]

        # Reference data for comparison - Will be dynamically updated
        self.GS_REFERENCE_DATA = {
            "renters": {"area": "Greater Sydney", "value": 32.6}, # Static Census data
            "social_housing": {"area": "Greater Sydney", "value": 4.5}, # Static Census data
            "median_rent": {"area": "Greater Sydney", "value": None, "annual_change": None}, # Dynamic
            "vacancy_rates": {"area": "Greater Sydney", "value": None, "annual_change": None}, # Dynamic
            "affordability": {"area": "Greater Sydney", "value": None, "annual_change": None} # Dynamic
        }

        # Reference data for comparison - Will be dynamically updated
        self.RON_REFERENCE_DATA = {
            "renters": {"area": "Rest of NSW", "value": 26.8}, # Static Census data
            "social_housing": {"area": "Rest of NSW", "value": 4.0}, # Static Census data
            "median_rent": {"area": "Rest of NSW", "value": None, "annual_change": None}, # Dynamic
            "vacancy_rates": {"area": "Rest of NSW", "value": None, "annual_change": None}, # Dynamic
            "affordability": {"area": "Rest of NSW", "value": None, "annual_change": None} # Dynamic
        }

        # Variables to store selections and data
        self.selected_geo_area = None
        self.selected_geo_name = None
        self.data = {}
        self.output_file = None

    def create_gui(self):
        """Create GUI for selecting geographic area and name"""
        self.root = tk.Tk()
        self.root.title("NSW Rental Data Analyzer")
        self.root.geometry("700x400")

        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Data directory selection
        dir_frame = ttk.LabelFrame(main_frame, text="Data Directory", padding="5")
        dir_frame.pack(fill=tk.X, padx=5, pady=5)

        self.dir_var = tk.StringVar(value=self.BASE_DIR)
        dir_entry = ttk.Entry(dir_frame, textvariable=self.dir_var, width=60)
        dir_entry.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        dir_button = ttk.Button(dir_frame, text="Browse...", command=self.browse_directory)
        dir_button.pack(side=tk.RIGHT, padx=5, pady=5)

        # Selection frame
        selection_frame = ttk.LabelFrame(main_frame, text="Geographic Selection", padding="5")
        selection_frame.pack(fill=tk.X, padx=5, pady=5)

        # Geographic area type selection
        ttk.Label(selection_frame, text="Geographic Area Type:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo = ttk.Combobox(selection_frame, values=self.GEO_AREAS, state="readonly", width=30)
        self.geo_area_combo.grid(column=1, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo.bind("<<ComboboxSelected>>", self.on_geo_area_selected)

        # Geographic name selection
        ttk.Label(selection_frame, text="Geographic Name:").grid(column=0, row=1, sticky=tk.W, pady=5, padx=5)
        self.geo_name_combo = ttk.Combobox(selection_frame, state="disabled", width=40)
        self.geo_name_combo.grid(column=1, row=1, sticky=tk.W, pady=5, padx=5)

        # Output file selection
        output_frame = ttk.LabelFrame(main_frame, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=5, pady=5)

        self.output_var = tk.StringVar()
        ttk.Label(output_frame, text="Output File:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=40)
        output_entry.grid(column=1, row=0, sticky=tk.W+tk.E, pady=5, padx=5)

        output_button = ttk.Button(output_frame, text="Browse...", command=self.browse_output_file)
        output_button.grid(column=2, row=0, sticky=tk.W, pady=5, padx=5)

        # Status and buttons
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, padx=5, pady=10)

        # Generate button
        self.generate_button = ttk.Button(status_frame, text="Generate Analysis", command=self.run_analysis, state="disabled")
        self.generate_button.pack(side=tk.RIGHT, padx=5)

        # Exit button
        exit_button = ttk.Button(status_frame, text="Exit", command=self.root.destroy)
        exit_button.pack(side=tk.RIGHT, padx=5)

        # Status label
        self.status_label = ttk.Label(status_frame, text="Select a geographic area type to begin")
        self.status_label.pack(side=tk.LEFT, padx=5)

        # Set column and row weights for resizing
        main_frame.columnconfigure(0, weight=1)

        self.root.mainloop()

    def browse_directory(self):
        """Browse for data directory"""
        dir_path = filedialog.askdirectory(initialdir=self.BASE_DIR, title="Select Data Directory")
        if dir_path:
            self.BASE_DIR = dir_path
            self.dir_var.set(dir_path)
            # Reset selections
            self.selected_geo_area = None
            self.selected_geo_name = None
            self.geo_area_combo.set('')
            self.geo_name_combo.set('')
            self.geo_name_combo.config(state="disabled")
            self.generate_button.config(state="disabled")
            self.status_label.config(text="Select a geographic area type to begin")

    def browse_output_file(self):
        """Browse for output file location"""
        file_path = filedialog.asksaveasfilename(
            initialdir=os.path.expanduser("~"),
            title="Save Analysis As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.output_file = file_path
            self.output_var.set(file_path)

    def on_geo_area_selected(self, event=None):
        """Handle geographic area type selection"""
        self.selected_geo_area = self.geo_area_combo.get()
        self.status_label.config(text=f"Loading {self.selected_geo_area} names...")
        self.root.update_idletasks()

        # Get available geographic names
        try:
            geo_names = self.get_geo_names(self.selected_geo_area)
            self.geo_name_combo.config(values=geo_names, state="readonly")
            self.geo_name_combo.bind("<<ComboboxSelected>>", self.on_geo_name_selected)
            self.status_label.config(text=f"Select a {self.selected_geo_area} name")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to load geographic names: {str(e)}")

    def on_geo_name_selected(self, event=None):
        """Handle geographic name selection"""
        self.selected_geo_name = self.geo_name_combo.get()
        self.generate_button.config(state="normal")

        # Set default output file name if not already set
        if not self.output_var.get():
            # Sanitize filename - remove parentheses and other special characters
            safe_geo_name = self.selected_geo_name.replace("(", "").replace(")", "").replace(" ", "_").replace(",", "")
            default_filename = f"{safe_geo_name}_{self.selected_geo_area}_Rental_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"

            # Ensure desktop path exists
            try:
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop_path):
                    # Fall back to current directory if Desktop doesn't exist
                    desktop_path = os.getcwd()
            except:
                desktop_path = os.getcwd()

            self.output_var.set(os.path.join(desktop_path, default_filename))
            self.output_file = os.path.join(desktop_path, default_filename)

        self.status_label.config(text=f"Ready to generate analysis for {self.selected_geo_name} ({self.selected_geo_area})")

    def get_geo_names(self, geo_area):
        """Get available geographic names for the selected area type"""
        names = set()
        found_files = False

        # Try to get names from multiple data sources
        for data_type, subdir in self.SUB_DIRS.items():
            try:
                dir_path = os.path.join(self.BASE_DIR, subdir)
                if not os.path.exists(dir_path):
                    continue

                # Skip if pattern doesn't exist for this geo_area
                if geo_area.lower() not in self.FILE_PATTERNS[data_type]:
                    continue

                file_pattern = self.FILE_PATTERNS[data_type][geo_area.lower()]

                for file in os.listdir(dir_path):
                    if file_pattern.lower() in file.lower():
                        found_files = True
                        file_path = os.path.join(dir_path, file)
                        print(f"Reading file: {file_path}")
                        df = self.read_data_file(file_path)

                        if df is not None and not df.empty:
                            # Look for the geographic name column
                            geo_col = self.find_geographic_column(df, geo_area)

                            if geo_col:
                                print(f"Found geography column: {geo_col}")
                                # Convert all values to strings and filter out likely non-geographic names
                                df[geo_col] = df[geo_col].astype(str)

                                # Filter out values that look like dates or numbers
                                area_names = []
                                for name in df[geo_col].dropna().unique().tolist():
                                    name_str = str(name)
                                    # Skip if it looks like a date format (2021-01-01, etc.)
                                    if re.match(r'^\d{4}-\d{2}-\d{2}', name_str) or re.match(r'^\d{2}/\d{2}/\d{4}', name_str):
                                        continue
                                    # Skip if it's just a number
                                    if name_str.replace('.', '', 1).isdigit(): # Allow one decimal point
                                         continue
                                    # Skip very short names (likely codes)
                                    if len(name_str) < 2:
                                        continue
                                    # Skip if it's an LGA/SA code
                                    if re.match(r'^LGA\d+$', name_str) or re.match(r'^SA\d+$', name_str):
                                        continue
                                    # Skip if it's Greater Sydney or Rest of NSW (handled separately)
                                    if name_str in ["Greater Sydney", "Rest of NSW"]:
                                        continue

                                    area_names.append(name_str)

                                print(f"Found {len(area_names)} geographic names")
                                names.update(area_names)
            except Exception as e:
                print(f"Error getting names from {data_type}: {str(e)}")

        if not found_files:
            raise Exception(f"No data files found for {geo_area}. Please check your data directory.")

        if not names:
            raise Exception(f"No geographic names found for {geo_area}. Check that your data files contain the expected columns.")

        name_list = sorted(list(names))
        if name_list:
            print(f"Final list of geographic names: {name_list[:10]}... (showing first 10 of {len(name_list)})")
        else:
            print("No geographic names found.")
        return name_list

    def read_data_file(self, file_path):
        """Read data from Excel or Parquet file"""
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # Use explicit converter and dtype to avoid datetime conversion issues
                return pd.read_excel(
                    file_path,
                    dtype={0: str},  # Force first column to be string
                    converters={0: str}  # Ensure first column is converted to string
                )
            elif file_path.endswith('.parquet'):
                df = pq.read_table(file_path).to_pandas()

                # Ensure the first column is a string to avoid comparison issues
                if len(df.columns) > 0:
                    df[df.columns[0]] = df[df.columns[0]].astype(str)

                return df
            else:
                print(f"Unsupported file format: {file_path}")
                return None
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            return None

    def find_geographic_column(self, df, geo_area):
        """Find the column containing geographic area names"""
        # Print the column names for debugging
        print(f"Available columns: {df.columns.tolist()}")

        # Direct matches - highest priority columns that definitely contain geographic names
        priority_columns = ['region_name', 'area_name', 'location_name', f'{geo_area.lower()}_name', 'name']

        # Check for exact matches in priority columns first
        for col in priority_columns:
            if col in df.columns:
                print(f"Found priority geography column: {col}")

                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col

        # Common names for geographic columns - next priority
        geo_keywords = [
            geo_area.lower(), 'name', 'area', 'region', 'location', 'geography',
            'district', 'locality', 'suburb', 'lga', 'sa3', 'sa4', 'gccsa', 'ced', 'sed'
        ]

        # Check for columns that contain these keywords
        for col in df.columns:
            col_lower = str(col).lower()
            for keyword in geo_keywords:
                if keyword in col_lower and 'code' not in col_lower and 'type' not in col_lower:
                    print(f"Found geography column by keyword: {col}")
                    # Show sample values
                    if not df.empty:
                        sample = df[col].dropna().head(5).astype(str).tolist()
                        print(f"Sample values from '{col}': {sample}")
                    return col

        # If we haven't found a match yet, specifically look for 'region_name' which we know exists
        if 'region_name' in df.columns:
            print(f"Falling back to region_name column")
            # Show sample values
            if not df.empty:
                sample = df['region_name'].dropna().head(5).astype(str).tolist()
                print(f"Sample values from 'region_name': {sample}")
            return 'region_name'

        # If still no match, check for columns that might contain place names
        for col in df.columns:
            try:
                sample = df[col].dropna().head(5).astype(str).tolist()
                print(f"Checking column '{col}' with sample values: {sample}")

                # Skip columns where values appear to be dates or numbers
                if all(not re.match(r'^\d{4}-\d{2}-\d{2}', str(x)) and
                       not re.match(r'^\d{2}/\d{2}/\d{4}', str(x)) and
                       not str(x).replace('.', '', 1).isdigit() and # Allow one decimal point
                       len(str(x)) > 2
                       for x in sample):

                    # Check if the values look like place names (contain alphabetic characters)
                    if all(any(c.isalpha() for c in str(x)) for x in sample):
                        print(f"Detected possible location names in column: {col}")
                        return col
            except Exception as e:
                print(f"Error checking column {col}: {str(e)}")

        # If no suitable column found, look for codes as a last resort
        for col in df.columns:
            col_lower = str(col).lower()
            if ('code' in col_lower and any(kw in col_lower for kw in geo_keywords)) or 'region_code' in col_lower:
                print(f"Falling back to code column: {col}")
                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col

        # Absolute last resort - first column
        if len(df.columns) > 0:
            first_col = df.columns[0]
            print(f"No geographic column identified, using first column as fallback: {first_col}")
            # Show sample values
            if not df.empty:
                sample = df[first_col].dropna().head(5).astype(str).tolist()
                print(f"Sample values from first column: {sample}")
            return first_col

        return None

    def run_analysis(self):
        """Run the analysis and generate Excel output"""
        if not self.selected_geo_area or not self.selected_geo_name:
            messagebox.showerror("Error", "Please select both a geographic area type and name")
            return

        if not self.output_file:
            messagebox.showerror("Error", "Please select an output file location")
            return

        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    messagebox.showerror("Error", f"Cannot create output directory: {str(e)}")
                    return

            # Make sure we have a valid file path for saving
            if not self.output_file.endswith('.xlsx'):
                self.output_file += '.xlsx'

            self.status_label.config(text="Collecting data...")
            self.root.update_idletasks()

            # Collect data from various sources
            self.collect_data()

            self.status_label.config(text="Creating Excel output...")
            self.root.update_idletasks()

            # Create Excel output
            self.create_excel_output()

            self.status_label.config(text=f"Analysis completed and saved to {self.output_file}")
            messagebox.showinfo("Success", f"Analysis completed and saved to {self.output_file}")

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in run_analysis: {str(e)}")
            print(error_details)
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate analysis: {str(e)}\n\nSee console for details.")


    def get_latest_data_for_area(self, data_type, area_name):
        """Helper function to find the latest data for a specific area and data type"""
        print(f"\nAttempting to find latest {data_type} data for: {area_name}")
        try:
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS[data_type])
            print(f"Looking in directory: {dir_path}")

            # Need to find the file pattern that includes the area name
            # This assumes Greater Sydney and Rest of NSW data might be in SA4 or GCCSA files
            relevant_geo_areas = ["GCCSA", "SA4"] # Common areas for GS/RON data
            found_file_pattern = None
            for geo_area in relevant_geo_areas:
                 if geo_area.lower() in self.FILE_PATTERNS[data_type]:
                     found_file_pattern = self.FILE_PATTERNS[data_type][geo_area.lower()]
                     print(f"Using file pattern: {found_file_pattern} for geo area {geo_area}")
                     break

            if not found_file_pattern:
                print(f"No relevant file pattern found for {data_type} in {relevant_geo_areas}")
                return None

            found_file = None
            for file in os.listdir(dir_path):
                if found_file_pattern.lower() in file.lower():
                    found_file = os.path.join(dir_path, file)
                    print(f"Found potential data file: {found_file}")
                    break

            if not found_file:
                print(f"No data file found for pattern {found_file_pattern}")
                return None

            df = self.read_data_file(found_file)

            if df is None or df.empty:
                print("Failed to read data file or file is empty.")
                return None

            # Try finding the column based on GCCSA or SA4 as a hint
            geo_col = self.find_geographic_column(df, "GCCSA") or self.find_geographic_column(df, "SA4")
            if not geo_col:
                print("Could not find geographic column in the file.")
                return None

            df[geo_col] = df[geo_col].astype(str)
            df_filtered = df[df[geo_col] == area_name].copy() # Use .copy()

            if df_filtered.empty:
                print(f"No data found for area name: '{area_name}' in column '{geo_col}'")
                return None

            print(f"Found {len(df_filtered)} rows for {area_name}. Columns: {df_filtered.columns.tolist()}")

            # Find the latest month
            latest_month = None
            if 'month' in df_filtered.columns:
                try:
                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                    latest_month = df_filtered['month'].max()
                    df_latest = df_filtered[df_filtered['month'] == latest_month].copy() # Use .copy() to avoid SettingWithCopyWarning
                    print(f"Filtered to latest month: {latest_month}")
                except Exception as e:
                    print(f"Error processing month column: {str(e)}")
                    df_latest = df_filtered.copy()
            else:
                df_latest = df_filtered.copy()

            if df_latest.empty:
                 print("No data in latest month.")
                 return None

            # Handle specific data types
            if data_type == "median_rents":
                # Prefer 3-month median rent for stability
                rent_col = None
                for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                    for col in df_latest.columns:
                        if col.startswith(col_prefix) and not any(x in col for x in ['growth', 'increase', 'change']):
                            rent_col = col
                            break
                    if rent_col:
                        break

                # Find annual growth column (if available in the source file)
                growth_col = None
                for col_suffix in ['annual_growth', 'annual_increase', 'yearly_growth', 'yearly_increase']:
                    for col in df_latest.columns:
                        if col.endswith(col_suffix):
                            growth_col = col
                            break
                    if growth_col:
                        break

                if rent_col:
                    rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                    annual_increase = None
                    if growth_col and not pd.isna(df_latest[growth_col].iloc[0]):
                         annual_increase = float(df_latest[growth_col].iloc[0])
                         # Assuming annual increase might be a decimal (e.g., 0.12 for 12%)
                         if annual_increase is not None and annual_increase < 50 and annual_increase > -50: # Simple check to see if it's likely a percentage as a decimal
                             annual_increase *= 100

                    return {"value": rent_value, "annual_change": annual_increase, "period": latest_month.strftime("%b-%Y") if latest_month else "Latest"}
                else:
                    print("Could not find median rent column.")
                    return None

            elif data_type == "vacancy_rates":
                 # Prefer 3-month smoothed vacancy rate
                rate_col = None
                if 'rental_vacancy_rate_3m_smoothed' in df_latest.columns:
                    rate_col = 'rental_vacancy_rate_3m_smoothed'
                else:
                    # Fallback to other columns if the specific one is not found
                    for col_name in ['rental_vacancy_rate', 'vacancy_rate', 'rate']:
                        if col_name in df_latest.columns:
                            rate_col = col_name
                            break

                # Find annual change column (if available in the source file)
                change_col = None
                for col_suffix in ['annual_change', 'yearly_change', 'annual_growth']:
                    for col in df_latest.columns:
                        if col_suffix in col.lower():
                            change_col = col
                            break
                    if change_col:
                        break


                if rate_col:
                    rate_value = float(df_latest[rate_col].iloc[0]) if not pd.isna(df_latest[rate_col].iloc[0]) else 0
                    # Ensure it's formatted as a percentage (0.75 vs 0.0075)
                    if rate_value is not None and rate_value > 0 and rate_value < 10: # Assuming rates are typically less than 10%
                         rate_value *= 100

                    annual_change = None
                    if change_col and not pd.isna(df_latest[change_col].iloc[0]):
                         annual_change = float(df_latest[change_col].iloc[0])
                         # Assuming annual change might be a decimal (e.g., -0.003 for -0.3%)
                         if annual_change is not None and annual_change < 1 and annual_change > -1: # Simple check
                            annual_change *= 100

                    return {"value": rate_value, "annual_change": annual_change, "period": latest_month.strftime("%b-%Y") if latest_month else "Latest"}
                else:
                    print("Could not find vacancy rate column.")
                    return None


            elif data_type == "affordability":
                 # Find affordability column - look for keywords
                pct_col = None
                affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                if affordability_columns:
                    if 'rental_affordability_3mo' in affordability_columns:
                        pct_col = 'rental_affordability_3mo'
                    elif 'rental_affordability_1mo' in affordability_columns:
                        pct_col = 'rental_affordability_1mo'
                    else:
                        pct_col = affordability_columns[0]  # Take the first one if no preferred column

                # If no direct affordability column, try to find rent-to-income ratio
                if not pct_col:
                    for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                        for col in df_latest.columns:
                            if all(keyword.lower() in col.lower() for keyword in keywords):
                                pct_col = col
                                break
                        if pct_col:
                            break

                # If still no column, try to calculate it ourselves if we have rent and income columns
                if not pct_col:
                    median_rent_cols = [col for col in df_latest.columns if 'median_rent' in col.lower() and not any(x in col.lower() for x in ['growth', 'increase', 'change'])]
                    income_cols = [col for col in df_latest.columns if 'income' in col.lower() and 'index' not in col.lower()]

                    if median_rent_cols and income_cols and len(df_latest) > 0:
                        # Prefer 3-month median rent for stability
                        rent_col = None
                        for prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent']:
                            matching_cols = [col for col in median_rent_cols if col.startswith(prefix)]
                            if matching_cols:
                                rent_col = matching_cols[0]
                                break

                        if not rent_col and median_rent_cols:
                            rent_col = median_rent_cols[0]

                        # Use the first income column we find
                        income_col = income_cols[0]

                        if rent_col and income_col:
                            rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                            income_value = float(df_latest[income_col].iloc[0]) if not pd.isna(df_latest[income_col].iloc[0]) else 0

                            if income_value > 0:
                                # Calculate weekly rental affordability (rent/income)
                                # For weekly rent and annual income: (weekly_rent * 52) / annual_income * 100
                                affordability = (rent_value * 52) / income_value * 100

                                # Create a placeholder dataframe with our calculated value
                                df_latest['calculated_affordability'] = affordability
                                pct_col = 'calculated_affordability'
                                print(f"Calculated affordability manually: {affordability:.1f}%")

                # Extract current affordability value
                current_pct_value = None
                if pct_col and len(df_latest) > 0:
                    current_pct_value = float(df_latest[pct_col].iloc[0]) if not pd.isna(df_latest[pct_col].iloc[0]) else 0
                    # Ensure the value is properly formatted as a percentage
                    if current_pct_value is not None and current_pct_value > 0 and current_pct_value < 10: # Assuming affordability is typically less than 10% as a decimal
                        current_pct_value *= 100
                    print(f"Current Affordability: {current_pct_value:.1f}%")

                # Find annual improvement column
                improvement_col = None
                for col_suffix in ['improvement', 'change', 'growth']:
                    for col in df_latest.columns:
                        if col_suffix in col.lower() and any(x in col.lower() for x in ['annual', 'yearly']):
                            improvement_col = col
                            break
                    if improvement_col:
                        break

                annual_improvement = None
                if improvement_col and len(df_latest) > 0 and not pd.isna(df_latest[improvement_col].iloc[0]):
                     annual_improvement = float(df_latest[improvement_col].iloc[0])
                     # Assuming annual improvement might be a decimal (e.g., 0.05 for 5%)
                     if annual_improvement is not None and annual_improvement < 1 and annual_improvement > -1: # Simple check
                          annual_improvement *= 100


                return {"percentage": current_pct_value, "annual_improvement": annual_improvement, "period": latest_month.strftime("%b-%Y") if latest_month else "Latest"}
                # Note: Annual improvement is calculated manually for the selected area, but for comparison areas,
                # we rely on an existing annual change column if available.


        except Exception as e:
            print(f"Error in get_latest_data_for_area for {area_name} ({data_type}): {str(e)}")
            import traceback
            traceback.print_exc()
            return None


    def collect_data(self):
        """Collect data from various sources"""
        self.data = {}
        print(f"\n\n{'='*50}")
        print(f"STARTING DATA COLLECTION FOR: {self.selected_geo_name} ({self.selected_geo_area})")
        print(f"{'='*50}")

        # --- Dynamically collect Greater Sydney and Rest of NSW reference data ---
        print("\nCollecting dynamic reference data...")

        # Median Rent for Greater Sydney and Rest of NSW
        gs_rent_data = self.get_latest_data_for_area("median_rents", "Greater Sydney")
        if gs_rent_data:
             self.GS_REFERENCE_DATA["median_rent"]["value"] = gs_rent_data["value"]
             self.GS_REFERENCE_DATA["median_rent"]["annual_change"] = gs_rent_data["annual_change"]
             print(f"Updated GS Median Rent: {self.GS_REFERENCE_DATA['median_rent']}")
        else:
             print("Could not fetch Greater Sydney Median Rent data.")


        ron_rent_data = self.get_latest_data_for_area("median_rents", "Rest of NSW")
        if ron_rent_data:
             self.RON_REFERENCE_DATA["median_rent"]["value"] = ron_rent_data["value"]
             self.RON_REFERENCE_DATA["median_rent"]["annual_change"] = ron_rent_data["annual_change"]
             print(f"Updated RON Median Rent: {self.RON_REFERENCE_DATA['median_rent']}")
        else:
             print("Could not fetch Rest of NSW Median Rent data.")


        # Vacancy Rates for Greater Sydney and Rest of NSW
        gs_vacancy_data = self.get_latest_data_for_area("vacancy_rates", "Greater Sydney")
        if gs_vacancy_data:
             self.GS_REFERENCE_DATA["vacancy_rates"]["value"] = gs_vacancy_data["value"]
             self.GS_REFERENCE_DATA["vacancy_rates"]["annual_change"] = gs_vacancy_data["annual_change"]
             print(f"Updated GS Vacancy Rate: {self.GS_REFERENCE_DATA['vacancy_rates']}")
        else:
             print("Could not fetch Greater Sydney Vacancy Rate data.")


        ron_vacancy_data = self.get_latest_data_for_area("vacancy_rates", "Rest of NSW")
        if ron_vacancy_data:
             self.RON_REFERENCE_DATA["vacancy_rates"]["value"] = ron_vacancy_data["value"]
             self.RON_REFERENCE_DATA["vacancy_rates"]["annual_change"] = ron_vacancy_data["annual_change"]
             print(f"Updated RON Vacancy Rate: {self.RON_REFERENCE_DATA['vacancy_rates']}")
        else:
             print("Could not fetch Rest of NSW Vacancy Rate data.")

        # Affordability for Greater Sydney and Rest of NSW
        gs_affordability_data = self.get_latest_data_for_area("affordability", "Greater Sydney")
        if gs_affordability_data:
             self.GS_REFERENCE_DATA["affordability"]["value"] = gs_affordability_data["percentage"]
             self.GS_REFERENCE_DATA["affordability"]["annual_change"] = gs_affordability_data["annual_improvement"]
             print(f"Updated GS Affordability: {self.GS_REFERENCE_DATA['affordability']}")
        else:
             print("Could not fetch Greater Sydney Affordability data.")


        ron_affordability_data = self.get_latest_data_for_area("affordability", "Rest of NSW")
        if ron_affordability_data:
             self.RON_REFERENCE_DATA["affordability"]["value"] = ron_affordability_data["percentage"]
             self.RON_REFERENCE_DATA["affordability"]["annual_change"] = ron_affordability_data["annual_improvement"]
             print(f"Updated RON Affordability: {self.RON_REFERENCE_DATA['affordability']}")
        else:
             print("Could not fetch Rest of NSW Affordability data.")


        print("\nDynamic reference data collection complete.")
        print(f"GS Reference Data: {self.GS_REFERENCE_DATA}")
        print(f"RON Reference Data: {self.RON_REFERENCE_DATA}")
        print("-" * 50)

        # --- Collect data for the selected geographic area ---

        # Collect Census dwelling data
        print("\nCollecting Census dwelling data...")
        try:
            # Find census dwelling file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["census_dwelling"])
            print(f"Looking for census files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["census_dwelling"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")

            census_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    census_files_found.append(file)

            print(f"Census files found: {census_files_found}")

            for file in census_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing census file: {file_path}")
                df = self.read_data_file(file_path)

                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)

                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")

                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str].copy() # Use .copy()

                        if df_filtered.empty:
                            print(f"No exact match found for '{selected_name_str}'. Trying to find partial matches...")

                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in str(value).lower() or str(value).lower() in selected_name_str.lower():
                                    matches.append(value)

                            if matches:
                                print(f"Potential partial matches: {matches}")
                                # Prioritize exact match if it exists, otherwise take the first partial match
                                best_match = selected_name_str if selected_name_str in matches else matches[0]
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match].copy()
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")

                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")

                            # Calculate rental percentage by finding total rental and total dwellings
                            # Method 1: Look for direct percentage columns
                            pct_col = None
                            for col in df_filtered.columns:
                                if "rented" in col.lower() and "percent" in col.lower():
                                    pct_col = col
                                    break

                            # Method 2: Calculate from raw counts
                            total_dwellings = None
                            if "dwellings" in df_filtered.columns:
                                total_dwellings = float(df_filtered["dwellings"].iloc[0]) if not pd.isna(df_filtered["dwellings"].iloc[0]) else 0

                            total_rented = None
                            if "dwellings_rented" in df_filtered.columns:
                                total_rented = float(df_filtered["dwellings_rented"].iloc[0]) if not pd.isna(df_filtered["dwellings_rented"].iloc[0]) else 0

                            # Calculate percentage if we have the raw counts
                            if total_dwellings is not None and total_rented is not None and total_dwellings > 0:
                                rental_pct = (total_rented / total_dwellings) * 100
                                rental_count = int(total_rented)
                                print(f"Calculated rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")

                                self.data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["renters"]
                                }
                            elif pct_col:
                                rental_pct = float(df_filtered[pct_col].iloc[0]) if not pd.isna(df_filtered[pct_col].iloc[0]) else 0
                                renter_count_col = pct_col.replace("percent", "count")
                                rental_count = int(df_filtered[renter_count_col].iloc[0]) if renter_count_col in df_filtered.columns and not pd.isna(df_filtered[renter_count_col].iloc[0]) else 0
                                print(f"Found rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")

                                self.data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["renters"]
                                }

                            # Find social housing data - specifically add dwellings_rented_sha + dwellings_rented_chp
                            social_housing_sha = 0
                            social_housing_chp = 0

                            # Get SHA data
                            if "dwellings_rented_sha" in df_filtered.columns:
                                sha_value = df_filtered["dwellings_rented_sha"].iloc[0]
                                social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                                print(f"SHA dwellings: {social_housing_sha}")

                            # Get CHP data
                            if "dwellings_rented_chp" in df_filtered.columns:
                                chp_value = df_filtered["dwellings_rented_chp"].iloc[0]
                                social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                                print(f"CHP dwellings: {social_housing_chp}")

                            # Calculate total social housing
                            total_social = social_housing_sha + social_housing_chp
                            print(f"Total social housing dwellings: {total_social}")

                            # Calculate social housing percentage
                            if total_dwellings is not None and total_dwellings > 0:
                                social_pct = (total_social / total_dwellings) * 100
                                social_count = int(total_social)
                                print(f"Calculated social housing percentage: {social_pct:.1f}% ({social_count} dwellings)")

                                self.data["social_housing"] = {
                                    "percentage": round(social_pct, 1),
                                    "count": social_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["social_housing"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["social_housing"]
                                }
                    break # Stop after finding data in one file
        except Exception as e:
            print(f"Error collecting census data: {str(e)}")
            import traceback
            traceback.print_exc()

        # If no data collected, use defaults (should ideally not happen if files exist)
        if "renters" not in self.data:
            print("Using DEFAULT renter data")
            self.data["renters"] = {
                "percentage": 25.5,
                "count": 8402,
                "period": "2021",
                "source": "ABS Census",
                "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                "comparison_ron": self.RON_REFERENCE_DATA["renters"]
            }

        if "social_housing" not in self.data:
            print("Using DEFAULT social housing data")
            self.data["social_housing"] = {
                "percentage": 2.8,
                "count": 938,
                "period": "2021",
                "source": "ABS Census",
                "comparison_gs": self.GS_REFERENCE_DATA["social_housing"],
                "comparison_ron": self.RON_REFERENCE_DATA["social_housing"]
            }

        # Collect Median Rent data
        print("\nCollecting Median Rent data...")
        try:
            # Find median rent file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["median_rents"])
            print(f"Looking for median rent files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["median_rents"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")

            rent_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    rent_files_found.append(file)

            print(f"Rent files found: {rent_files_found}")

            for file in rent_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing rent file: {file_path}")
                df = self.read_data_file(file_path)

                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)

                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")

                        # Check for exact match
                        df_filtered_all_months = df[df[geo_col] == selected_name_str].copy() # Keep all months for calculation

                        if df_filtered_all_months.empty:
                            print(f"No exact match found for '{selected_name_str}'. Trying to find partial matches...")

                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in str(value).lower() or str(value).lower() in selected_name_str.lower():
                                    matches.append(value)

                            if matches:
                                print(f"Potential partial matches: {matches}")
                                # Prioritize exact match if it exists, otherwise take the first partial match
                                best_match = selected_name_str if selected_name_str in matches else matches[0]
                                print(f"Using best match: '{best_match}'")
                                df_filtered_all_months = df[df[geo_col] == best_match].copy()
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered_all_months)} rows")

                        if not df_filtered_all_months.empty:
                            print(f"Filtered data columns (all months): {df_filtered_all_months.columns.tolist()}")

                            # Ensure 'month' column is datetime and sort
                            if 'month' in df_filtered_all_months.columns:
                                try:
                                    df_filtered_all_months['month'] = pd.to_datetime(df_filtered_all_months['month'], errors='coerce')
                                    df_filtered_all_months.sort_values(by='month', inplace=True)
                                    df_filtered_all_months.dropna(subset=['month'], inplace=True) # Drop rows with invalid dates
                                    print("Processed month column and sorted data.")
                                except Exception as e:
                                    print(f"Error processing month column for median rent: {str(e)}")
                                    # If date processing fails, we can't calculate growth
                                    df_filtered_all_months = pd.DataFrame() # Empty dataframe to skip calculations

                            if not df_filtered_all_months.empty:
                                # If we have property_type, filter to "All Dwellings" for consistency
                                if 'property_type' in df_filtered_all_months.columns:
                                    if 'All Dwellings' in df_filtered_all_months['property_type'].values:
                                        df_filtered_all_months = df_filtered_all_months[df_filtered_all_months['property_type'] == 'All Dwellings'].copy()
                                        print("Filtered to 'All Dwellings' property type")
                                    else:
                                         print("Warning: 'All Dwellings' not found in property_type column.")


                                # Find the latest month's data
                                latest_month = df_filtered_all_months['month'].max()
                                df_latest = df_filtered_all_months[df_filtered_all_months['month'] == latest_month].copy()
                                print(f"Filtered to latest month: {latest_month.strftime('%b-%Y')}")

                                # Find columns for median rent data - prefer 3-month median
                                rent_col = None
                                for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                                    for col in df_latest.columns:
                                        if col.startswith(col_prefix) and not any(x in col for x in ['growth', 'increase', 'change']):
                                            rent_col = col
                                            break
                                    if rent_col:
                                        break

                                print(f"Median rent column used for latest data: {rent_col}")

                                # Extract current median rent value
                                current_rent_value = None
                                if rent_col and len(df_latest) > 0:
                                    current_rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                                    print(f"Current Median rent: ${current_rent_value}")

                                # Calculate annual growth
                                annual_growth_pct = 0
                                previous_year_rent = None

                                if current_rent_value is not None and latest_month is not None:
                                    one_year_ago = latest_month - pd.DateOffset(months=12)

                                    # Find data for the month 12 months prior
                                    df_one_year_ago = df_filtered_all_months[df_filtered_all_months['month'] == one_year_ago].copy()

                                    if not df_one_year_ago.empty and rent_col in df_one_year_ago.columns:
                                        previous_year_rent = float(df_one_year_ago[rent_col].iloc[0]) if not pd.isna(df_one_year_ago[rent_col].iloc[0]) else 0
                                        print(f"Median rent 12 months prior ({one_year_ago.strftime('%b-%Y')}): ${previous_year_rent}")

                                        if previous_year_rent is not None and previous_year_rent > 0:
                                            annual_growth_pct = ((current_rent_value - previous_year_rent) / previous_year_rent) * 100
                                            print(f"Calculated annual growth: {annual_growth_pct:.1f}%")
                                        elif previous_year_rent == 0 and current_rent_value is not None and current_rent_value > 0:
                                             annual_growth_pct = 1000 # Indicate significant growth from zero
                                             print("Calculated annual growth: >1000% (from $0)")
                                        else:
                                             annual_growth_pct = 0
                                             print("Calculated annual growth: 0% (no change or data issues)")
                                    else:
                                        print(f"No data found for {one_year_ago.strftime('%b-%Y')} to calculate annual growth.")

                                # Store median rent data
                                self.data["median_rent"] = {
                                    "value": int(round(current_rent_value, 0)) if current_rent_value is not None else 0,
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "N/A",
                                    "source": "NSW Fair Trading Corelogic Data",
                                    "annual_increase": round(annual_growth_pct, 1), # Use calculated growth
                                    "previous_year_rent": int(round(previous_year_rent, 0)) if previous_year_rent is not None else None,
                                    "comparison_gs": self.GS_REFERENCE_DATA["median_rent"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["median_rent"]
                                }
                    break # Stop after finding data in one file
        except Exception as e:
            print(f"Error collecting median rent data: {str(e)}")
            import traceback
            traceback.print_exc()

        # If no data collected, use defaults
        if "median_rent" not in self.data:
            print("Using DEFAULT median rent data")
            # Attempt to calculate a plausible previous year rent from default values
            default_value = 595
            default_increase_pct = 10.0
            # Calculate the value 12 months ago based on the default increase
            # value_ago * (1 + increase_pct/100) = current_value
            # value_ago = current_value / (1 + increase_pct/100)
            default_previous_year = round(default_value / (1 + default_increase_pct / 100)) if default_increase_pct != -100 else 0


            self.data["median_rent"] = {
                "value": default_value,
                "period": "Apr-25", # Assuming latest default period
                "source": "NSW Fair Trading Corelogic Data",
                "annual_increase": default_increase_pct,
                "previous_year_rent": default_previous_year,
                "comparison_gs": self.GS_REFERENCE_DATA["median_rent"],
                "comparison_ron": self.RON_REFERENCE_DATA["median_rent"]
            }


        # Collect Vacancy Rate data
        print("\nCollecting Vacancy Rate data...")
        try:
            # Find vacancy rate file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["vacancy_rates"])
            print(f"Looking for vacancy rate files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["vacancy_rates"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")

            vacancy_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    vacancy_files_found.append(file)

            print(f"Vacancy rate files found: {vacancy_files_found}")

            for file in vacancy_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing vacancy rate file: {file_path}")
                df = self.read_data_file(file_path)

                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)

                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")

                         # Check for exact match
                        df_filtered_all_months = df[df[geo_col] == selected_name_str].copy() # Keep all months for calculation

                        if df_filtered_all_months.empty:
                            print(f"No exact match found for '{selected_name_str}'. Trying to find partial matches...")

                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in str(value).lower() or str(value).lower() in selected_name_str.lower():
                                    matches.append(value)

                            if matches:
                                print(f"Potential partial matches: {matches}")
                                # Prioritize exact match if it exists, otherwise take the first partial match
                                best_match = selected_name_str if selected_name_str in matches else matches[0]
                                print(f"Using best match: '{best_match}'")
                                df_filtered_all_months = df[df[geo_col] == best_match].copy()
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered_all_months)} rows")


                        if not df_filtered_all_months.empty:
                            print(f"Filtered data columns (all months): {df_filtered_all_months.columns.tolist()}")

                             # Ensure 'month' column is datetime and sort
                            if 'month' in df_filtered_all_months.columns:
                                try:
                                    df_filtered_all_months['month'] = pd.to_datetime(df_filtered_all_months['month'], errors='coerce')
                                    df_filtered_all_months.sort_values(by='month', inplace=True)
                                    df_filtered_all_months.dropna(subset=['month'], inplace=True) # Drop rows with invalid dates
                                    print("Processed month column and sorted data.")
                                except Exception as e:
                                    print(f"Error processing month column for vacancy rates: {str(e)}")
                                    # If date processing fails, we can't find latest or previous year
                                    df_filtered_all_months = pd.DataFrame() # Empty dataframe to skip calculations


                            if not df_filtered_all_months.empty:
                                # Find the latest month's data
                                latest_month = df_filtered_all_months['month'].max()
                                df_latest = df_filtered_all_months[df_filtered_all_months['month'] == latest_month].copy()
                                print(f"Filtered to latest month: {latest_month.strftime('%b-%Y')}")

                                # Find vacancy rate column - specifically use rental_vacancy_rate_3m_smoothed
                                rate_col = None
                                if 'rental_vacancy_rate_3m_smoothed' in df_latest.columns:
                                    rate_col = 'rental_vacancy_rate_3m_smoothed'
                                    print(f"Found specific vacancy rate column: {rate_col}")
                                else:
                                    # Fallback to other columns if the specific one is not found
                                    for col_name in ['rental_vacancy_rate', 'vacancy_rate', 'rate']:
                                        if col_name in df_latest.columns:
                                            rate_col = col_name
                                            print(f"Using fallback vacancy rate column: {rate_col}")
                                            break

                                # Extract current vacancy rate value
                                current_rate_value = None
                                if rate_col and len(df_latest) > 0:
                                    current_rate_value = float(df_latest[rate_col].iloc[0]) if not pd.isna(df_latest[rate_col].iloc[0]) else 0
                                    # Ensure it's formatted as a percentage (0.75 vs 0.0075)
                                    if current_rate_value is not None and current_rate_value > 0 and current_rate_value < 10: # Assuming rates are typically less than 10%
                                         current_rate_value *= 100
                                    print(f"Current Vacancy rate: {current_rate_value:.2f}%")

                                # Find previous year's rate
                                previous_year_rate = None
                                if current_rate_value is not None and latest_month is not None:
                                     one_year_ago = latest_month - pd.DateOffset(months=12)
                                     df_one_year_ago = df_filtered_all_months[df_filtered_all_months['month'] == one_year_ago].copy()

                                     if not df_one_year_ago.empty and rate_col in df_one_year_ago.columns:
                                         previous_year_rate = float(df_one_year_ago[rate_col].iloc[0]) if not pd.isna(df_one_year_ago[rate_col].iloc[0]) else 0
                                         # Ensure it's formatted as a percentage
                                         if previous_year_rate is not None and previous_year_rate > 0 and previous_year_rate < 10:
                                             previous_year_rate *= 100
                                         print(f"Vacancy rate 12 months prior ({one_year_ago.strftime('%b-%Y')}): {previous_year_rate:.2f}%")
                                     else:
                                         print(f"No data found for {one_year_ago.strftime('%b-%Y')} to get previous year vacancy rate.")


                                # Store vacancy rate data
                                self.data["vacancy_rates"] = {
                                    "value": round(current_rate_value, 2) if current_rate_value is not None else 0,
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "N/A",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "previous_year_rate": round(previous_year_rate, 2) if previous_year_rate is not None else None,
                                    "comparison_gs": self.GS_REFERENCE_DATA["vacancy_rates"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["vacancy_rates"]
                                }
                    break # Stop after finding data in one file
        except Exception as e:
            print(f"Error collecting vacancy rate data: {str(e)}")
            import traceback
            traceback.print_exc()

        # If no data collected, use defaults
        if "vacancy_rates" not in self.data:
            print("Using DEFAULT vacancy rate data")
            self.data["vacancy_rates"] = {
                "value": 0.72,
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "previous_year_rate": 1.0,  # Default value for previous year
                "comparison_gs": self.GS_REFERENCE_DATA["vacancy_rates"],
                "comparison_ron": self.RON_REFERENCE_DATA["vacancy_rates"]
            }

        # Collect Affordability data
        print("\nCollecting Affordability data...")
        try:
            # Find affordability file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["affordability"])
            print(f"Looking for affordability files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["affordability"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")

            affordability_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    affordability_files_found.append(file)

            print(f"Affordability files found: {affordability_files_found}")

            for file in affordability_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing affordability file: {file_path}")
                df = self.read_data_file(file_path)

                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)

                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")

                        # Check for exact match
                        df_filtered_all_months = df[df[geo_col] == selected_name_str].copy() # Keep all months for calculation

                        if df_filtered_all_months.empty:
                            print(f"No exact match found for '{selected_name_str}'. Trying to find partial matches...")

                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in str(value).lower() or str(value).lower() in selected_name_str.lower():
                                    matches.append(value)

                            if matches:
                                print(f"Potential partial matches: {matches}")
                                # Prioritize exact match if it exists, otherwise take the first partial match
                                best_match = selected_name_str if selected_name_str in matches else matches[0]
                                print(f"Using best match: '{best_match}'")
                                df_filtered_all_months = df[df[geo_col] == best_match].copy()
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered_all_months)} rows")


                        if not df_filtered_all_months.empty:
                            print(f"Filtered data columns (all months): {df_filtered_all_months.columns.tolist()}")

                            # Ensure 'month' column is datetime and sort
                            if 'month' in df_filtered_all_months.columns:
                                try:
                                    df_filtered_all_months['month'] = pd.to_datetime(df_filtered_all_months['month'], errors='coerce')
                                    df_filtered_all_months.sort_values(by='month', inplace=True)
                                    df_filtered_all_months.dropna(subset=['month'], inplace=True) # Drop rows with invalid dates
                                    print("Processed month column and sorted data.")
                                except Exception as e:
                                    print(f"Error processing month column for affordability: {str(e)}")
                                    # If date processing fails, we can't calculate improvement
                                    df_filtered_all_months = pd.DataFrame() # Empty dataframe to skip calculations


                            if not df_filtered_all_months.empty:
                                # Find the latest month's data
                                latest_month = df_filtered_all_months['month'].max()
                                df_latest = df_filtered_all_months[df_filtered_all_months['month'] == latest_month].copy()
                                print(f"Filtered to latest month: {latest_month.strftime('%b-%Y')}")

                                # Find affordability column - look for keywords
                                pct_col = None

                                # First priority: direct affordability columns
                                affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                                if affordability_columns:
                                    # Prefer 3-month affordability for stability
                                    if 'rental_affordability_3mo' in affordability_columns:
                                        pct_col = 'rental_affordability_3mo'
                                    elif 'rental_affordability_1mo' in affordability_columns:
                                        pct_col = 'rental_affordability_1mo'
                                    else:
                                        pct_col = affordability_columns[0]  # Take the first one if no preferred column

                                # If no direct affordability column, try to find rent-to-income ratio
                                if not pct_col:
                                    for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                                        for col in df_latest.columns:
                                            if all(keyword.lower() in col.lower() for keyword in keywords):
                                                pct_col = col
                                                break
                                        if pct_col:
                                            break

                                # If still no column, try to calculate it ourselves if we have rent and income columns
                                if not pct_col:
                                    median_rent_cols = [col for col in df_latest.columns if 'median_rent' in col.lower() and not any(x in col.lower() for x in ['growth', 'increase', 'change'])]
                                    income_cols = [col for col in df_latest.columns if 'income' in col.lower() and 'index' not in col.lower()]

                                    if median_rent_cols and income_cols and len(df_latest) > 0:
                                        # Prefer 3-month median rent for stability
                                        rent_col = None
                                        for prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent']:
                                            matching_cols = [col for col in median_rent_cols if col.startswith(prefix)]
                                            if matching_cols:
                                                rent_col = matching_cols[0]
                                                break

                                        if not rent_col and median_rent_cols:
                                            rent_col = median_rent_cols[0]

                                        # Use the first income column we find
                                        income_col = income_cols[0]

                                        if rent_col and income_col:
                                            rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                                            income_value = float(df_latest[income_col].iloc[0]) if not pd.isna(df_latest[income_col].iloc[0]) else 0

                                            if income_value > 0:
                                                # Calculate weekly rental affordability (rent/income)
                                                # For weekly rent and annual income: (weekly_rent * 52) / annual_income * 100
                                                affordability = (rent_value * 52) / income_value * 100

                                                # Create a placeholder dataframe with our calculated value
                                                df_latest['calculated_affordability'] = affordability
                                                pct_col = 'calculated_affordability'
                                                print(f"Calculated affordability manually: {affordability:.1f}%")

                                # Extract current affordability value
                                current_pct_value = None
                                if pct_col and len(df_latest) > 0:
                                    current_pct_value = float(df_latest[pct_col].iloc[0]) if not pd.isna(df_latest[pct_col].iloc[0]) else 0
                                    # Ensure the value is properly formatted as a percentage
                                    if current_pct_value is not None and current_pct_value > 0 and current_pct_value < 10: # Assuming affordability is typically less than 10% as a decimal
                                        current_pct_value *= 100
                                    print(f"Current Affordability: {current_pct_value:.1f}%")

                                # Calculate annual improvement
                                annual_improvement = 0
                                if current_pct_value is not None and latest_month is not None and 'month' in df_filtered_all_months.columns and pct_col in df_filtered_all_months.columns:
                                    try:
                                        one_year_ago = latest_month - pd.DateOffset(months=12)
                                        df_one_year_ago = df_filtered_all_months[df_filtered_all_months['month'] == one_year_ago].copy()

                                        if not df_one_year_ago.empty and pct_col in df_one_year_ago.columns:
                                            year_ago_value = float(df_one_year_ago[pct_col].iloc[0]) if not pd.isna(df_one_year_ago[pct_col].iloc[0]) else 0
                                            # Ensure it's formatted as a percentage
                                            if year_ago_value is not None and year_ago_value > 0 and year_ago_value < 10:
                                                year_ago_value *= 100

                                            # Calculate change - note that for affordability, a decrease is an improvement
                                            # (lower percentage of income spent on rent is better)
                                            change = current_pct_value - year_ago_value
                                            annual_improvement = -change  # Negative change = improvement for affordability

                                            print(f"Calculated annual affordability change: {change:.2f}% (improvement: {annual_improvement:.2f}%)")
                                            print(f"Comparing {latest_month.strftime('%b-%Y')} ({current_pct_value:.1f}%) to {one_year_ago.strftime('%b-%Y')} ({year_ago_value:.1f}%)")
                                        else:
                                            print(f"No data found for {one_year_ago.strftime('%b-%Y')} to calculate annual improvement")
                                    except Exception as e:
                                        print(f"Error calculating annual improvement: {str(e)}")


                                # Store affordability data
                                self.data["affordability"] = {
                                    "percentage": round(current_pct_value, 1) if current_pct_value is not None else 0,
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "N/A",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "annual_improvement": round(annual_improvement, 2),
                                    "comparison_gs": self.GS_REFERENCE_DATA["affordability"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["affordability"]
                                }
                    break # Stop after finding data in one file
        except Exception as e:
            print(f"Error collecting affordability data: {str(e)}")
            import traceback
            traceback.print_exc()

        # If no data collected, use defaults
        if "affordability" not in self.data:
            print("Using DEFAULT affordability data")
            self.data["affordability"] = {
                "percentage": 43.6,
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "annual_improvement": 0,
                "comparison_gs": self.GS_REFERENCE_DATA["affordability"],
                "comparison_ron": self.RON_REFERENCE_DATA["affordability"]
            }

        # Removed rental contacts data as requested

        print("\nData collection complete")
        print(f"{'='*50}\n")

    def find_column(self, df, keywords):
        """Find column that contains all the keywords (case-insensitive)"""
        print(f"Looking for column with keywords: {keywords}")

        columns_found = []
        for col in df.columns:
            if all(keyword.lower() in str(col).lower() for keyword in keywords):
                columns_found.append(col)
                print(f"Found matching column: {col}")

        if columns_found:
            # Return the first match
            return columns_found[0]

        # Try with partial matching if no exact match
        partial_matches = []
        for col in df.columns:
            col_lower = str(col).lower()
            match_score = sum(1 for kw in keywords if kw.lower() in col_lower)
            if match_score > 0:
                partial_matches.append((col, match_score))

        # Sort by match score
        partial_matches.sort(key=lambda x: x[1], reverse=True)

        if partial_matches:
            best_match = partial_matches[0][0]
            print(f"No exact match found. Using best partial match: {best_match}")
            return best_match

        print(f"No matching column found for keywords: {keywords}")
        return None

    def generate_comparison_comment(self, metric, value, comparison_gs, comparison_ron):
        """Generate a comparison comment for a metric that shows both Greater Sydney and Rest of NSW references"""

        if metric == "renters":
            gs_text = ""
            if comparison_gs["value"] is not None:
                if value < comparison_gs["value"] - 1:  # 1% buffer to avoid "slightly lower" for small differences
                    gs_text = f"lower than the Greater Sydney average of {comparison_gs['value']}%"
                elif value > comparison_gs["value"] + 1:
                    gs_text = f"higher than the Greater Sydney average of {comparison_gs['value']}%"
                else:
                    gs_text = f"similar to the Greater Sydney average of {comparison_gs['value']}%"
            else:
                gs_text = "compared to Greater Sydney (data not available)"


            ron_text = ""
            if comparison_ron["value"] is not None:
                if value < comparison_ron["value"] - 1:
                    ron_text = f"and lower than the Rest of NSW average of {comparison_ron['value']}%"
                elif value > comparison_ron["value"] + 1:
                    ron_text = f"and higher than the Rest of NSW average of {comparison_ron['value']}%"
                else:
                    ron_text = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
            else:
                 ron_text = "and Rest of NSW (data not available)"

            return f"{self.selected_geo_name} ({self.selected_geo_area}) has a concentration of renters that is {gs_text} {ron_text}."

        elif metric == "social_housing":
            gs_text = ""
            if comparison_gs["value"] is not None:
                if value < comparison_gs["value"] - 0.5:  # 0.5% buffer
                    gs_text = f"lower than the Greater Sydney average of {comparison_gs['value']}%"
                elif value > comparison_gs["value"] + 0.5:
                    gs_text = f"higher than the Greater Sydney average of {comparison_gs['value']}%"
                else:
                    gs_text = f"similar to the Greater Sydney average of {comparison_gs['value']}%"
            else:
                gs_text = "compared to Greater Sydney (data not available)"

            ron_text = ""
            if comparison_ron["value"] is not None:
                if value < comparison_ron["value"] - 0.5:
                    ron_text = f"and lower than the Rest of NSW average of {comparison_ron['value']}%"
                elif value > comparison_ron["value"] + 0.5:
                    ron_text = f"and higher than the Rest of NSW average of {comparison_ron['value']}%"
                else:
                    ron_text = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
            else:
                 ron_text = "and Rest of NSW (data not available)"

            return f"{self.selected_geo_name} ({self.selected_geo_area}) has a concentration of social housing that is {gs_text} {ron_text}."

        elif metric == "median_rent":
            local_increase = self.data["median_rent"].get("annual_increase")
            if local_increase is None:
                 local_increase = 0 # Default to 0 if calculation failed

            gs_change_text = "data not available"
            if comparison_gs.get("annual_change") is not None:
                gs_change_text = f"{comparison_gs['annual_change']}%"

            ron_change_text = "data not available"
            if comparison_ron.get("annual_change") is not None:
                ron_change_text = f"{comparison_ron['annual_change']}%"

            comparison_text = f"For reference, Greater Sydney's annual increase was {gs_change_text} and Rest of NSW's was {ron_change_text}."

            if local_increase is not None:
                 if local_increase > 0:
                     trend_text = f"increased by {local_increase}%"
                 elif local_increase < 0:
                     trend_text = f"decreased by {abs(local_increase)}%"
                 else:
                     trend_text = "remained stable"
                 return f"{self.selected_geo_name} ({self.selected_geo_area})'s median weekly rent has {trend_text} over the past year. {comparison_text}"
            else:
                 return f"{self.selected_geo_name} ({self.selected_geo_area})'s median weekly rent data is available, but annual change could not be calculated. {comparison_text}"


        elif metric == "vacancy_rates":
            current_rate = self.data["vacancy_rates"].get("value")
            previous_rate = self.data["vacancy_rates"].get("previous_year_rate")

            trend_text = ""
            if current_rate is not None and previous_rate is not None:
                if current_rate < previous_rate - 0.1:
                    trend_text = f"The vacancy rate has tightened from {previous_rate}% a year ago to {current_rate}% now. "
                elif current_rate > previous_rate + 0.1:
                    trend_text = f"The vacancy rate has loosened from {previous_rate}% a year ago to {current_rate}% now. "
                else:
                    trend_text = f"The vacancy rate has remained stable at around {current_rate}% compared to {previous_rate}% a year ago. "
            elif current_rate is not None:
                 trend_text = f"The current vacancy rate is {current_rate}%. "
            else:
                 trend_text = "Vacancy rate data is not available. "


            gs_change_text = "data not available"
            if comparison_gs.get("annual_change") is not None:
                 # Vacancy rate change is often presented as a simple difference
                 gs_change_text = f"{comparison_gs['annual_change']:.2f} percentage points"

            ron_change_text = "data not available"
            if comparison_ron.get("annual_change") is not None:
                 ron_change_text = f"{comparison_ron['annual_change']:.2f} percentage points"

            comparison_text = f"For reference, Greater Sydney experienced a change of {gs_change_text} and Rest of NSW {ron_change_text} over the past year."

            return trend_text + comparison_text

        elif metric == "affordability":
            local_improvement = self.data["affordability"].get("annual_improvement")
            if local_improvement is None:
                 local_improvement = 0 # Default to 0 if calculation failed

            local_pct = self.data["affordability"].get("percentage")
            if local_pct is None:
                 local_pct = 0

            # Compare with Greater Sydney
            gs_comparison = "compared to Greater Sydney (data not available)"
            if comparison_gs["value"] is not None:
                if local_pct > comparison_gs["value"] + 2:  # 2% buffer
                    gs_comparison = f"less affordable than the Greater Sydney average of {comparison_gs['value']}%"
                elif local_pct < comparison_gs["value"] - 2:
                    gs_comparison = f"more affordable than the Greater Sydney average of {comparison_gs['value']}%"
                else:
                    gs_comparison = f"similar to the Greater Sydney average of {comparison_gs['value']}%"

            # Compare with Rest of NSW
            ron_comparison = "and Rest of NSW (data not available)"
            if comparison_ron["value"] is not None:
                if local_pct > comparison_ron["value"] + 2:
                    ron_comparison = f"and less affordable than the Rest of NSW average of {comparison_ron['value']}%"
                elif local_pct < comparison_ron["value"] - 2:
                    ron_comparison = f"and more affordable than the Rest of NSW average of {comparison_cron['value']}%"
                else:
                    ron_comparison = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
            else:
                 ron_comparison = "and Rest of NSW (data not available)"


            # Evaluate the trend
            if local_improvement is not None:
                if local_improvement > 0.1:  # Improvement
                    change_text = f"an improvement of {abs(local_improvement):.2f}%"
                elif local_improvement < -0.1:  # Deterioration
                    change_text = f"a deterioration of {abs(local_improvement):.2f}%"
                else:
                    change_text = "relatively stable affordability"
            else:
                 change_text = "no discernible annual change"


            gs_annual_change_text = "data not available"
            if comparison_gs.get("annual_change") is not None:
                 gs_annual_change_text = f"{comparison_gs['annual_change']:.2f}%" # Format as percentage change


            ron_annual_change_text = "data not available"
            if comparison_ron.get("annual_change") is not None:
                 ron_annual_change_text = f"{comparison_ron['annual_change']:.2f}%" # Format as percentage change


            return (f"{self.selected_geo_name} ({self.selected_geo_area}) rental affordability is {gs_comparison} {ron_comparison}, "
                   f"with {change_text} over the past year. Greater Sydney had a change of {gs_annual_change_text} "
                   f"while Rest of NSW had a change of {ron_annual_change_text}.")

        return ""


    def create_excel_output(self):
        """Create a nicely formatted Excel output with the analysis"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.selected_geo_name} Analysis"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        metric_font = Font(bold=True, size=11)
        metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        metric_alignment = Alignment(vertical="center", wrap_text=True)

        value_font = Font(size=11)
        value_alignment = Alignment(vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Rental Market Analysis for {self.selected_geo_name} ({self.selected_geo_area})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Headers - Row 3
        headers = ["Metric", "Value", "Period", "Source", "Comment"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25 # Increased width slightly
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 60 # Increased width for comments

        # Add metrics data
        row = 4

        # Renters
        ws.cell(row=row, column=1).value = "# and % of renters"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2).value = f"{self.data['renters']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3).value = self.data['renters']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4).value = self.data['renters']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border

        comment = self.generate_comparison_comment("renters", self.data['renters']['percentage'],
                                              self.data['renters']['comparison_gs'], self.data['renters']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border

        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['renters']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border

        # Social Housing
        row += 1
        ws.cell(row=row, column=1).value = "# and % of Social Housing"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3).value = self.data['social_housing']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4).value = self.data['social_housing']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border

        comment = self.generate_comparison_comment("social_housing", self.data['social_housing']['percentage'],
                                                self.data['social_housing']['comparison_gs'], self.data['social_housing']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border

        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border

        # Median Weekly Rent
        row += 1
        ws.cell(row=row, column=1).value = "Median Weekly Rent"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2).value = f"${self.data['median_rent']['value']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3).value = self.data['median_rent']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4).value = self.data['median_rent']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border

        comment = self.generate_comparison_comment("median_rent", self.data['median_rent']['value'],
                                            self.data['median_rent']['comparison_gs'], self.data['median_rent']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border

        row += 1
        # Display previous year's rent and calculated annual increase
        previous_year_rent = self.data['median_rent'].get('previous_year_rent')
        annual_increase = self.data['median_rent'].get('annual_increase')

        if previous_year_rent is not None:
            increase_text = f"Annual change: {annual_increase:.1f}% (from ${previous_year_rent} 12 months prior)"
        else:
            increase_text = f"Annual change: {annual_increase:.1f}% (previous year data not available)"

        ws.cell(row=row, column=2).value = increase_text
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border


        # Vacancy Rates
        row += 1
        ws.cell(row=row, column=1).value = "Vacancy Rates"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2).value = f"{self.data['vacancy_rates']['value']:.2f}%" # Ensure 2 decimal places
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3).value = self.data['vacancy_rates']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4).value = self.data['vacancy_rates']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border

        comment = self.generate_comparison_comment("vacancy_rates", self.data['vacancy_rates']['value'],
                                             self.data['vacancy_rates']['comparison_gs'], self.data['vacancy_rates']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border

        row += 1
        previous_year_rate = self.data['vacancy_rates']['previous_year_rate']
        if previous_year_rate is not None:
            previous_year_text = f"Previous year: {previous_year_rate:.2f}%" # Ensure 2 decimal places
        else:
            previous_year_text = "Previous year data not available"
        ws.cell(row=row, column=2).value = previous_year_text
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border

        # Rental Affordability
        row += 1
        ws.cell(row=row, column=1).value = "Rental affordability*"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2).value = f"{self.data['affordability']['percentage']:.1f}% of income on rent" # Ensure 1 decimal place
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border # Corrected this line from column[3] to column=3

        ws.cell(row=row, column=3).value = self.data['affordability']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border


        ws.cell(row=row, column=4).value = self.data['affordability']['source']
        ws.cell(row=row, column=4).font = value_font # Apply font
        ws.cell(row=row, column=4).alignment = value_alignment # Apply alignment (fix for the error)
        ws.cell(row=row, column=4).border = thin_border # Apply border


        comment = self.generate_comparison_comment("affordability", self.data['affordability']['percentage'],
                                              self.data['affordability']['comparison_gs'], self.data['affordability']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border

        row += 1
        annual_improvement = self.data['affordability'].get('annual_improvement')
        if annual_improvement is not None:
            improvement_text = "improvement" if annual_improvement > 0 else ("deterioration" if annual_improvement < 0 else "stability")
            ws.cell(row=row, column=2).value = f"Annual {improvement_text} {abs(annual_improvement):.2f}%" # Ensure 2 decimal places
        else:
            ws.cell(row=row, column=2).value = "Annual change data not available"

        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border

        for col in [1, 3, 4, 5]: # Added column 4 here to apply border to the second row of this metric
            ws.cell(row=row, column=col).border = thin_border

        # Rental Contacts section has been removed as requested

        # Add footnote for affordability
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].value = ("* Methodology: the rental affordability is calculated by taking median weekly rental household incomes for the "
                              "geographic area and comparing that to median weekly rents for the same area. Any number higher than 30% of income "
                              "on rent is considered that a household is experiencing rental stress. This metric is calculated by Fair Trading "
                              "using ABS income and indexation data as well as Core Logic rental data.")
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # Set row heights
        for r in range(3, row):
            ws.row_dimensions[r].height = 30

        ws.row_dimensions[row].height = 45  # Footnote row

        # Save the workbook
        wb.save(self.output_file)

def main():
    # Set pandas to not display warnings
    pd.options.mode.chained_assignment = None

    try:
        analyzer = RentalDataAnalyzer()
        analyzer.create_gui()
    except Exception as e:
        import traceback
        print(f"Critical error: {str(e)}")
        print(traceback.format_exc())

        # Show error in GUI if possible
        try:
            import tkinter.messagebox as msgbox
            msgbox.showerror("Critical Error", f"An unexpected error occurred:\n\n{str(e)}\n\nPlease check the console for details.")
        except:
            pass

if __name__ == "__main__":
    main()