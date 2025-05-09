import os
import pandas as pd
import numpy as np
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import re

class RentalDataAnalyzer:
    def __init__(self):
        """Initialize the Rental Data Analyzer with paths and reference data"""
        # Base directory - can be modified by user via GUI
        self.BASE_DIR = r"C:\Users\joshu\OneDrive - NSWGOV\Rental Data Model\Calculated rental data"
        
        # Geographic areas
        self.GEO_AREAS = ["CED", "GCCSA", "LGA", "SA3", "SA4", "SED", "Suburb"]
        
        # Data subdirectories
        self.SUB_DIRS = {
            "median_rents": os.path.join("Median rents", "output data"),
            "census_dwelling": os.path.join("Census data", "output data", "dwellings"),
            "census_demographics": os.path.join("Census data", "output data", "demographics"),
            "affordability": os.path.join("Affordability", "output data"),
            "vacancy_rates": os.path.join("Rental vacancy rates", "output data")
        }
        
        # File patterns for different geographic areas
        self.FILE_PATTERNS = {
            "median_rents": {area.lower(): f"{area.lower()}_rent_data" for area in self.GEO_AREAS},
            "affordability": {area.lower(): f"{area.lower()}_affordability" for area in self.GEO_AREAS},
            "vacancy_rates": {area.lower(): f"{area.lower()}_vacancy_rate" for area in self.GEO_AREAS},
            "census_dwelling": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_dwelling_tenure" for area in self.GEO_AREAS},
            "census_demographics": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_demographics" for area in self.GEO_AREAS}
        }
        
        # Greater Sydney LGAs
        self.GREATER_SYDNEY_LGAS = [
            "Bayside (NSW)", "Blacktown", "Blue Mountains", "Burwood", "Camden", "Campbelltown (NSW)", 
            "Canada Bay", "Canterbury-Bankstown", "Cumberland", "Fairfield", "Georges River", 
            "Hawkesbury", "Hornsby", "Hunters Hill", "Inner West", "Ku-ring-gai", "Lane Cove", 
            "Liverpool", "Mosman", "North Sydney", "Northern Beaches", "Parramatta", "Penrith", 
            "Randwick", "Ryde", "Strathfield", "Sutherland Shire", "Sydney", "The Hills Shire", 
            "Waverley", "Willoughby", "Woollahra", "Wollondilly"
        ]
        
        # Reference data for comparison - Greater Sydney
        self.GS_REFERENCE_DATA = {
            "renters": {"area": "Greater Sydney", "value": 32.6},
            "social_housing": {"area": "Greater Sydney", "value": 4.5},
            "median_rent": {"area": "Greater Sydney", "value": 12.1},
            "vacancy_rates": {"area": "Greater Sydney", "value": -0.3},
            "affordability": {"area": "Greater Sydney", "value": 45.2, "annual_change": 5.1}
        }
        
        # Reference data for comparison - Rest of NSW
        self.RON_REFERENCE_DATA = {
            "renters": {"area": "Rest of NSW", "value": 26.8},
            "social_housing": {"area": "Rest of NSW", "value": 4.0},
            "median_rent": {"area": "Rest of NSW", "value": 8.6},
            "vacancy_rates": {"area": "Rest of NSW", "value": -0.1},
            "affordability": {"area": "Rest of NSW", "value": 41.7, "annual_change": 3.5}
        }
        
        # Variables to store selections and data
        self.selected_geo_area = None
        self.selected_geo_name = None
        self.data = {}
        self.output_file = None

    def create_gui(self):
        """Create GUI for selecting geographic area and name"""
        self.root = tk.Tk()
        self.root.title("NSW Rental Data Analyzer")
        self.root.geometry("700x400")
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Data directory selection
        dir_frame = ttk.LabelFrame(main_frame, text="Data Directory", padding="5")
        dir_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.dir_var = tk.StringVar(value=self.BASE_DIR)
        dir_entry = ttk.Entry(dir_frame, textvariable=self.dir_var, width=60)
        dir_entry.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        dir_button = ttk.Button(dir_frame, text="Browse...", command=self.browse_directory)
        dir_button.pack(side=tk.RIGHT, padx=5, pady=5)
        
        # Selection frame
        selection_frame = ttk.LabelFrame(main_frame, text="Geographic Selection", padding="5")
        selection_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Geographic area type selection
        ttk.Label(selection_frame, text="Geographic Area Type:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo = ttk.Combobox(selection_frame, values=self.GEO_AREAS, state="readonly", width=30)
        self.geo_area_combo.grid(column=1, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo.bind("<<ComboboxSelected>>", self.on_geo_area_selected)
        
        # Geographic name selection
        ttk.Label(selection_frame, text="Geographic Name:").grid(column=0, row=1, sticky=tk.W, pady=5, padx=5)
        self.geo_name_combo = ttk.Combobox(selection_frame, state="disabled", width=40)
        self.geo_name_combo.grid(column=1, row=1, sticky=tk.W, pady=5, padx=5)
        
        # Output file selection
        output_frame = ttk.LabelFrame(main_frame, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.output_var = tk.StringVar()
        ttk.Label(output_frame, text="Output File:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=40)
        output_entry.grid(column=1, row=0, sticky=tk.W+tk.E, pady=5, padx=5)
        
        output_button = ttk.Button(output_frame, text="Browse...", command=self.browse_output_file)
        output_button.grid(column=2, row=0, sticky=tk.W, pady=5, padx=5)
        
        # Status and buttons
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, padx=5, pady=10)
        
        # Generate button
        self.generate_button = ttk.Button(status_frame, text="Generate Analysis", command=self.run_analysis, state="disabled")
        self.generate_button.pack(side=tk.RIGHT, padx=5)
        
        # Exit button
        exit_button = ttk.Button(status_frame, text="Exit", command=self.root.destroy)
        exit_button.pack(side=tk.RIGHT, padx=5)
        
        # Status label
        self.status_label = ttk.Label(status_frame, text="Select a geographic area type to begin")
        self.status_label.pack(side=tk.LEFT, padx=5)
        
        # Set column and row weights for resizing
        main_frame.columnconfigure(0, weight=1)
        
        self.root.mainloop()
    
    def browse_directory(self):
        """Browse for data directory"""
        dir_path = filedialog.askdirectory(initialdir=self.BASE_DIR, title="Select Data Directory")
        if dir_path:
            self.BASE_DIR = dir_path
            self.dir_var.set(dir_path)
            # Reset selections
            self.selected_geo_area = None
            self.selected_geo_name = None
            self.geo_area_combo.set('')
            self.geo_name_combo.set('')
            self.geo_name_combo.config(state="disabled")
            self.generate_button.config(state="disabled")
            self.status_label.config(text="Select a geographic area type to begin")
    
    def browse_output_file(self):
        """Browse for output file location"""
        file_path = filedialog.asksaveasfilename(
            initialdir=os.path.expanduser("~"), 
            title="Save Analysis As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.output_file = file_path
            self.output_var.set(file_path)
    
    def on_geo_area_selected(self, event=None):
        """Handle geographic area type selection"""
        self.selected_geo_area = self.geo_area_combo.get()
        self.status_label.config(text=f"Loading {self.selected_geo_area} names...")
        self.root.update_idletasks()
        
        # Get available geographic names
        try:
            geo_names = self.get_geo_names(self.selected_geo_area)
            self.geo_name_combo.config(values=geo_names, state="readonly")
            self.geo_name_combo.bind("<<ComboboxSelected>>", self.on_geo_name_selected)
            self.status_label.config(text=f"Select a {self.selected_geo_area} name")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to load geographic names: {str(e)}")
    
    def on_geo_name_selected(self, event=None):
        """Handle geographic name selection"""
        self.selected_geo_name = self.geo_name_combo.get()
        self.generate_button.config(state="normal")
        
        # Set default output file name if not already set
        if not self.output_var.get():
            # Sanitize filename - remove parentheses and other special characters
            safe_geo_name = self.selected_geo_name.replace("(", "").replace(")", "").replace(" ", "_")
            default_filename = f"{safe_geo_name}_{self.selected_geo_area}_Rental_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # Ensure desktop path exists
            try:
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop_path):
                    # Fall back to current directory if Desktop doesn't exist
                    desktop_path = os.getcwd()
            except:
                desktop_path = os.getcwd()
                
            self.output_var.set(os.path.join(desktop_path, default_filename))
            self.output_file = os.path.join(desktop_path, default_filename)
        
        self.status_label.config(text=f"Ready to generate analysis for {self.selected_geo_name} ({self.selected_geo_area})")
    
    def get_geo_names(self, geo_area):
        """Get available geographic names for the selected area type"""
        names = set()
        found_files = False
        
        # Try to get names from multiple data sources
        for data_type, subdir in self.SUB_DIRS.items():
            try:
                dir_path = os.path.join(self.BASE_DIR, subdir)
                if not os.path.exists(dir_path):
                    continue
                
                # Skip if pattern doesn't exist for this geo_area
                if geo_area.lower() not in self.FILE_PATTERNS[data_type]:
                    continue
                    
                file_pattern = self.FILE_PATTERNS[data_type][geo_area.lower()]
                
                for file in os.listdir(dir_path):
                    if file_pattern.lower() in file.lower():
                        found_files = True
                        file_path = os.path.join(dir_path, file)
                        print(f"Reading file: {file_path}")
                        df = self.read_data_file(file_path)
                        
                        if df is not None and not df.empty:
                            # Look for the geographic name column
                            geo_col = self.find_geographic_column(df, geo_area)
                            
                            if geo_col:
                                print(f"Found geography column: {geo_col}")
                                # Convert all values to strings and filter out likely non-geographic names
                                df[geo_col] = df[geo_col].astype(str)
                                
                                # Filter out values that look like dates or numbers
                                area_names = []
                                for name in df[geo_col].dropna().unique().tolist():
                                    name_str = str(name)
                                    # Skip if it looks like a date format (2021-01-01, etc.)
                                    if re.match(r'^\d{4}-\d{2}-\d{2}', name_str) or re.match(r'^\d{2}/\d{2}/\d{4}', name_str):
                                        continue
                                    # Skip if it's just a number
                                    if name_str.isdigit():
                                        continue
                                    # Skip very short names (likely codes)
                                    if len(name_str) < 2:
                                        continue
                                    # Skip if it's an LGA/SA code
                                    if re.match(r'^LGA\d+$', name_str) or re.match(r'^SA\d+$', name_str):
                                        continue
                                    
                                    area_names.append(name_str)
                                
                                print(f"Found {len(area_names)} geographic names")
                                names.update(area_names)
            except Exception as e:
                print(f"Error getting names from {data_type}: {str(e)}")
        
        if not found_files:
            raise Exception(f"No data files found for {geo_area}. Please check your data directory.")
        
        if not names:
            raise Exception(f"No geographic names found for {geo_area}. Check that your data files contain the expected columns.")
            
        name_list = sorted(list(names))
        if name_list:
            print(f"Final list of geographic names: {name_list[:10]}... (showing first 10 of {len(name_list)})")
        else:
            print("No geographic names found.")
        return name_list
    
    def read_data_file(self, file_path):
        """Read data from Excel or Parquet file"""
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # Use explicit converter and dtype to avoid datetime conversion issues
                return pd.read_excel(
                    file_path,
                    dtype={0: str},  # Force first column to be string
                    converters={0: str}  # Ensure first column is converted to string
                )
            elif file_path.endswith('.parquet'):
                df = pq.read_table(file_path).to_pandas()
                
                # Ensure the first column is a string to avoid comparison issues
                if len(df.columns) > 0:
                    df[df.columns[0]] = df[df.columns[0]].astype(str)
                    
                return df
            else:
                print(f"Unsupported file format: {file_path}")
                return None
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            return None
    
    def find_geographic_column(self, df, geo_area):
        """Find the column containing geographic area names"""
        # Print the column names for debugging
        print(f"Available columns: {df.columns.tolist()}")
        
        # Direct matches - highest priority columns that definitely contain geographic names
        priority_columns = ['region_name', 'area_name', 'location_name', f'{geo_area.lower()}_name', 'name']
        
        # Check for exact matches in priority columns first
        for col in priority_columns:
            if col in df.columns:
                print(f"Found priority geography column: {col}")
                
                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col
        
        # Common names for geographic columns - next priority
        geo_keywords = [
            geo_area.lower(), 'name', 'area', 'region', 'location', 'geography',
            'district', 'locality', 'suburb', 'lga', 'sa3', 'sa4', 'gccsa', 'ced', 'sed'
        ]
        
        # Check for columns that contain these keywords
        for col in df.columns:
            col_lower = str(col).lower()
            for keyword in geo_keywords:
                if keyword in col_lower and 'code' not in col_lower and 'type' not in col_lower:
                    print(f"Found geography column by keyword: {col}")
                    # Show sample values
                    if not df.empty:
                        sample = df[col].dropna().head(5).astype(str).tolist()
                        print(f"Sample values from '{col}': {sample}")
                    return col
        
        # If we haven't found a match yet, specifically look for 'region_name' which we know exists
        if 'region_name' in df.columns:
            print(f"Falling back to region_name column")
            # Show sample values
            if not df.empty:
                sample = df['region_name'].dropna().head(5).astype(str).tolist()
                print(f"Sample values from 'region_name': {sample}")
            return 'region_name'
            
        # If still no match, check for columns that might contain place names
        for col in df.columns:
            try:
                sample = df[col].dropna().head(5).astype(str).tolist()
                print(f"Checking column '{col}' with sample values: {sample}")
                
                # Skip columns where values appear to be dates or numbers
                if all(not re.match(r'^\d{4}-\d{2}-\d{2}', str(x)) and 
                       not re.match(r'^\d{2}/\d{2}/\d{4}', str(x)) and
                       not str(x).replace('.', '').isdigit() and
                       len(str(x)) > 2
                       for x in sample):
                    
                    # Check if the values look like place names (contain alphabetic characters)
                    if all(any(c.isalpha() for c in str(x)) for x in sample):
                        print(f"Detected possible location names in column: {col}")
                        return col
            except Exception as e:
                print(f"Error checking column {col}: {str(e)}")
        
        # If no suitable column found, look for codes as a last resort
        for col in df.columns:
            col_lower = str(col).lower()
            if ('code' in col_lower and any(kw in col_lower for kw in geo_keywords)) or 'region_code' in col_lower:
                print(f"Falling back to code column: {col}")
                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col
                
        # Absolute last resort - first column
        if len(df.columns) > 0:
            first_col = df.columns[0]
            print(f"No geographic column identified, using first column as fallback: {first_col}")
            # Show sample values
            if not df.empty:
                sample = df[first_col].dropna().head(5).astype(str).tolist()
                print(f"Sample values from first column: {sample}")
            return first_col
        
        return None
    
    def run_analysis(self):
        """Run the analysis and generate Excel output"""
        if not self.selected_geo_area or not self.selected_geo_name:
            messagebox.showerror("Error", "Please select both a geographic area type and name")
            return
        
        if not self.output_file:
            messagebox.showerror("Error", "Please select an output file location")
            return
        
        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    messagebox.showerror("Error", f"Cannot create output directory: {str(e)}")
                    return
                    
            # Make sure we have a valid file path for saving
            if not self.output_file.endswith('.xlsx'):
                self.output_file += '.xlsx'
                
            self.status_label.config(text="Collecting data...")
            self.root.update_idletasks()
            
            # Collect data from various sources
            self.collect_data()
            
            self.status_label.config(text="Creating Excel output...")
            self.root.update_idletasks()
            
            # Create Excel output
            self.create_excel_output()
            
            self.status_label.config(text=f"Analysis completed and saved to {self.output_file}")
            messagebox.showinfo("Success", f"Analysis completed and saved to {self.output_file}")
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in run_analysis: {str(e)}")
            print(error_details)
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate analysis: {str(e)}\n\nSee console for details.")
    
    def collect_data(self):
        """Collect data from various sources"""
        self.data = {}
        print(f"\n\n{'='*50}")
        print(f"STARTING DATA COLLECTION FOR: {self.selected_geo_name} ({self.selected_geo_area})")
        print(f"{'='*50}")
        
        # Determine if this is Greater Sydney or Rest of NSW
        is_greater_sydney = self.selected_geo_name in self.GREATER_SYDNEY_LGAS
        reference_data = self.GS_REFERENCE_DATA if is_greater_sydney else self.RON_REFERENCE_DATA
        print(f"Using reference data for: {reference_data['renters']['area']}")
        
        # Collect Census dwelling data
        print("\nCollecting Census dwelling data...")
        try:
            # Find census dwelling file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["census_dwelling"])
            print(f"Looking for census files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["census_dwelling"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            census_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    census_files_found.append(file)
            
            print(f"Census files found: {census_files_found}")
            
            for file in census_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing census file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Print a few sample values from the column
                        sample_values = df[geo_col].dropna().unique()[:10].tolist()
                        print(f"Sample values in column: {sample_values}")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # Calculate rental percentage by finding total rental and total dwellings
                            # Method 1: Look for direct percentage columns
                            pct_col = None
                            for col in df_filtered.columns:
                                if "rented" in col.lower() and "percent" in col.lower():
                                    pct_col = col
                                    break
                            
                            # Method 2: Calculate from raw counts
                            if pct_col is None:
                                total_dwellings = None
                                total_rented = None
                                
                                # Find total dwellings
                                if "dwellings" in df_filtered.columns:
                                    total_dwellings = float(df_filtered["dwellings"].values[0])
                                
                                # Find total rented
                                if "dwellings_rented" in df_filtered.columns:
                                    total_rented = float(df_filtered["dwellings_rented"].values[0])
                                
                                # Calculate percentage
                                if total_dwellings is not None and total_rented is not None and total_dwellings > 0:
                                    rental_pct = (total_rented / total_dwellings) * 100
                                    rental_count = int(total_rented)
                                    print(f"Calculated rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")
                                    
                                    self.data["renters"] = {
                                        "percentage": round(rental_pct, 1),
                                        "count": rental_count,
                                        "period": "2021",
                                        "source": "ABS Census",
                                        "comparison": reference_data["renters"]
                                    }
                            else:
                                rental_pct = float(df_filtered[pct_col].values[0])
                                renter_count_col = pct_col.replace("percent", "count")
                                rental_count = int(df_filtered[renter_count_col].values[0]) if renter_count_col in df_filtered.columns else 0
                                print(f"Found rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")
                                
                                self.data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison": reference_data["renters"]
                                }
                            
                            # Find social housing data - combine SHA, CHP columns
                            total_social = 0
                            social_cols = [col for col in df_filtered.columns if "sha" in col.lower() or "chp" in col.lower()]
                            
                            for col in social_cols:
                                if "count" in col.lower() or not any(x in col.lower() for x in ["percent", "pct", "ratio"]):
                                    if not pd.isna(df_filtered[col].values[0]):
                                        total_social += float(df_filtered[col].values[0])
                            
                            # Calculate social housing percentage
                            if total_dwellings is not None and total_dwellings > 0:
                                social_pct = (total_social / total_dwellings) * 100
                                social_count = int(total_social)
                                print(f"Calculated social housing percentage: {social_pct:.1f}% ({social_count} dwellings)")
                                
                                self.data["social_housing"] = {
                                    "percentage": round(social_pct, 1),
                                    "count": social_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison": reference_data["social_housing"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting census data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "renters" not in self.data:
            print("Using DEFAULT renter data")
            self.data["renters"] = {
                "percentage": 25.5,
                "count": 8402,
                "period": "2021",
                "source": "ABS Census",
                "comparison": reference_data["renters"]
            }
            
        if "social_housing" not in self.data:
            print("Using DEFAULT social housing data")
            self.data["social_housing"] = {
                "percentage": 2.8,
                "count": 938,
                "period": "2021",
                "source": "ABS Census",
                "comparison": reference_data["social_housing"]
            }
        
        # Collect Median Rent data
        print("\nCollecting Median Rent data...")
        try:
            # Find median rent file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["median_rents"])
            print(f"Looking for median rent files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["median_rents"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            rent_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    rent_files_found.append(file)
            
            print(f"Rent files found: {rent_files_found}")
            
            for file in rent_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing rent file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Print a few sample values from the column
                        sample_values = df[geo_col].dropna().unique()[:10].tolist()
                        print(f"Sample values in column: {sample_values}")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            latest_month = None
                            if 'month' in df_filtered.columns:
                                try:
                                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                    latest_month = df_filtered['month'].max()
                                    
                                    # Format latest_month for output
                                    latest_month_str = latest_month.strftime("%b-%Y")  # e.g., "Apr-2025"
                                    
                                    df_latest = df_filtered[df_filtered['month'] == latest_month]
                                    print(f"Filtered to latest month: {latest_month} ({latest_month_str})")
                                except Exception as e:
                                    print(f"Error filtering to latest month: {str(e)}")
                                    df_latest = df_filtered
                            else:
                                df_latest = df_filtered
                                
                            # If we have property_type, get the "All Dwellings" type
                            if 'property_type' in df_latest.columns:
                                if 'All Dwellings' in df_latest['property_type'].values:
                                    df_latest = df_latest[df_latest['property_type'] == 'All Dwellings']
                                    print("Filtered to 'All Dwellings' property type")
                            
                            # Find columns for median rent data - prefer 3-month median
                            rent_col = None
                            for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                                for col in df_latest.columns:
                                    if col.startswith(col_prefix) and not any(x in col for x in ['growth', 'increase', 'change']):
                                        rent_col = col
                                        break
                                if rent_col:
                                    break
                                    
                            # Find annual growth column
                            growth_col = None
                            for col_suffix in ['annual_growth', 'annual_increase', 'yearly_growth', 'yearly_increase']:
                                for col in df_latest.columns:
                                    if col.endswith(col_suffix):
                                        growth_col = col
                                        break
                                if growth_col:
                                    break
                            
                            print(f"Median rent column: {rent_col}")
                            print(f"Annual increase column: {growth_col}")
                            
                            # Extract data
                            if rent_col:
                                # Get the median rent value
                                if len(df_latest) > 0:
                                    rent_value = float(df_latest[rent_col].values[0])
                                    print(f"Median rent: ${rent_value}")
                                    
                                    # Get annual increase if available
                                    annual_increase = None
                                    if growth_col and len(df_latest) > 0:
                                        annual_increase_value = df_latest[growth_col].values[0]
                                        if not pd.isna(annual_increase_value):
                                            annual_increase = float(annual_increase_value) * 100 if float(annual_increase_value) < 1 else float(annual_increase_value)
                                            print(f"Annual increase: {annual_increase:.1f}%")
                                    
                                    self.data["median_rent"] = {
                                        "value": int(round(rent_value, 0)),
                                        "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                        "source": "NSW Fair Trading Corelogic Data",
                                        "annual_increase": round(annual_increase, 1) if annual_increase is not None else reference_data["median_rent"]["value"],
                                        "comparison": reference_data["median_rent"]
                                    }
                    break
        except Exception as e:
            print(f"Error collecting median rent data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "median_rent" not in self.data:
            print("Using DEFAULT median rent data")
            self.data["median_rent"] = {
                "value": 595,
                "period": "Apr-25",
                "source": "NSW Fair Trading Corelogic Data",
                "annual_increase": reference_data["median_rent"]["value"],
                "comparison": reference_data["median_rent"]
            }
        
        # Collect Vacancy Rate data
        print("\nCollecting Vacancy Rate data...")
        try:
            # Find vacancy rate file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["vacancy_rates"])
            print(f"Looking for vacancy rate files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["vacancy_rates"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            vacancy_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    vacancy_files_found.append(file)
            
            print(f"Vacancy rate files found: {vacancy_files_found}")
            
            for file in vacancy_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing vacancy rate file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            if 'month' in df_filtered.columns:
                                df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                latest_month = df_filtered['month'].max()
                                df_latest = df_filtered[df_filtered['month'] == latest_month]
                                print(f"Filtered to latest month: {latest_month}")
                            else:
                                df_latest = df_filtered
                                
                            # Find vacancy rate column
                            rate_col = None
                            for col_name in ['rental_vacancy_rate', 'vacancy_rate', 'rate']:
                                if col_name in df_latest.columns:
                                    rate_col = col_name
                                    break
                                    
                            # Find annual change column
                            change_col = None
                            for col_suffix in ['annual_change', 'yearly_change', 'annual_growth']:
                                for col in df_latest.columns:
                                    if col_suffix in col.lower():
                                        change_col = col
                                        break
                                if change_col:
                                    break
                            
                            # If we don't have an annual change column but have multiple months of data,
                            # try to calculate the annual change ourselves
                            if not change_col and 'month' in df_filtered.columns and rate_col:
                                try:
                                    # Get current month's value
                                    current_value = float(df_latest[rate_col].values[0])
                                    
                                    # Try to find data from a year ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    year_ago_data = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    if not year_ago_data.empty:
                                        year_ago_value = float(year_ago_data[rate_col].values[0])
                                        annual_change = current_value - year_ago_value
                                        
                                        # Create a placeholder dataframe column with our calculated value
                                        df_latest['calculated_annual_change'] = annual_change
                                        change_col = 'calculated_annual_change'
                                        print(f"Calculated annual change manually: {annual_change:.2f}%")
                                except Exception as e:
                                    print(f"Error calculating annual change: {str(e)}")
                            
                            print(f"Vacancy rate column: {rate_col}")
                            print(f"Annual change column: {change_col}")
                            
                            # Extract data
                            if rate_col and len(df_latest) > 0:
                                rate_value = float(df_latest[rate_col].values[0])
                                
                                # Ensure it's formatted as a percentage
                                if rate_value > 0 and rate_value < 1:
                                    rate_value = rate_value * 100  # Convert decimal to percentage
                                
                                # Get annual change if available
                                annual_change = None
                                if change_col and len(df_latest) > 0:
                                    annual_change_value = df_latest[change_col].values[0]
                                    if not pd.isna(annual_change_value):
                                        annual_change = float(annual_change_value) * 100 if float(annual_change_value) < 1 else float(annual_change_value)
                                
                                # Default to no change if we couldn't determine annual change
                                if annual_change is None:
                                    annual_change = 0.0
                                    print(f"Using default annual change value of {annual_change}")
                                
                                self.data["vacancy_rates"] = {
                                    "value": round(rate_value, 2),
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "annual_change": round(annual_change, 2),
                                    "comparison": reference_data["vacancy_rates"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting vacancy rate data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "vacancy_rates" not in self.data:
            print("Using DEFAULT vacancy rate data")
            self.data["vacancy_rates"] = {
                "value": 0.72,
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "annual_change": reference_data["vacancy_rates"]["value"],
                "comparison": reference_data["vacancy_rates"]
            }
        
        # Collect Affordability data
        print("\nCollecting Affordability data...")
        try:
            # Find affordability file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["affordability"])
            print(f"Looking for affordability files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["affordability"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            affordability_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    affordability_files_found.append(file)
            
            print(f"Affordability files found: {affordability_files_found}")
            
            for file in affordability_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing affordability file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            if 'month' in df_filtered.columns:
                                df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                latest_month = df_filtered['month'].max()
                                df_latest = df_filtered[df_filtered['month'] == latest_month]
                                print(f"Filtered to latest month: {latest_month}")
                            else:
                                df_latest = df_filtered
                                
                            # Find affordability column - look for keywords
                            pct_col = None
                            
                            # First priority: direct affordability columns
                            affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                            if affordability_columns:
                                # Prefer 3-month affordability for stability
                                if 'rental_affordability_3mo' in affordability_columns:
                                    pct_col = 'rental_affordability_3mo'
                                elif 'rental_affordability_1mo' in affordability_columns:
                                    pct_col = 'rental_affordability_1mo'
                                else:
                                    pct_col = affordability_columns[0]  # Take the first one if no preferred column
                            
                            # If no direct affordability column, try to find rent-to-income ratio
                            if not pct_col:
                                for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                                    for col in df_latest.columns:
                                        if all(keyword.lower() in col.lower() for keyword in keywords):
                                            pct_col = col
                                            break
                                    if pct_col:
                                        break
                            
                            # If still no column, try to calculate it ourselves if we have rent and income columns
                            if not pct_col:
                                median_rent_cols = [col for col in df_latest.columns if 'median_rent' in col.lower() and not any(x in col.lower() for x in ['growth', 'increase', 'change'])]
                                income_cols = [col for col in df_latest.columns if 'income' in col.lower() and 'index' not in col.lower()]
                                
                                if median_rent_cols and income_cols and len(df_latest) > 0:
                                    # Prefer 3-month median rent for stability
                                    rent_col = None
                                    for prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent']:
                                        matching_cols = [col for col in median_rent_cols if col.startswith(prefix)]
                                        if matching_cols:
                                            rent_col = matching_cols[0]
                                            break
                                    
                                    if not rent_col and median_rent_cols:
                                        rent_col = median_rent_cols[0]
                                    
                                    # Use the first income column we find
                                    income_col = income_cols[0]
                                    
                                    if rent_col and income_col:
                                        rent_value = float(df_latest[rent_col].values[0])
                                        income_value = float(df_latest[income_col].values[0])
                                        
                                        if income_value > 0:
                                            # Calculate weekly rental affordability (rent/income)
                                            # For weekly rent and annual income: (weekly_rent * 52) / annual_income * 100
                                            affordability = (rent_value * 52) / income_value * 100
                                            
                                            # Create a placeholder dataframe with our calculated value
                                            df_latest['calculated_affordability'] = affordability
                                            pct_col = 'calculated_affordability'
                                            print(f"Calculated affordability manually: {affordability:.1f}%")
                            
                            # Find improvement column
                            improvement_col = None
                            for col_suffix in ['improvement', 'change', 'growth']:
                                for col in df_latest.columns:
                                    if col_suffix in col.lower() and any(x in col.lower() for x in ['annual', 'yearly']):
                                        improvement_col = col
                                        break
                                if improvement_col:
                                    break
                            
                            print(f"Affordability column: {pct_col}")
                            print(f"Annual improvement column: {improvement_col}")
                            
                            # Extract data
                            if pct_col and len(df_latest) > 0:
                                pct_value = float(df_latest[pct_col].values[0])
                                
                                # Ensure the value is properly formatted as a percentage
                                if pct_value > 0 and pct_value < 1:
                                    pct_value = pct_value * 100  # Convert decimal to percentage
                                
                                # Get annual improvement if available
                                annual_improvement = None
                                if improvement_col and len(df_latest) > 0:
                                    annual_improvement_value = df_latest[improvement_col].values[0]
                                    if not pd.isna(annual_improvement_value):
                                        annual_improvement = float(annual_improvement_value) * 100 if float(annual_improvement_value) < 1 else float(annual_improvement_value)
                                
                                # If we couldn't find an annual improvement column but have multiple months of data,
                                # try to calculate it manually by comparing with previous year
                                if annual_improvement is None and 'month' in df_filtered.columns:
                                    try:
                                        # Get current month's value
                                        current_value = pct_value
                                        
                                        # Try to find data from a year ago
                                        one_year_ago = latest_month - pd.DateOffset(months=12)
                                        one_year_ago_str = one_year_ago.strftime("%b-%Y")  # e.g., "Apr-2024"
                                        
                                        year_ago_data = df_filtered[df_filtered['month'] == one_year_ago]
                                        
                                        if not year_ago_data.empty and pct_col in year_ago_data.columns:
                                            year_ago_value = float(year_ago_data[pct_col].values[0])
                                            if year_ago_value > 0 and year_ago_value < 1:
                                                year_ago_value = year_ago_value * 100  # Convert decimal to percentage
                                            
                                            # Calculate change - note that for affordability, a decrease is an improvement
                                            # (lower percentage of income spent on rent is better)
                                            change = current_value - year_ago_value
                                            annual_improvement = -change  # Negative change = improvement for affordability
                                            
                                            print(f"Calculated annual affordability change: {change:.2f}% (improvement: {annual_improvement:.2f}%)")
                                            print(f"Comparing {latest_month_str} ({current_value:.1f}%) to {one_year_ago_str} ({year_ago_value:.1f}%)")
                                        else:
                                            print(f"No data found for {one_year_ago_str} to calculate annual improvement")
                                    except Exception as e:
                                        print(f"Error calculating annual improvement: {str(e)}")
                                
                                # Use default annual change if we couldn't calculate it
                                if annual_improvement is None:
                                    annual_improvement = reference_data["affordability"]["annual_change"]
                                
                                self.data["affordability"] = {
                                    "percentage": round(pct_value, 1),
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "annual_improvement": round(annual_improvement, 2) if annual_improvement is not None else reference_data["affordability"]["annual_change"],
                                    "comparison": reference_data["affordability"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting affordability data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "affordability" not in self.data:
            print("Using DEFAULT affordability data")
            self.data["affordability"] = {
                "percentage": 43.6,
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "annual_improvement": reference_data["affordability"]["annual_change"],
                "comparison": reference_data["affordability"]
            }
        
        # Add rental contacts data (placeholder)
        self.data["rental_contacts"] = {
            "total": 85,
            "enquiries": 47,
            "complaints": 38,
            "period": "FY23/24 & FY24/25 (up to April'25)",
            "source": "NSW Fair Trading",
            "details": (f"Matters relating to rent and charges & repairs and maintenance was the top "
                       f"enquiries (8 each) while matters relating to repairs and maintenance was the "
                       f"top complaint (16) to Fair Trading. * Data is as of and up to April'25. "
                       f"Suburbs in the {self.selected_geo_name} {self.selected_geo_area.lower()} that "
                       f"has data includes Bonny hills, Dunbogan, Hannam Vale, Lake Cathie, "
                       f"Port Macquarie, Thrumster and Upsalls Creek.")
        }
        
        print("\nData collection complete")
        print(f"{'='*50}\n")
    
    def find_column(self, df, keywords):
        """Find column that contains all the keywords (case-insensitive)"""
        print(f"Looking for column with keywords: {keywords}")
        
        columns_found = []
        for col in df.columns:
            if all(keyword.lower() in str(col).lower() for keyword in keywords):
                columns_found.append(col)
                print(f"Found matching column: {col}")
        
        if columns_found:
            # Return the first match
            return columns_found[0]
            
        # Try with partial matching if no exact match
        partial_matches = []
        for col in df.columns:
            col_lower = str(col).lower()
            match_score = sum(1 for kw in keywords if kw.lower() in col_lower)
            if match_score > 0:
                partial_matches.append((col, match_score))
        
        # Sort by match score
        partial_matches.sort(key=lambda x: x[1], reverse=True)
        
        if partial_matches:
            best_match = partial_matches[0][0]
            print(f"No exact match found. Using best partial match: {best_match}")
            return best_match
            
        print(f"No matching column found for keywords: {keywords}")
        return None
    
    def generate_comparison_comment(self, metric, value, comparison_value):
        """Generate a comparison comment for a metric"""
        comparison_area = comparison_value['area']
        
        if metric == "renters":
            if value < comparison_value["value"] - 1:  # 1% buffer to avoid "slightly lower" for small differences
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a lower concentration of renters than the {comparison_area} average of {comparison_value['value']}%."
            elif value > comparison_value["value"] + 1:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a higher concentration of renters than the {comparison_area} average of {comparison_value['value']}%."
            else:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a similar concentration of renters to the {comparison_area} average of {comparison_value['value']}%."
        
        elif metric == "social_housing":
            if value < comparison_value["value"] - 0.5:  # 0.5% buffer
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a lower concentration of social housing than the {comparison_area} average of {comparison_value['value']}%."
            elif value > comparison_value["value"] + 0.5:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a higher concentration of social housing than the {comparison_area} average of {comparison_value['value']}%."
            else:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has a similar concentration of social housing to the {comparison_area} average of {comparison_value['value']}%."
        
        elif metric == "median_rent":
            local_increase = self.data["median_rent"]["annual_increase"]
            if pd.isna(local_increase):
                local_increase = 0
                
            if local_increase < comparison_value["value"] - 1:  # 1% buffer
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has experienced lower median annual rental increase than {comparison_area} which experienced an annual median rental increase of {comparison_value['value']}%."
            elif local_increase > comparison_value["value"] + 1:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has experienced higher median annual rental increase than {comparison_area} which experienced an annual median rental increase of {comparison_value['value']}%."
            else:
                return f"{self.selected_geo_name} ({self.selected_geo_area}) has experienced a similar median annual rental increase to {comparison_area} which is {comparison_value['value']}%."
        
        elif metric == "vacancy_rates":
            local_change = self.data["vacancy_rates"]["annual_change"]
            if pd.isna(local_change):
                local_change = 0
                
            if local_change > 0.1:  # Improvement threshold
                return f"Vacancy rates for {self.selected_geo_name} ({self.selected_geo_area}) has improved over the past 12 months, compared to {comparison_area} which saw a change in vacancy rates of {comparison_value['value']}%."
            elif local_change < -0.1:  # Deterioration threshold
                return f"Vacancy rates for {self.selected_geo_name} ({self.selected_geo_area}) has deteriorated over the past 12 months, compared to {comparison_area} which saw a change in vacancy rates of {comparison_value['value']}%."
            else:
                return f"Vacancy rates for {self.selected_geo_name} ({self.selected_geo_area}) has remained stable over the past 12 months, compared to {comparison_area} which saw a change in vacancy rates of {comparison_value['value']}%."
        
        elif metric == "affordability":
            local_improvement = self.data["affordability"]["annual_improvement"]
            if pd.isna(local_improvement):
                local_improvement = 0
                
            local_pct = self.data["affordability"]["percentage"]
            ref_value = comparison_value["value"]
            ref_change = comparison_value["annual_change"]
            
            # Compare the actual affordability percentages
            affordability_comparison = ""
            if local_pct > ref_value + 2:  # 2% buffer
                affordability_comparison = f"less affordable than the {comparison_area} average"
            elif local_pct < ref_value - 2:
                affordability_comparison = f"more affordable than the {comparison_area} average"
            else:
                affordability_comparison = f"similar to the {comparison_area} average"
            
            # Evaluate the trend
            if local_improvement > 0.1:  # Improvement
                change_text = "improvement"
            elif local_improvement < -0.1:  # Deterioration
                change_text = "deterioration"
            else:
                change_text = "stable affordability"
            
            return (f"{self.selected_geo_name} ({self.selected_geo_area}) rental affordability is {affordability_comparison} "
                   f"({ref_value}% of income on rent), with a {change_text} of {abs(local_improvement)}% over the past year "
                   f"compared to {comparison_area}'s {abs(ref_change)}%.")
        
        return ""
    
    def create_excel_output(self):
        """Create a nicely formatted Excel output with the analysis"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.selected_geo_name} Analysis"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        metric_font = Font(bold=True, size=11)
        metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        metric_alignment = Alignment(vertical="center", wrap_text=True)
        
        value_font = Font(size=11)
        value_alignment = Alignment(vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Rental Market Analysis for {self.selected_geo_name} ({self.selected_geo_area})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers - Row 3
        headers = ["Metric", "Value", "Period", "Source", "Comment"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        
        # Add metrics data
        row = 4
        
        # Renters
        ws.cell(row=row, column=1).value = "# and % of renters"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['renters']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['renters']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['renters']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("renters", self.data['renters']['percentage'], self.data['renters']['comparison'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['renters']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Social Housing
        row += 1
        ws.cell(row=row, column=1).value = "# and % of Social Housing"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['social_housing']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['social_housing']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("social_housing", self.data['social_housing']['percentage'], self.data['social_housing']['comparison'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Median Weekly Rent
        row += 1
        ws.cell(row=row, column=1).value = "Median Weekly Rent"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"${self.data['median_rent']['value']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['median_rent']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['median_rent']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("median_rent", self.data['median_rent']['value'], self.data['median_rent']['comparison'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"Annual increase {self.data['median_rent']['annual_increase']}%"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Vacancy Rates
        row += 1
        ws.cell(row=row, column=1).value = "Vacancy Rates"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['vacancy_rates']['value']}%"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['vacancy_rates']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['vacancy_rates']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("vacancy_rates", self.data['vacancy_rates']['value'], self.data['vacancy_rates']['comparison'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        annual_change = self.data['vacancy_rates']['annual_change']
        annual_change_text = f"Annual change {annual_change}%"
        ws.cell(row=row, column=2).value = annual_change_text
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Rental Affordability
        row += 1
        ws.cell(row=row, column=1).value = "Rental affordability*"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['affordability']['percentage']}% of income on rent"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['affordability']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['affordability']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("affordability", self.data['affordability']['percentage'], self.data['affordability']['comparison'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        annual_improvement = self.data['affordability']['annual_improvement']
        improvement_text = "improvement" if annual_improvement > 0 else "deterioration"
        ws.cell(row=row, column=2).value = f"Annual {improvement_text} {abs(annual_improvement)}%"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Rental Contacts
        row += 1
        ws.cell(row=row, column=1).value = "Rental Contacts (Complaints & Enquiries)"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"Total: {self.data['rental_contacts']['total']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['rental_contacts']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['rental_contacts']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        ws.cell(row=row, column=5).value = self.data['rental_contacts']['details']
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"Enquiries: {self.data['rental_contacts']['enquiries']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"Complaints: {self.data['rental_contacts']['complaints']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Add footnote for affordability
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].value = ("* Methodology: the rental affordability is calculated by taking median weekly rental household incomes for the "
                              "geographic area and comparing that to median weekly rents for the same area. Any number high than 30% of income "
                              "on rent is considered that a household is experiencing rental stress. This metric is calculated by Fair Trading "
                              "using ABS income and indexation data as well as Core Logic rental data.")
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Set row heights
        for r in range(3, row):
            ws.row_dimensions[r].height = 30
        
        ws.row_dimensions[row].height = 45  # Footnote row
        
        # Save the workbook
        wb.save(self.output_file)

def main():
    # Set pandas to not display warnings
    pd.options.mode.chained_assignment = None
    
    try:
        analyzer = RentalDataAnalyzer()
        analyzer.create_gui()
    except Exception as e:
        import traceback
        print(f"Critical error: {str(e)}")
        print(traceback.format_exc())
        
        # Show error in GUI if possible
        try:
            import tkinter.messagebox as msgbox
            msgbox.showerror("Critical Error", f"An unexpected error occurred:\n\n{str(e)}\n\nPlease check the console for details.")
        except:
            pass

if __name__ == "__main__":
    main()