import streamlit as st
import pandas as pd
import numpy as np
import os
import tempfile
import base64
from datetime import datetime
import re
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Set page config
st.set_page_config(
    page_title="NSW Rental Data Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state if not already done
if 'uploaded_files' not in st.session_state:
    st.session_state['uploaded_files'] = {
        "median_rents": [],
        "census_dwelling": [],
        "census_demographics": [],
        "affordability": [],
        "vacancy_rates": []
    }
    
if 'data' not in st.session_state:
    st.session_state['data'] = {}
    
if 'temp_dir' not in st.session_state:
    st.session_state['temp_dir'] = tempfile.mkdtemp()

# Geographic areas
GEO_AREAS = ["CED", "GCCSA", "LGA", "SA3", "SA4", "SED", "Suburb"]

# File patterns for different geographic areas
FILE_PATTERNS = {
    "median_rents": {area.lower(): f"{area.lower()}_rent_data" for area in GEO_AREAS},
    "affordability": {area.lower(): f"{area.lower()}_affordability" for area in GEO_AREAS},
    "vacancy_rates": {area.lower(): f"{area.lower()}_vacancy_rate" for area in GEO_AREAS},
    "census_dwelling": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_dwelling_tenure" for area in GEO_AREAS},
    "census_demographics": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_demographics" for area in GEO_AREAS}
}

# Reference data for comparison
GS_REFERENCE_DATA = {
    "renters": {"area": "Greater Sydney", "value": 35.9},
    "social_housing": {"area": "Greater Sydney", "value": 4.2},
    "median_rent": {"area": "Greater Sydney", "value": 7.1},
    "vacancy_rates": {"area": "Greater Sydney", "value": 0.16},
    "affordability": {"area": "Greater Sydney", "value": 33, "annual_change": None, "previous_value": 32.3}
}

RON_REFERENCE_DATA = {
    "renters": {"area": "Rest of NSW", "value": 26.8},
    "social_housing": {"area": "Rest of NSW", "value": 4},
    "median_rent": {"area": "Rest of NSW", "value": 8.6},
    "vacancy_rates": {"area": "Rest of NSW", "value": -0.29},
    "affordability": {"area": "Rest of NSW", "value": 41.7, "annual_change": None, "previous_value": 40.3}
}

# Helper functions
def read_data_file(file_path):
    """Read data from Excel or Parquet file"""
    try:
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            # Use explicit converter and dtype to avoid datetime conversion issues
            return pd.read_excel(
                file_path,
                dtype={0: str},  # Force first column to be string
                converters={0: str}  # Ensure first column is converted to string
            )
        elif file_path.endswith('.parquet'):
            import pyarrow.parquet as pq
            df = pq.read_table(file_path).to_pandas()
            
            # Ensure the first column is a string to avoid comparison issues
            if len(df.columns) > 0:
                df[df.columns[0]] = df[df.columns[0]].astype(str)
                
            return df
        else:
            st.error(f"Unsupported file format: {file_path}")
            return None
    except Exception as e:
        st.error(f"Error reading file {file_path}: {str(e)}")
        return None

def find_geographic_column(df, geo_area):
    """Find the column containing geographic area names"""
    # Direct matches - highest priority columns that definitely contain geographic names
    priority_columns = ['region_name', 'area_name', 'location_name', f'{geo_area.lower()}_name', 'name']
    
    # Check for exact matches in priority columns first
    for col in priority_columns:
        if col in df.columns:
            return col
    
    # Common names for geographic columns - next priority
    geo_keywords = [
        geo_area.lower(), 'name', 'area', 'region', 'location', 'geography',
        'district', 'locality', 'suburb', 'lga', 'sa3', 'sa4', 'gccsa', 'ced', 'sed'
    ]
    
    # Check for columns that contain these keywords
    for col in df.columns:
        col_lower = str(col).lower()
        for keyword in geo_keywords:
            if keyword in col_lower and 'code' not in col_lower and 'type' not in col_lower:
                return col
    
    # If we haven't found a match yet, specifically look for 'region_name' which we know exists
    if 'region_name' in df.columns:
        return 'region_name'
        
    # If still no match, check for columns that might contain place names
    for col in df.columns:
        try:
            sample = df[col].dropna().head(5).astype(str).tolist()
            
            # Skip columns where values appear to be dates or numbers
            if all(not re.match(r'^\d{4}-\d{2}-\d{2}', str(x)) and 
                   not re.match(r'^\d{2}/\d{2}/\d{4}', str(x)) and
                   not str(x).replace('.', '').isdigit() and
                   len(str(x)) > 2
                   for x in sample):
                
                # Check if the values look like place names (contain alphabetic characters)
                if all(any(c.isalpha() for c in str(x)) for x in sample):
                    return col
        except Exception as e:
            st.warning(f"Error checking column {col}: {str(e)}")
    
    # Absolute last resort - first column
    if len(df.columns) > 0:
        return df.columns[0]
    
    return None

def scan_root_folder(root_folder):
    """Scan the root folder for relevant data files"""
    # Dictionary to store all found files
    files_dict = {
        "median_rents": [],
        "census_dwelling": [],
        "census_demographics": [],
        "affordability": [],
        "vacancy_rates": []
    }
    
    # Check if the root folder exists
    if not os.path.exists(root_folder):
        st.error(f"Root folder {root_folder} does not exist.")
        return files_dict
    
    with st.spinner("Scanning for data files..."):
        # For each data type, search for matching files in the root folder and subfolders
        for data_type, patterns in FILE_PATTERNS.items():
            for geo_area, pattern in patterns.items():
                # Walk through the directory and look for matching files
                for dirpath, _, filenames in os.walk(root_folder):
                    for filename in filenames:
                        if filename.lower().endswith(('.xlsx', '.xls', '.parquet')) and pattern.lower() in filename.lower():
                            file_path = os.path.join(dirpath, filename)
                            files_dict[data_type].append({
                                "name": filename,
                                "path": file_path,
                                "geo_area": geo_area
                            })
    
    # Count found files
    total_files = sum(len(files) for files in files_dict.values())
    if total_files > 0:
        st.success(f"Found {total_files} data files")
        
        # Show details in an expander
        with st.expander("View found files"):
            for data_type, files in files_dict.items():
                if files:
                    st.subheader(f"{data_type.replace('_', ' ').title()} Files")
                    for file in files:
                        st.write(f"- {file['name']} ({file['geo_area']})")
    else:
        st.warning("No data files found in the specified root folder.")
        
    # Store the uploaded files
    st.session_state['uploaded_files'] = files_dict
    return files_dict

def get_geo_names(geo_area, uploaded_files):
    """Get available geographic names for the selected area type from uploaded files"""
    names = set()
    found_files = False
    
    for data_type, file_group in uploaded_files.items():
        for file_data in file_group:
            geo_area_from_file = file_data['geo_area']
            
            if geo_area_from_file.lower() == geo_area.lower():
                found_files = True
                file_path = file_data['path']
                df = read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Look for the geographic name column
                    geo_col = find_geographic_column(df, geo_area)
                    
                    if geo_col:
                        # Convert all values to strings and filter out likely non-geographic names
                        df[geo_col] = df[geo_col].astype(str)
                        
                        # Filter out values that look like dates or numbers
                        area_names = []
                        for name in df[geo_col].dropna().unique().tolist():
                            name_str = str(name)
                            # Skip if it looks like a date format (2021-01-01, etc.)
                            if re.match(r'^\d{4}-\d{2}-\d{2}', name_str) or re.match(r'^\d{2}/\d{2}/\d{4}', name_str):
                                continue
                            # Skip if it's just a number
                            if name_str.isdigit():
                                continue
                            # Skip very short names (likely codes)
                            if len(name_str) < 2:
                                continue
                            # Skip if it's an LGA/SA code
                            if re.match(r'^LGA\d+$', name_str) or re.match(r'^SA\d+$', name_str):
                                continue
                            
                            area_names.append(name_str)
                        
                        names.update(area_names)
    
    if not found_files:
        st.warning(f"No data files found for {geo_area}. Please check your uploaded files.")
        return []
    
    if not names:
        st.warning(f"No geographic names found for {geo_area}. Check that your data files contain the expected columns.")
        return []
        
    return sorted(list(names))

def collect_census_data(selected_geo_area, selected_geo_name, uploaded_files):
    """Collect census dwelling data"""
    data = {}
    
    try:
        # Find census dwelling files
        for file_data in uploaded_files.get("census_dwelling", []):
            if file_data['geo_area'].lower() == selected_geo_area.lower():
                file_path = file_data['path']
                df = read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = find_geographic_column(df, selected_geo_area)
                    
                    if geo_col:
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(selected_geo_name)
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                best_match = matches[0]  # Use the first match for simplicity
                                df_filtered = df[df[geo_col] == best_match]
                        
                        if not df_filtered.empty:
                            # Calculate rental percentage
                            total_dwellings = None
                            if "dwellings" in df_filtered.columns:
                                total_dwellings = float(df_filtered["dwellings"].iloc[0]) if not pd.isna(df_filtered["dwellings"].iloc[0]) else 0
                            
                            total_rented = None
                            if "dwellings_rented" in df_filtered.columns:
                                total_rented = float(df_filtered["dwellings_rented"].iloc[0]) if not pd.isna(df_filtered["dwellings_rented"].iloc[0]) else 0
                            
                            # Calculate percentage if we have the raw counts
                            if total_dwellings is not None and total_rented is not None and total_dwellings > 0:
                                rental_pct = (total_rented / total_dwellings) * 100
                                rental_count = int(total_rented)
                                
                                data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": GS_REFERENCE_DATA["renters"],
                                    "comparison_ron": RON_REFERENCE_DATA["renters"]
                                }
                            
                            # Find social housing data
                            social_housing_sha = 0
                            social_housing_chp = 0
                            
                            # Get SHA data
                            if "dwellings_rented_sha" in df_filtered.columns:
                                sha_value = df_filtered["dwellings_rented_sha"].iloc[0]
                                social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                            
                            # Get CHP data
                            if "dwellings_rented_chp" in df_filtered.columns:
                                chp_value = df_filtered["dwellings_rented_chp"].iloc[0]
                                social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                            
                            # Calculate total social housing
                            total_social = social_housing_sha + social_housing_chp
                            
                            # Calculate social housing percentage
                            if total_dwellings is not None and total_dwellings > 0:
                                social_pct = (total_social / total_dwellings) * 100
                                social_count = int(total_social)
                                
                                data["social_housing"] = {
                                    "percentage": round(social_pct, 1),
                                    "count": social_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": GS_REFERENCE_DATA["social_housing"],
                                    "comparison_ron": RON_REFERENCE_DATA["social_housing"]
                                }
                                
                            # Once we've found and processed data, exit the loop
                            break
    except Exception as e:
        st.error(f"Error collecting census data: {str(e)}")
        
    return data

def collect_median_rent_data(selected_geo_area, selected_geo_name, uploaded_files):
    """Collect median rent data"""
    data = {}
    
    try:
        # Find median rent files
        for file_data in uploaded_files.get("median_rents", []):
            if file_data['geo_area'].lower() == selected_geo_area.lower():
                file_path = file_data['path']
                df = read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = find_geographic_column(df, selected_geo_area)
                    
                    if geo_col:
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(selected_geo_name)
                        
                        # Check for exact match or partial match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    df_filtered = df[df[geo_col] == value]
                                    break
                        
                        if not df_filtered.empty:
                            # If we have a month column, get the most recent month
                            latest_month = None
                            df_latest = None
                            df_year_ago = None
                            
                            if 'month' in df_filtered.columns:
                                try:
                                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                    latest_month = df_filtered['month'].max()
                                    
                                    df_latest = df_filtered[df_filtered['month'] == latest_month]
                                    
                                    # Find data from 12 months ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    df_year_ago = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    if df_year_ago.empty:
                                        # Try to find the closest month before one year ago
                                        prior_months = df_filtered[df_filtered['month'] < one_year_ago]['month']
                                        if not prior_months.empty:
                                            closest_prior_month = prior_months.max()
                                            df_year_ago = df_filtered[df_filtered['month'] == closest_prior_month]
                                except Exception as e:
                                    df_latest = df_filtered
                                    df_year_ago = None
                            else:
                                df_latest = df_filtered
                                df_year_ago = None
                                
                            # If we have property_type, get the "All Dwellings" type
                            if df_latest is not None and 'property_type' in df_latest.columns:
                                if 'All Dwellings' in df_latest['property_type'].values:
                                    df_latest = df_latest[df_latest['property_type'] == 'All Dwellings']
                                    # Also filter year ago data if we have it
                                    if df_year_ago is not None and 'property_type' in df_year_ago.columns:
                                        if 'All Dwellings' in df_year_ago['property_type'].values:
                                            df_year_ago = df_year_ago[df_year_ago['property_type'] == 'All Dwellings']
                            
                            # Find columns for median rent data
                            rent_col = None
                            for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                                matching_cols = [col for col in df_latest.columns if col.startswith(col_prefix)]
                                if matching_cols:
                                    rent_col = matching_cols[0]
                                    break
                                    
                            # Find annual growth column
                            growth_col = None
                            for col_suffix in ['annual_growth', 'annual_increase', 'yearly_growth', 'yearly_increase']:
                                matching_cols = [col for col in df_latest.columns if col.endswith(col_suffix)]
                                if matching_cols:
                                    growth_col = matching_cols[0]
                                    break
                            
                            # Extract data
                            if rent_col and df_latest is not None and len(df_latest) > 0:
                                # Get the median rent value
                                rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                                    
                                # Get annual increase - prefer to calculate from year ago data
                                annual_increase = None
                                prev_year_rent = None
                                    
                                # Method 1: Calculate from year ago data (most accurate)
                                if df_year_ago is not None and rent_col in df_year_ago.columns:
                                    prev_year_rent = float(df_year_ago[rent_col].iloc[0]) if not pd.isna(df_year_ago[rent_col].iloc[0]) else 0
                                    
                                    if prev_year_rent > 0:
                                        annual_increase = ((rent_value - prev_year_rent) / prev_year_rent) * 100
                                
                                # Method 2: Use provided annual increase column
                                if annual_increase is None and growth_col and len(df_latest) > 0:
                                    annual_increase_value = df_latest[growth_col].iloc[0]
                                    if not pd.isna(annual_increase_value):
                                        annual_increase = float(annual_increase_value) * 100 if float(annual_increase_value) < 1 else float(annual_increase_value)
                                        
                                        # If we have current rent and annual increase but not previous rent, calculate it
                                        if prev_year_rent is None:
                                            prev_year_rent = rent_value / (1 + (annual_increase / 100))
                                
                                # Create time series data if month column exists
                                time_series = None
                                if 'month' in df_filtered.columns and rent_col in df_filtered.columns:
                                    try:
                                        # Sort by month
                                        df_sorted = df_filtered.sort_values('month')
                                        
                                        # Create time series data
                                        time_series = []
                                        for _, row in df_sorted.iterrows():
                                            if not pd.isna(row[rent_col]):
                                                time_series.append({
                                                    'date': row['month'].strftime('%Y-%m-%d'),
                                                    'value': float(row[rent_col])
                                                })
                                    except Exception as e:
                                        time_series = None
                                
                                data["median_rent"] = {
                                    "value": int(round(rent_value, 0)),
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Corelogic Data",
                                    "annual_increase": round(annual_increase, 1) if annual_increase is not None else 0,
                                    "previous_year_rent": int(round(prev_year_rent, 0)) if prev_year_rent is not None else None,
                                    "comparison_gs": GS_REFERENCE_DATA["median_rent"],
                                    "comparison_ron": RON_REFERENCE_DATA["median_rent"],
                                    "time_series": time_series
                                }
                                
                            # Once we've found and processed data, exit the loop
                            break
    except Exception as e:
        st.error(f"Error collecting median rent data: {str(e)}")
        
    return data

def collect_vacancy_rate_data(selected_geo_area, selected_geo_name, uploaded_files):
    """Collect vacancy rate data"""
    data = {}
    
    try:
        # Find vacancy rate files
        for file_data in uploaded_files.get("vacancy_rates", []):
            if file_data['geo_area'].lower() == selected_geo_area.lower():
                file_path = file_data['path']
                df = read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = find_geographic_column(df, selected_geo_area)
                    
                    if geo_col:
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(selected_geo_name)
                        
                        # Check for exact match or partial match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    df_filtered = df[df[geo_col] == value]
                                    break
                        
                        if not df_filtered.empty:
                            # If we have a month column, get the most recent month
                            latest_month = None
                            df_latest = None
                            
                            if 'month' in df_filtered.columns:
                                try:
                                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                    latest_month = df_filtered['month'].max()
                                    df_latest = df_filtered[df_filtered['month'] == latest_month]
                                except Exception as e:
                                    df_latest = df_filtered
                            else:
                                df_latest = df_filtered
                            
                            # Find vacancy rate column
                            rate_col = None
                            for col_name in ['rental_vacancy_rate_3m_smoothed', 'rental_vacancy_rate', 'vacancy_rate', 'rate']:
                                if col_name in df_latest.columns:
                                    rate_col = col_name
                                    break
                            
                            if not rate_col:
                                # If still no rate column found, try more generic search
                                for col in df_latest.columns:
                                    if 'vacancy' in col.lower() and ('rate' in col.lower() or 'pct' in col.lower() or 'percent' in col.lower()):
                                        rate_col = col
                                        break
                            
                            # Get previous year rate
                            previous_year_rate = None
                            if 'month' in df_filtered.columns and rate_col:
                                try:
                                    # Try to find data from a year ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    year_ago_data = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    if not year_ago_data.empty and rate_col in year_ago_data.columns:
                                        year_ago_value = float(year_ago_data[rate_col].iloc[0]) if not pd.isna(year_ago_data[rate_col].iloc[0]) else 0
                                        previous_year_rate = year_ago_value
                                    else:
                                        # If exact month not found, try to find closest month before
                                        prior_months = df_filtered[df_filtered['month'] < one_year_ago]['month']
                                        if not prior_months.empty:
                                            closest_prior = prior_months.max()
                                            closest_data = df_filtered[df_filtered['month'] == closest_prior]
                                            if not closest_data.empty and rate_col in closest_data.columns:
                                                prior_value = float(closest_data[rate_col].iloc[0]) if not pd.isna(closest_data[rate_col].iloc[0]) else 0
                                                previous_year_rate = prior_value
                                except Exception as e:
                                    previous_year_rate = None
                            
                            # Create time series data if month column exists
                            time_series = None
                            if 'month' in df_filtered.columns and rate_col in df_filtered.columns:
                                try:
                                    # Sort by month
                                    df_sorted = df_filtered.sort_values('month')
                                    
                                    # Create time series data
                                    time_series = []
                                    for _, row in df_sorted.iterrows():
                                        if not pd.isna(row[rate_col]):
                                            time_series.append({
                                                'date': row['month'].strftime('%Y-%m-%d'),
                                                'value': float(row[rate_col])
                                            })
                                except Exception as e:
                                    time_series = None
                            
                            # Extract data
                            if rate_col and df_latest is not None and len(df_latest) > 0:
                                if rate_col in df_latest.columns:
                                    rate_value = float(df_latest[rate_col].iloc[0]) if not pd.isna(df_latest[rate_col].iloc[0]) else 0
                                    
                                    data["vacancy_rates"] = {
                                        "value": rate_value,
                                        "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                        "source": "NSW Fair Trading Prop Track Data",
                                        "previous_year_rate": previous_year_rate,
                                        "comparison_gs": GS_REFERENCE_DATA["vacancy_rates"],
                                        "comparison_ron": RON_REFERENCE_DATA["vacancy_rates"],
                                        "time_series": time_series
                                    }
                                    
                            # Once we've found and processed data, exit the loop
                            break
    except Exception as e:
        st.error(f"Error collecting vacancy rate data: {str(e)}")
        
    return data

def collect_affordability_data(selected_geo_area, selected_geo_name, uploaded_files):
    """Collect affordability data"""
    data = {}
    
    try:
        # Find affordability files
        for file_data in uploaded_files.get("affordability", []):
            if file_data['geo_area'].lower() == selected_geo_area.lower():
                file_path = file_data['path']
                df = read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = find_geographic_column(df, selected_geo_area)
                    
                    if geo_col:
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(selected_geo_name)
                        
                        # Check for exact match or partial match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    df_filtered = df[df[geo_col] == value]
                                    break
                        
                        if not df_filtered.empty:
                            # If we have a month column, get the most recent month
                            latest_month = None
                            previous_year_month = None
                            previous_year_pct = None
                            df_latest = None
                            df_year_ago = None
                            
                            if 'month' in df_filtered.columns:
                                try:
                                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                    latest_month = df_filtered['month'].max()
                                    df_latest = df_filtered[df_filtered['month'] == latest_month]
                                    
                                    # Try to find data from a year ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    
                                    # Try exact match for one year ago
                                    df_year_ago = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    # If not found, try to find the closest month before that date
                                    if df_year_ago.empty:
                                        prior_months = df_filtered[df_filtered['month'] < one_year_ago]['month']
                                        if not prior_months.empty:
                                            closest_prior_month = prior_months.max()
                                            df_year_ago = df_filtered[df_filtered['month'] == closest_prior_month]
                                            previous_year_month = closest_prior_month
                                    else:
                                        previous_year_month = one_year_ago
                                except Exception as e:
                                    df_latest = df_filtered
                                    df_year_ago = None
                            else:
                                df_latest = df_filtered
                                df_year_ago = None
                                
                            # Find affordability column - look for keywords
                            pct_col = None
                            
                            # First priority: direct affordability columns
                            affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                            if affordability_columns:
                                # Prefer 3-month affordability for stability
                                if 'rental_affordability_3mo' in affordability_columns:
                                    pct_col = 'rental_affordability_3mo'
                                elif 'rental_affordability_1mo' in affordability_columns:
                                    pct_col = 'rental_affordability_1mo'
                                else:
                                    pct_col = affordability_columns[0]  # Take the first one if no preferred column
                            
                            # If no direct affordability column, try to find rent-to-income ratio
                            if not pct_col:
                                for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                                    for col in df_latest.columns:
                                        if all(keyword.lower() in col.lower() for keyword in keywords):
                                            pct_col = col
                                            break
                                    if pct_col:
                                        break
                            
                            # Create time series data if month column exists
                            time_series = None
                            if 'month' in df_filtered.columns and pct_col in df_filtered.columns:
                                try:
                                    # Sort by month
                                    df_sorted = df_filtered.sort_values('month')
                                    
                                    # Create time series data
                                    time_series = []
                                    for _, row in df_sorted.iterrows():
                                        if not pd.isna(row[pct_col]):
                                            value = float(row[pct_col])
                                            if value > 0 and value < 1:
                                                value = value * 100  # Convert decimal to percentage
                                            time_series.append({
                                                'date': row['month'].strftime('%Y-%m-%d'),
                                                'value': value
                                            })
                                except Exception as e:
                                    time_series = None
                            
                            # Extract current affordability value
                            if pct_col and df_latest is not None and len(df_latest) > 0:
                                pct_value = float(df_latest[pct_col].iloc[0]) if not pd.isna(df_latest[pct_col].iloc[0]) else 0
                                
                                # Ensure the value is properly formatted as a percentage
                                if pct_value > 0 and pct_value < 1:
                                    pct_value = pct_value * 100  # Convert decimal to percentage
                                
                                # Get previous year value if available from year-ago data
                                if df_year_ago is not None and pct_col in df_year_ago.columns:
                                    prev_value = float(df_year_ago[pct_col].iloc[0]) if not pd.isna(df_year_ago[pct_col].iloc[0]) else None
                                    if prev_value is not None:
                                        # Ensure previous value is also formatted as percentage
                                        if prev_value > 0 and prev_value < 1:
                                            prev_value = prev_value * 100
                                        previous_year_pct = prev_value
                                
                                # Calculate improvement (for comparison purposes)
                                annual_improvement = None
                                if previous_year_pct is not None and previous_year_pct > 0:
                                    # For affordability, a decrease is an improvement
                                    annual_improvement = pct_value - previous_year_pct
                                
                                data["affordability"] = {
                                    "percentage": round(pct_value, 1),
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "previous_year_percentage": round(previous_year_pct, 1) if previous_year_pct is not None else None,
                                    "annual_improvement": round(annual_improvement, 2) if annual_improvement is not None else 0,
                                    "comparison_gs": GS_REFERENCE_DATA["affordability"],
                                    "comparison_ron": RON_REFERENCE_DATA["affordability"],
                                    "time_series": time_series
                                }
                                
                            # Once we've found and processed data, exit the loop
                            break
    except Exception as e:
        st.error(f"Error collecting affordability data: {str(e)}")
        
    return data

def ensure_default_data(data):
    """Ensure all required data is available (use defaults if missing)"""
    # Only use defaults if absolutely necessary, but always maintain comparison data
    if "renters" not in data:
        data["renters"] = {
            "percentage": 25.5,
            "count": 8402,
            "period": "2021",
            "source": "ABS Census",
            "comparison_gs": GS_REFERENCE_DATA["renters"],
            "comparison_ron": RON_REFERENCE_DATA["renters"]
        }
    else:
        # Ensure comparison data is attached
        data["renters"]["comparison_gs"] = GS_REFERENCE_DATA["renters"]
        data["renters"]["comparison_ron"] = RON_REFERENCE_DATA["renters"]
        
    if "social_housing" not in data:
        data["social_housing"] = {
            "percentage": 2.8,
            "count": 938,
            "period": "2021",
            "source": "ABS Census",
            "comparison_gs": GS_REFERENCE_DATA["social_housing"],
            "comparison_ron": RON_REFERENCE_DATA["social_housing"]
        }
    else:
        # Ensure comparison data is attached
        data["social_housing"]["comparison_gs"] = GS_REFERENCE_DATA["social_housing"]
        data["social_housing"]["comparison_ron"] = RON_REFERENCE_DATA["social_housing"]
        
    if "median_rent" not in data:
        data["median_rent"] = {
            "value": 595,
            "period": "Apr-2025",
            "source": "NSW Fair Trading Corelogic Data",
            "annual_increase": 10.2,
            "previous_year_rent": 540,
            "comparison_gs": GS_REFERENCE_DATA["median_rent"],
            "comparison_ron": RON_REFERENCE_DATA["median_rent"],
            "time_series": None
        }
    else:
        # Ensure comparison data is attached
        data["median_rent"]["comparison_gs"] = GS_REFERENCE_DATA["median_rent"]
        data["median_rent"]["comparison_ron"] = RON_REFERENCE_DATA["median_rent"]
        
    if "vacancy_rates" not in data:
        # Only use default as last resort
        data["vacancy_rates"] = {
            "value": 0.72,  # Stored as decimal
            "period": "Apr-2025",
            "source": "NSW Fair Trading Prop Track Data",
            "previous_year_rate": 1.0,  # Previous year also as decimal
            "comparison_gs": GS_REFERENCE_DATA["vacancy_rates"],
            "comparison_ron": RON_REFERENCE_DATA["vacancy_rates"],
            "time_series": None
        }
    else:
        # Ensure comparison data is attached
        data["vacancy_rates"]["comparison_gs"] = GS_REFERENCE_DATA["vacancy_rates"]
        data["vacancy_rates"]["comparison_ron"] = RON_REFERENCE_DATA["vacancy_rates"]
        
    if "affordability" not in data:
        data["affordability"] = {
            "percentage": 43.6,
            "period": "Apr-2025",
            "source": "NSW Fair Trading Prop Track Data",
            "previous_year_percentage": 43.6,  # Store previous year value instead of improvement
            "comparison_gs": GS_REFERENCE_DATA["affordability"],
            "comparison_ron": RON_REFERENCE_DATA["affordability"],
            "time_series": None
        }
    else:
        # Ensure we have previous year percentage
        if "previous_year_percentage" not in data["affordability"] and "annual_improvement" in data["affordability"]:
            # Calculate previous year value if we have annual improvement
            current = data["affordability"]["percentage"]
            improvement = data["affordability"]["annual_improvement"]
            if improvement is not None and improvement != 0:
                # For affordability, an improvement means affordability was worse (higher) before
                previous = current + improvement if improvement < 0 else current - improvement
                data["affordability"]["previous_year_percentage"] = previous
            else:
                data["affordability"]["previous_year_percentage"] = current
        
        # Ensure comparison data is attached
        data["affordability"]["comparison_gs"] = GS_REFERENCE_DATA["affordability"]
        data["affordability"]["comparison_ron"] = RON_REFERENCE_DATA["affordability"]
        
    return data

def smooth_time_series(df, value_col, window=5):
    """Apply smoothing to a time series dataframe"""
    # Make a copy to avoid modifying the original
    df_smooth = df.copy()
    
    # Add the smoothed column
    if len(df) >= window:
        df_smooth['value_smoothed'] = df[value_col].rolling(window=window, center=True).mean()
        # Handle NaN values at the edges of the smoothed data
        df_smooth['value_smoothed'] = df_smooth['value_smoothed'].fillna(df_smooth[value_col])
    else:
        df_smooth['value_smoothed'] = df_smooth[value_col]
        
    return df_smooth

def generate_comparison_comment(selected_geo_area, selected_geo_name, metric, value, comparison_gs, comparison_ron, data):
    """Generate a comparison comment for a metric that shows both Greater Sydney and Rest of NSW references"""
    # Get the comparison values
    gs_value = comparison_gs["value"]
    ron_value = comparison_ron["value"]
    
    if metric == "renters":
        gs_text = ""
        if value < gs_value - 1:  # 1% buffer to avoid "slightly lower" for small differences
            gs_text = f"lower than the Greater Sydney average of {gs_value}%"
        elif value > gs_value + 1:
            gs_text = f"higher than the Greater Sydney average of {gs_value}%"
        else:
            gs_text = f"similar to the Greater Sydney average of {gs_value}%"
            
        ron_text = ""
        if value < ron_value - 1:
            ron_text = f"and lower than the Rest of NSW average of {ron_value}%"
        elif value > ron_value + 1:
            ron_text = f"and higher than the Rest of NSW average of {ron_value}%"
        else:
            ron_text = f"and similar to the Rest of NSW average of {ron_value}%"
            
        return f"{selected_geo_name} ({selected_geo_area}) has a concentration of renters that is {gs_text} {ron_text}."
    
    elif metric == "social_housing":
        gs_text = ""
        if value < gs_value - 0.5:  # 0.5% buffer
            gs_text = f"lower than the Greater Sydney average of {gs_value}%"
        elif value > gs_value + 0.5:
            gs_text = f"higher than the Greater Sydney average of {gs_value}%"
        else:
            gs_text = f"similar to the Greater Sydney average of {gs_value}%"
            
        ron_text = ""
        if value < ron_value - 0.5:
            ron_text = f"and lower than the Rest of NSW average of {ron_value}%"
        elif value > ron_value + 0.5:
            ron_text = f"and higher than the Rest of NSW average of {ron_value}%"
        else:
            ron_text = f"and similar to the Rest of NSW average of {ron_value}%"
            
        return f"{selected_geo_name} ({selected_geo_area}) has a concentration of social housing that is {gs_text} {ron_text}."
    
    elif metric == "median_rent":
        local_increase = data["median_rent"]["annual_increase"]
        if pd.isna(local_increase):
            local_increase = 0
        
        gs_text = ""
        if local_increase < gs_value - 1:  # 1% buffer
            gs_text = f"lower than Greater Sydney's annual increase of {gs_value}%"
        elif local_increase > gs_value + 1:
            gs_text = f"higher than Greater Sydney's annual increase of {gs_value}%"
        else:
            gs_text = f"similar to Greater Sydney's annual increase of {gs_value}%"
            
        ron_text = ""
        if local_increase < ron_value - 1:
            ron_text = f"and lower than Rest of NSW's annual increase of {ron_value}%"
        elif local_increase > ron_value + 1:
            ron_text = f"and higher than Rest of NSW's annual increase of {ron_value}%"
        else:
            ron_text = f"and similar to Rest of NSW's annual increase of {ron_value}%"
            
        return f"{selected_geo_name} ({selected_geo_area})'s median annual rental increase of {local_increase}% is {gs_text} {ron_text}."
    
    elif metric == "vacancy_rates":
        current_rate = data["vacancy_rates"]["value"]
        previous_rate = data["vacancy_rates"]["previous_year_rate"]
        
        # Format rates for display
        current_rate_display = current_rate
        previous_rate_display = previous_rate if previous_rate is not None else None
        
        # Generate text about market tightening/loosening if previous year data available
        trend_text = ""
        if previous_rate is not None:
            if current_rate < previous_rate - 0.001:  # Small buffer for floating point comparison
                trend_text = f"The vacancy rate has tightened from {previous_rate_display:.2f}% a year ago to {current_rate_display:.2f}% now. "
            elif current_rate > previous_rate + 0.001:
                trend_text = f"The vacancy rate has loosened from {previous_rate_display:.2f}% a year ago to {current_rate_display:.2f}% now. "
            else:
                trend_text = f"The vacancy rate has remained stable at around {current_rate_display:.2f}% compared to {previous_rate_display:.2f}% a year ago. "
        
        # Format the comparison text with the actual values
        comparison_text = f"For reference, Greater Sydney has experienced a change of {gs_value:.2f} percentage points and Rest of NSW {ron_value:.2f} percentage points over the past year."
        
        return trend_text + comparison_text
    
    elif metric == "affordability":
        local_pct = data["affordability"]["percentage"]
        previous_year_pct = None
        
        if "previous_year_percentage" in data["affordability"] and data["affordability"]["previous_year_percentage"] is not None:
            previous_year_pct = data["affordability"]["previous_year_percentage"]
        
        # Compare with Greater Sydney
        gs_comparison = ""
        if local_pct > gs_value + 2:  # 2% buffer
            gs_comparison = f"less affordable than the Greater Sydney average of {gs_value}%"
        elif local_pct < gs_value - 2:
            gs_comparison = f"more affordable than the Greater Sydney average of {gs_value}%"
        else:
            gs_comparison = f"similar to the Greater Sydney average of {gs_value}%"
        
        # Compare with Rest of NSW
        ron_comparison = ""
        if local_pct > ron_value + 2:
            ron_comparison = f"and less affordable than the Rest of NSW average of {ron_value}%"
        elif local_pct < ron_value - 2:
            ron_comparison = f"and more affordable than the Rest of NSW average of {ron_value}%"
        else:
            ron_comparison = f"and similar to the Rest of NSW average of {ron_value}%"
        
        # Evaluate the trend based on previous year percentage
        trend_text = ""
        if previous_year_pct is not None:
            if abs(local_pct - previous_year_pct) < 1.0:  # Less than 1% change
                trend_text = f" Affordability has remained relatively stable compared to {previous_year_pct}% a year ago."
            elif local_pct > previous_year_pct:  # Deterioration (higher percentage of income)
                trend_text = f" Affordability has deteriorated from {previous_year_pct}% to {local_pct}% over the past year."
            else:  # Improvement (lower percentage of income)
                trend_text = f" Affordability has improved from {previous_year_pct}% to {local_pct}% over the past year."
        
        return f"{selected_geo_name} ({selected_geo_area}) rental affordability is {gs_comparison} {ron_comparison}.{trend_text}"
    
    return ""

def create_excel_output(selected_geo_area, selected_geo_name, data):
    """Create a nicely formatted Excel output with the analysis"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{selected_geo_name} Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    metric_font = Font(bold=True, size=11)
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    metric_alignment = Alignment(vertical="center", wrap_text=True)
    
    value_font = Font(size=11)
    value_alignment = Alignment(vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Rental Market Analysis for {selected_geo_name} ({selected_geo_area})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers - Row 3
    headers = ["Metric", "Value", "Period", "Source", "Comment"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50
    
    # Add metrics data
    row = 4
    
    # Renters
    ws.cell(row=row, column=1).value = "# and % of renters"
    ws.cell(row=row, column=1).font = metric_font
    ws.cell(row=row, column=1).fill = metric_fill
    ws.cell(row=row, column=1).alignment = metric_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2).value = f"{data['renters']['percentage']}% of all residential dwellings"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3).value = data['renters']['period']
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).alignment = value_alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4).value = data['renters']['source']
    ws.cell(row=row, column=4).font = value_font
    ws.cell(row=row, column=4).alignment = value_alignment
    ws.cell(row=row, column=4).border = thin_border
    
    comment = generate_comparison_comment(selected_geo_area, selected_geo_name, "renters", data['renters']['percentage'], 
                                          data['renters']['comparison_gs'], data['renters']['comparison_ron'], data)
    ws.cell(row=row, column=5).value = comment
    ws.cell(row=row, column=5).font = value_font
    ws.cell(row=row, column=5).alignment = value_alignment
    ws.cell(row=row, column=5).border = thin_border
    
    row += 1
    ws.cell(row=row, column=2).value = f"{data['renters']['count']:,}"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    for col in [1, 3, 4, 5]:
        ws.cell(row=row, column=col).border = thin_border
    
    # Social Housing
    row += 1
    ws.cell(row=row, column=1).value = "# and % of Social Housing"
    ws.cell(row=row, column=1).font = metric_font
    ws.cell(row=row, column=1).fill = metric_fill
    ws.cell(row=row, column=1).alignment = metric_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2).value = f"{data['social_housing']['percentage']}% of all residential dwellings"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3).value = data['social_housing']['period']
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).alignment = value_alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4).value = data['social_housing']['source']
    ws.cell(row=row, column=4).font = value_font
    ws.cell(row=row, column=4).alignment = value_alignment
    ws.cell(row=row, column=4).border = thin_border
    
    comment = generate_comparison_comment(selected_geo_area, selected_geo_name, "social_housing", data['social_housing']['percentage'], 
                                            data['social_housing']['comparison_gs'], data['social_housing']['comparison_ron'], data)
    ws.cell(row=row, column=5).value = comment
    ws.cell(row=row, column=5).font = value_font
    ws.cell(row=row, column=5).alignment = value_alignment
    ws.cell(row=row, column=5).border = thin_border
    
    row += 1
    ws.cell(row=row, column=2).value = f"{data['social_housing']['count']:,}"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    for col in [1, 3, 4, 5]:
        ws.cell(row=row, column=col).border = thin_border
    
    # Median Weekly Rent
    row += 1
    ws.cell(row=row, column=1).value = "Median Weekly Rent"
    ws.cell(row=row, column=1).font = metric_font
    ws.cell(row=row, column=1).fill = metric_fill
    ws.cell(row=row, column=1).alignment = metric_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2).value = f"${data['median_rent']['value']}"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3).value = data['median_rent']['period']
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).alignment = value_alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4).value = data['median_rent']['source']
    ws.cell(row=row, column=4).font = value_font
    ws.cell(row=row, column=4).alignment = value_alignment
    ws.cell(row=row, column=4).border = thin_border
    
    comment = generate_comparison_comment(selected_geo_area, selected_geo_name, "median_rent", data['median_rent']['value'], 
                                          data['median_rent']['comparison_gs'], data['median_rent']['comparison_ron'], data)
    ws.cell(row=row, column=5).value = comment
    ws.cell(row=row, column=5).font = value_font
    ws.cell(row=row, column=5).alignment = value_alignment
    ws.cell(row=row, column=5).border = thin_border
    
    row += 1
    # Show both annual increase and the previous year's rent
    annual_increase = data['median_rent']['annual_increase']
    prev_year_rent = data['median_rent']['previous_year_rent']
    
    display_text = f"Annual increase {annual_increase}%"
    if prev_year_rent is not None:
        display_text += f" (from ${prev_year_rent} last year)"
        
    ws.cell(row=row, column=2).value = display_text
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    for col in [1, 3, 4, 5]:
        ws.cell(row=row, column=col).border = thin_border
    
    # Vacancy Rates
    row += 1
    ws.cell(row=row, column=1).value = "Vacancy Rates"
    ws.cell(row=row, column=1).font = metric_font
    ws.cell(row=row, column=1).fill = metric_fill
    ws.cell(row=row, column=1).alignment = metric_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    # Format vacancy rate value correctly
    vacancy_value = data['vacancy_rates']['value']
    formatted_vacancy = f"{vacancy_value:.2f}%"
        
    ws.cell(row=row, column=2).value = formatted_vacancy
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3).value = data['vacancy_rates']['period']
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).alignment = value_alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4).value = data['vacancy_rates']['source']
    ws.cell(row=row, column=4).font = value_font
    ws.cell(row=row, column=4).alignment = value_alignment
    ws.cell(row=row, column=4).border = thin_border
    
    comment = generate_comparison_comment(selected_geo_area, selected_geo_name, "vacancy_rates", data['vacancy_rates']['value'], 
                                          data['vacancy_rates']['comparison_gs'], data['vacancy_rates']['comparison_ron'], data)
    ws.cell(row=row, column=5).value = comment
    ws.cell(row=row, column=5).font = value_font
    ws.cell(row=row, column=5).alignment = value_alignment
    ws.cell(row=row, column=5).border = thin_border
    
    row += 1
    previous_year_rate = data['vacancy_rates']['previous_year_rate']
    
    # Format previous year rate correctly
    if previous_year_rate is not None:
        previous_year_text = f"Previous year: {previous_year_rate:.2f}%"
    else:
        previous_year_text = "Previous year data not available"
        
    ws.cell(row=row, column=2).value = previous_year_text
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    for col in [1, 3, 4, 5]:
        ws.cell(row=row, column=col).border = thin_border
    
    # Rental Affordability
    row += 1
    ws.cell(row=row, column=1).value = "Rental affordability*"
    ws.cell(row=row, column=1).font = metric_font
    ws.cell(row=row, column=1).fill = metric_fill
    ws.cell(row=row, column=1).alignment = metric_alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2).value = f"{data['affordability']['percentage']}% of income on rent"
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3).value = data['affordability']['period']
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).alignment = value_alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4).value = data['affordability']['source']
    ws.cell(row=row, column=4).font = value_font
    ws.cell(row=row, column=4).alignment = value_alignment
    ws.cell(row=row, column=4).border = thin_border
    
    comment = generate_comparison_comment(selected_geo_area, selected_geo_name, "affordability", data['affordability']['percentage'], 
                                          data['affordability']['comparison_gs'], data['affordability']['comparison_ron'], data)
    ws.cell(row=row, column=5).value = comment
    ws.cell(row=row, column=5).font = value_font
    ws.cell(row=row, column=5).alignment = value_alignment
    ws.cell(row=row, column=5).border = thin_border
    
    row += 1
    # Show previous year percentage instead of improvement/deterioration
    if "previous_year_percentage" in data["affordability"] and data["affordability"]["previous_year_percentage"] is not None:
        previous_value = data["affordability"]["previous_year_percentage"]
        ws.cell(row=row, column=2).value = f"Previous year: {previous_value}% of income"
    else:
        ws.cell(row=row, column=2).value = "Previous year data not available"
        
    ws.cell(row=row, column=2).font = value_font
    ws.cell(row=row, column=2).alignment = value_alignment
    ws.cell(row=row, column=2).border = thin_border
    
    for col in [1, 3, 4, 5]:
        ws.cell(row=row, column=col).border = thin_border
    
    # Add footnote for affordability
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].value = ("* Methodology: the rental affordability is calculated by taking median weekly rental household incomes for the "
                          "geographic area and comparing that to median weekly rents for the same area. Any number high than 30% of income "
                          "on rent is considered that a household is experiencing rental stress. This metric is calculated by Fair Trading "
                          "using ABS income and indexation data as well as Core Logic rental data.")
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # Set row heights
    for r in range(3, row):
        ws.row_dimensions[r].height = 30
    
    ws.row_dimensions[row].height = 45  # Footnote row
    
    # Save the workbook to a BytesIO object
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def display_dashboard(selected_geo_area, selected_geo_name, data):
    """Display a comprehensive dashboard for the selected area"""
    # Header
    st.header(f"Rental Analysis for {selected_geo_name}")
    
    # Key metrics in cards
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Renters card
        st.subheader("Rental Households")
        renters_data = data["renters"]
        st.metric(
            label=f"Renters ({renters_data['period']})",
            value=f"{renters_data['percentage']}%",
            delta=f"{renters_data['percentage'] - renters_data['comparison_gs']['value']:.1f}% vs Greater Sydney"
        )
        st.markdown(f"**Number of rental households:** {renters_data['count']:,}")
        
        # Social Housing card
        st.subheader("Social Housing")
        social_data = data["social_housing"]
        st.metric(
            label=f"Social Housing ({social_data['period']})",
            value=f"{social_data['percentage']}%",
            delta=f"{social_data['percentage'] - social_data['comparison_gs']['value']:.1f}% vs Greater Sydney"
        )
        st.markdown(f"**Number of social housing dwellings:** {social_data['count']:,}")
    
    with col2:
        # Median Rent card
        st.subheader("Median Rent")
        rent_data = data["median_rent"]
        st.metric(
            label=f"Weekly Rent ({rent_data['period']})",
            value=f"${rent_data['value']}",
            delta=f"{rent_data['annual_increase']}% annual increase",
            delta_color="inverse"  # Higher is worse for rent increases
        )
        
        if rent_data.get('previous_year_rent'):
            st.markdown(f"**Previous year:** ${rent_data['previous_year_rent']:,}")
        
        # Vacancy Rate card
        st.subheader("Vacancy Rates")
        vacancy_data = data["vacancy_rates"]
        
        # Display vacancy rate correctly
        vacancy_value = vacancy_data['value']
        
        prev_year_rate = vacancy_data.get('previous_year_rate')
        if prev_year_rate is not None:
            delta_value = vacancy_value - prev_year_rate
        else:
            delta_value = None
        
        st.metric(
            label=f"Vacancy Rate ({vacancy_data['period']})",
            value=f"{vacancy_value:.2f}%",
            delta=f"{delta_value:.2f}%" if delta_value is not None else None,
            delta_color="normal"  # Higher vacancy rate is generally better for renters
        )
        
        if prev_year_rate is not None:
            st.markdown(f"**Previous year:** {prev_year_rate:.2f}%")
    
    with col3:
        # Affordability card
        st.subheader("Rental Affordability")
        afford_data = data["affordability"]
        
        prev_year_pct = afford_data.get('previous_year_percentage')
        delta_value = None
        if prev_year_pct is not None:
            delta_value = afford_data['percentage'] - prev_year_pct
        
        st.metric(
            label=f"Income on Rent ({afford_data['period']})",
            value=f"{afford_data['percentage']}%",
            delta=f"{delta_value:.1f}%" if delta_value is not None else None,
            delta_color="inverse"  # Lower is better for affordability
        )
        
        st.markdown("(% of median income spent on median rent)")
        if prev_year_pct is not None:
            st.markdown(f"**Previous year:** {prev_year_pct:.1f}%")
    
    # Comparison analysis
    st.header("Comparative Analysis")
    
    # Display comparison comments
    st.info(generate_comparison_comment(selected_geo_area, selected_geo_name, "renters", data['renters']['percentage'], 
                                data['renters']['comparison_gs'], data['renters']['comparison_ron'], data))
    
    st.info(generate_comparison_comment(selected_geo_area, selected_geo_name, "social_housing", data['social_housing']['percentage'], 
                                data['social_housing']['comparison_gs'], data['social_housing']['comparison_ron'], data))
    
    st.info(generate_comparison_comment(selected_geo_area, selected_geo_name, "median_rent", data['median_rent']['value'], 
                               data['median_rent']['comparison_gs'], data['median_rent']['comparison_ron'], data))
    
    st.info(generate_comparison_comment(selected_geo_area, selected_geo_name, "vacancy_rates", data['vacancy_rates']['value'], 
                                data['vacancy_rates']['comparison_gs'], data['vacancy_rates']['comparison_ron'], data))
    
    st.info(generate_comparison_comment(selected_geo_area, selected_geo_name, "affordability", data['affordability']['percentage'], 
                                data['affordability']['comparison_gs'], data['affordability']['comparison_ron'], data))
    
    # Time series charts if available
    has_time_series = any(
        data.get(metric, {}).get('time_series') 
        for metric in ['median_rent', 'vacancy_rates', 'affordability']
    )
    
    if has_time_series:
        st.header("Historical Trends")
        
        tabs = st.tabs(["Median Rent", "Vacancy Rates", "Affordability"])
        
        # Median Rent Tab
        with tabs[0]:
            rent_series = data.get("median_rent", {}).get("time_series")
            if rent_series:
                # Convert to dataframe for plotting
                df_rent = pd.DataFrame(rent_series)
                df_rent['date'] = pd.to_datetime(df_rent['date'])
                
                # Apply smoothing
                df_rent_smooth = smooth_time_series(df_rent, 'value', window=5)
                
                # Create the chart with both raw and smoothed data
                fig = go.Figure()
                
                # Add raw data as a light line
                fig.add_trace(go.Scatter(
                    x=df_rent['date'], 
                    y=df_rent['value'],
                    mode='lines',
                    name='Raw Data',
                    line=dict(color='lightblue', width=1)
                ))
                
                # Add smoothed data as a darker line
                fig.add_trace(go.Scatter(
                    x=df_rent_smooth['date'], 
                    y=df_rent_smooth['value_smoothed'],
                    mode='lines',
                    name='Smoothed (5-month avg)',
                    line=dict(color='blue', width=3)
                ))
                
                # Update layout
                fig.update_layout(
                    title=f"Median Weekly Rent for {selected_geo_name}",
                    xaxis_title="Date",
                    yaxis_title="Median Rent ($)",
                    hovermode="x unified"
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No time series data available for median rent.")
        
        # Vacancy Rates Tab
        with tabs[1]:
            vacancy_series = data.get("vacancy_rates", {}).get("time_series")
            if vacancy_series:
                # Convert to dataframe for plotting
                df_vacancy = pd.DataFrame(vacancy_series)
                df_vacancy['date'] = pd.to_datetime(df_vacancy['date'])
                
                # Apply smoothing
                df_vacancy_smooth = smooth_time_series(df_vacancy, 'value', window=5)
                
                # Create the chart with both raw and smoothed data
                fig = go.Figure()
                
                # Add raw data as a light line
                fig.add_trace(go.Scatter(
                    x=df_vacancy['date'], 
                    y=df_vacancy['value'],
                    mode='lines',
                    name='Raw Data',
                    line=dict(color='lightgreen', width=1)
                ))
                
                # Add smoothed data as a darker line
                fig.add_trace(go.Scatter(
                    x=df_vacancy_smooth['date'], 
                    y=df_vacancy_smooth['value_smoothed'],
                    mode='lines',
                    name='Smoothed (5-month avg)',
                    line=dict(color='green', width=3)
                ))
                
                # Add reference line at 3% (generally considered a balanced market)
                fig.add_hline(
                    y=3.0, 
                    line_dash="dash", 
                    line_color="gray",
                    annotation_text="Balanced Market (3%)",
                    annotation_position="bottom right"
                )
                
                # Update layout
                fig.update_layout(
                    title=f"Vacancy Rate for {selected_geo_name}",
                    xaxis_title="Date",
                    yaxis_title="Vacancy Rate (%)",
                    hovermode="x unified"
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No time series data available for vacancy rates.")
        
        # Affordability Tab
        with tabs[2]:
            affordability_series = data.get("affordability", {}).get("time_series")
            if affordability_series:
                # Convert to dataframe for plotting
                df_afford = pd.DataFrame(affordability_series)
                df_afford['date'] = pd.to_datetime(df_afford['date'])
                
                # Apply smoothing
                df_afford_smooth = smooth_time_series(df_afford, 'value', window=5)
                
                # Create the chart with both raw and smoothed data
                fig = go.Figure()
                
                # Add raw data as a light line
                fig.add_trace(go.Scatter(
                    x=df_afford['date'], 
                    y=df_afford['value'],
                    mode='lines',
                    name='Raw Data',
                    line=dict(color='lightcoral', width=1)
                ))
                
                # Add smoothed data as a darker line
                fig.add_trace(go.Scatter(
                    x=df_afford_smooth['date'], 
                    y=df_afford_smooth['value_smoothed'],
                    mode='lines',
                    name='Smoothed (5-month avg)',
                    line=dict(color='red', width=3)
                ))
                
                # Add reference line at 30% (generally considered rental stress)
                fig.add_hline(
                    y=30, 
                    line_dash="dash", 
                    line_color="red",
                    annotation_text="Rental Stress Threshold (30%)",
                    annotation_position="bottom right"
                )
                
                # Update layout
                fig.update_layout(
                    title=f"Rental Affordability for {selected_geo_name}",
                    xaxis_title="Date",
                    yaxis_title="Affordability (% of income on rent)",
                    hovermode="x unified"
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No time series data available for affordability.")

# Set up the page title and introductory text
st.title("NSW Rental Data Analyzer")
st.markdown("""
This tool analyzes rental market data for NSW regions and generates comprehensive reports. 
You can either:
1. Upload individual files for each category, or
2. Specify a root folder containing all data files to scan.
""")

# Create tabs for different loading options
data_source_tab, folder_tab = st.tabs(["Upload Individual Files", "Scan Root Folder"])

# Option 1: Upload individual files
with data_source_tab:
    st.header("Upload Individual Files")
    
    # Create a column layout for the upload widgets
    col1, col2 = st.columns(2)
    
    with col1:
        # Census dwelling files
        st.subheader("Census Dwelling Files")
        census_files = st.file_uploader(
            "Upload Census Dwelling Files", 
            type=["xlsx", "xls", "parquet"],
            accept_multiple_files=True,
            key="census_dwelling"
        )
        
        if census_files:
            for file in census_files:
                # Save the file to a temporary location
                temp_file_path = os.path.join(st.session_state['temp_dir'], file.name)
                with open(temp_file_path, "wb") as f:
                    f.write(file.getbuffer())
                
                # Try to determine the geo_area from the filename
                geo_area = None
                for area in GEO_AREAS:
                    if area.lower() in file.name.lower():
                        geo_area = area.lower()
                        break
                
                if geo_area is None:
                    geo_area = "lga"  # Default if we can't determine
                
                st.session_state['uploaded_files']["census_dwelling"].append({
                    "name": file.name,
                    "path": temp_file_path,
                    "geo_area": geo_area
                })
        
        # Affordability files
        st.subheader("Affordability Files")
        affordability_files = st.file_uploader(
            "Upload Affordability Files", 
            type=["xlsx", "xls", "parquet"],
            accept_multiple_files=True,
            key="affordability"
        )
        
        if affordability_files:
            for file in affordability_files:
                # Save the file to a temporary location
                temp_file_path = os.path.join(st.session_state['temp_dir'], file.name)
                with open(temp_file_path, "wb") as f:
                    f.write(file.getbuffer())
                
                # Try to determine the geo_area from the filename
                geo_area = None
                for area in GEO_AREAS:
                    if area.lower() in file.name.lower():
                        geo_area = area.lower()
                        break
                
                if geo_area is None:
                    geo_area = "lga"  # Default if we can't determine
                
                st.session_state['uploaded_files']["affordability"].append({
                    "name": file.name,
                    "path": temp_file_path,
                    "geo_area": geo_area
                })
    
    with col2:
        # Median rent files
        st.subheader("Median Rent Files")
        rent_files = st.file_uploader(
            "Upload Median Rent Files", 
            type=["xlsx", "xls", "parquet"],
            accept_multiple_files=True,
            key="median_rents"
        )
        
        if rent_files:
            for file in rent_files:
                # Save the file to a temporary location
                temp_file_path = os.path.join(st.session_state['temp_dir'], file.name)
                with open(temp_file_path, "wb") as f:
                    f.write(file.getbuffer())
                
                # Try to determine the geo_area from the filename
                geo_area = None
                for area in GEO_AREAS:
                    if area.lower() in file.name.lower():
                        geo_area = area.lower()
                        break
                
                if geo_area is None:
                    geo_area = "lga"  # Default if we can't determine
                
                st.session_state['uploaded_files']["median_rents"].append({
                    "name": file.name,
                    "path": temp_file_path,
                    "geo_area": geo_area
                })
        
        # Vacancy rate files
        st.subheader("Vacancy Rate Files")
        vacancy_files = st.file_uploader(
            "Upload Vacancy Rate Files", 
            type=["xlsx", "xls", "parquet"],
            accept_multiple_files=True,
            key="vacancy_rates"
        )
        
        if vacancy_files:
            for file in vacancy_files:
                # Save the file to a temporary location
                temp_file_path = os.path.join(st.session_state['temp_dir'], file.name)
                with open(temp_file_path, "wb") as f:
                    f.write(file.getbuffer())
                
                # Try to determine the geo_area from the filename
                geo_area = None
                for area in GEO_AREAS:
                    if area.lower() in file.name.lower():
                        geo_area = area.lower()
                        break
                
                if geo_area is None:
                    geo_area = "lga"  # Default if we can't determine
                
                st.session_state['uploaded_files']["vacancy_rates"].append({
                    "name": file.name,
                    "path": temp_file_path,
                    "geo_area": geo_area
                })
    
    # Count total files uploaded
    total_files = sum(len(files) for files in st.session_state['uploaded_files'].values())
    
    if total_files > 0:
        st.success(f"{total_files} files uploaded successfully!")
        
        # Show details in an expander
        with st.expander("View uploaded files"):
            for data_type, files in st.session_state['uploaded_files'].items():
                if files:
                    st.subheader(f"{data_type.replace('_', ' ').title()} Files")
                    for file in files:
                        st.write(f"- {file['name']} ({file['geo_area']})")
    else:
        st.warning("No files uploaded. Please upload at least one data file.")

# Option 2: Scan a root folder
with folder_tab:
    st.header("Scan Root Folder")
    
    st.markdown("""
    Specify a folder containing all your data files, and we'll automatically scan for relevant files.
    
    Expected folder structure (flexible):
    - Census data/output data/dwellings
    - Median rents/output data
    - Affordability/output data
    - Rental vacancy rates/output data
    
    If your files are in a different structure, we'll try to find them anyway based on filenames.
    """)
    
    # Input for root folder path
    root_folder = st.text_input(
        "Root folder path:",
        placeholder="e.g., C:\\Rental_Data or /home/user/rental_data"
    )
    
    # Button to scan the folder
    if st.button("Scan Folder", key="scan_folder_button"):
        if root_folder:
            # Scan the root folder for data files
            scan_root_folder(root_folder)

# Proceed only if files are found
has_files = sum(len(files) for files in st.session_state['uploaded_files'].values()) > 0

if has_files:
    st.header("Geographic Selection")
    
    # Get unique geographic areas from all files
    available_geo_areas = set()
    for data_type, files in st.session_state['uploaded_files'].items():
        for file in files:
            available_geo_areas.add(file["geo_area"])
    
    # If no geo areas found, show an error
    if not available_geo_areas:
        st.error("No geographic areas found in the uploaded files.")
    else:
        # Create a selectbox for geo areas
        selected_geo_area = st.selectbox(
            "Select Geographic Area Type:",
            sorted(list(available_geo_areas))
        )
        
        # Get geographic names for the selected area
        geo_names = get_geo_names(selected_geo_area, st.session_state['uploaded_files'])
        
        if geo_names:
            # Create a selectbox for geo names
            selected_geo_name = st.selectbox(
                f"Select {selected_geo_area.upper()} Name:",
                geo_names
            )
            
            # Add a button to generate the analysis
            if st.button("Generate Analysis", type="primary"):
                with st.spinner(f"Analyzing data for {selected_geo_name}..."):
                    # Collect data for the selected area
                    data = {}
                    
                    # Collect Census dwelling data
                    census_data = collect_census_data(selected_geo_area, selected_geo_name, st.session_state['uploaded_files'])
                    data.update(census_data)
                    
                    # Collect Median Rent data
                    rent_data = collect_median_rent_data(selected_geo_area, selected_geo_name, st.session_state['uploaded_files'])
                    data.update(rent_data)
                    
                    # Collect Vacancy Rate data
                    vacancy_data = collect_vacancy_rate_data(selected_geo_area, selected_geo_name, st.session_state['uploaded_files'])
                    data.update(vacancy_data)
                    
                    # Collect Affordability data
                    affordability_data = collect_affordability_data(selected_geo_area, selected_geo_name, st.session_state['uploaded_files'])
                    data.update(affordability_data)
                    
                    # Ensure all required data is available (use defaults if missing)
                    data = ensure_default_data(data)
                    
                    # Store data in session state
                    st.session_state['data'] = data
                    
                    # Display the dashboard
                    display_dashboard(selected_geo_area, selected_geo_name, data)
                    
                    # Add Excel export option
                    st.header("Export Data")
                    excel_data = create_excel_output(selected_geo_area, selected_geo_name, data)
                    
                    st.download_button(
                        label="Download Excel Report",
                        data=excel_data,
                        file_name=f"{selected_geo_name}_{selected_geo_area}_Rental_Analysis.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.warning(f"No geographic names found for {selected_geo_area}. Try selecting a different geographic area type.")

# Add footer with info
st.markdown("---")
st.caption("* Methodology: Rental affordability is calculated by comparing median weekly rental household incomes to median weekly rents. Values above 30% indicate rental stress.")
st.caption("Sources: ABS Census, NSW Fair Trading using Corelogic and PropTrack data")
