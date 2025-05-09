import os
import pandas as pd
import numpy as np
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import re

class RentalDataAnalyzer:
    def __init__(self):
        """Initialize the Rental Data Analyzer with paths and reference data"""
        # Base directory - can be modified by user via GUI
        self.BASE_DIR = r"C:\Users\joshu\OneDrive - NSWGOV\Rental Data Model\Calculated rental data"
        
        # Geographic areas
        self.GEO_AREAS = ["CED", "GCCSA", "LGA", "SA3", "SA4", "SED", "Suburb"]
        
        # Data subdirectories
        self.SUB_DIRS = {
            "median_rents": os.path.join("Median rents", "output data"),
            "census_dwelling": os.path.join("Census data", "output data", "dwellings"),
            "census_demographics": os.path.join("Census data", "output data", "demographics"),
            "affordability": os.path.join("Affordability", "output data"),
            "vacancy_rates": os.path.join("Rental vacancy rates", "output data")
        }
        
        # File patterns for different geographic areas
        self.FILE_PATTERNS = {
            "median_rents": {area.lower(): f"{area.lower()}_rent_data" for area in self.GEO_AREAS},
            "affordability": {area.lower(): f"{area.lower()}_affordability" for area in self.GEO_AREAS},
            "vacancy_rates": {area.lower(): f"{area.lower()}_vacancy_rate" for area in self.GEO_AREAS},
            "census_dwelling": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_dwelling_tenure" for area in self.GEO_AREAS},
            "census_demographics": {area.lower(): f"census_2021_{area.upper() if area != 'Suburb' else area}_demographics" for area in self.GEO_AREAS}
        }
        
        # Greater Sydney LGAs
        self.GREATER_SYDNEY_LGAS = [
            "Bayside (NSW)", "Blacktown", "Blue Mountains", "Burwood", "Camden", "Campbelltown (NSW)", 
            "Canada Bay", "Canterbury-Bankstown", "Cumberland", "Fairfield", "Georges River", 
            "Hawkesbury", "Hornsby", "Hunters Hill", "Inner West", "Ku-ring-gai", "Lane Cove", 
            "Liverpool", "Mosman", "North Sydney", "Northern Beaches", "Parramatta", "Penrith", 
            "Randwick", "Ryde", "Strathfield", "Sutherland Shire", "Sydney", "The Hills Shire", 
            "Waverley", "Willoughby", "Woollahra", "Wollondilly"
        ]
        
        # Reference data for comparison - will be updated dynamically
        self.GS_REFERENCE_DATA = {
            "renters": {"area": "Greater Sydney", "value": None},
            "social_housing": {"area": "Greater Sydney", "value": None},
            "median_rent": {"area": "Greater Sydney", "value": None},
            "vacancy_rates": {"area": "Greater Sydney", "value": None},
            "affordability": {"area": "Greater Sydney", "value": None, "annual_change": None}
        }
        
        # Reference data for comparison - will be updated dynamically
        self.RON_REFERENCE_DATA = {
            "renters": {"area": "Rest of NSW", "value": None},
            "social_housing": {"area": "Rest of NSW", "value": None},
            "median_rent": {"area": "Rest of NSW", "value": None},
            "vacancy_rates": {"area": "Rest of NSW", "value": None},
            "affordability": {"area": "Rest of NSW", "value": None, "annual_change": None}
        }
        
        # Variables to store selections and data
        self.selected_geo_area = None
        self.selected_geo_name = None
        self.data = {}
        self.output_file = None

    def create_gui(self):
        """Create GUI for selecting geographic area and name"""
        self.root = tk.Tk()
        self.root.title("NSW Rental Data Analyzer")
        self.root.geometry("700x400")
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Data directory selection
        dir_frame = ttk.LabelFrame(main_frame, text="Data Directory", padding="5")
        dir_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.dir_var = tk.StringVar(value=self.BASE_DIR)
        dir_entry = ttk.Entry(dir_frame, textvariable=self.dir_var, width=60)
        dir_entry.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        dir_button = ttk.Button(dir_frame, text="Browse...", command=self.browse_directory)
        dir_button.pack(side=tk.RIGHT, padx=5, pady=5)
        
        # Selection frame
        selection_frame = ttk.LabelFrame(main_frame, text="Geographic Selection", padding="5")
        selection_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Geographic area type selection
        ttk.Label(selection_frame, text="Geographic Area Type:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo = ttk.Combobox(selection_frame, values=self.GEO_AREAS, state="readonly", width=30)
        self.geo_area_combo.grid(column=1, row=0, sticky=tk.W, pady=5, padx=5)
        self.geo_area_combo.bind("<<ComboboxSelected>>", self.on_geo_area_selected)
        
        # Geographic name selection
        ttk.Label(selection_frame, text="Geographic Name:").grid(column=0, row=1, sticky=tk.W, pady=5, padx=5)
        self.geo_name_combo = ttk.Combobox(selection_frame, state="disabled", width=40)
        self.geo_name_combo.grid(column=1, row=1, sticky=tk.W, pady=5, padx=5)
        
        # Output file selection
        output_frame = ttk.LabelFrame(main_frame, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.output_var = tk.StringVar()
        ttk.Label(output_frame, text="Output File:").grid(column=0, row=0, sticky=tk.W, pady=5, padx=5)
        output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=40)
        output_entry.grid(column=1, row=0, sticky=tk.W+tk.E, pady=5, padx=5)
        
        output_button = ttk.Button(output_frame, text="Browse...", command=self.browse_output_file)
        output_button.grid(column=2, row=0, sticky=tk.W, pady=5, padx=5)
        
        # Status and buttons
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, padx=5, pady=10)
        
        # Generate button
        self.generate_button = ttk.Button(status_frame, text="Generate Analysis", command=self.run_analysis, state="disabled")
        self.generate_button.pack(side=tk.RIGHT, padx=5)
        
        # Exit button
        exit_button = ttk.Button(status_frame, text="Exit", command=self.root.destroy)
        exit_button.pack(side=tk.RIGHT, padx=5)
        
        # Status label
        self.status_label = ttk.Label(status_frame, text="Select a geographic area type to begin")
        self.status_label.pack(side=tk.LEFT, padx=5)
        
        # Set column and row weights for resizing
        main_frame.columnconfigure(0, weight=1)
        
        self.root.mainloop()
    
    def browse_directory(self):
        """Browse for data directory"""
        dir_path = filedialog.askdirectory(initialdir=self.BASE_DIR, title="Select Data Directory")
        if dir_path:
            self.BASE_DIR = dir_path
            self.dir_var.set(dir_path)
            # Reset selections
            self.selected_geo_area = None
            self.selected_geo_name = None
            self.geo_area_combo.set('')
            self.geo_name_combo.set('')
            self.geo_name_combo.config(state="disabled")
            self.generate_button.config(state="disabled")
            self.status_label.config(text="Select a geographic area type to begin")
    
    def browse_output_file(self):
        """Browse for output file location"""
        file_path = filedialog.asksaveasfilename(
            initialdir=os.path.expanduser("~"), 
            title="Save Analysis As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.output_file = file_path
            self.output_var.set(file_path)
    
    def on_geo_area_selected(self, event=None):
        """Handle geographic area type selection"""
        self.selected_geo_area = self.geo_area_combo.get()
        self.status_label.config(text=f"Loading {self.selected_geo_area} names...")
        self.root.update_idletasks()
        
        # Get available geographic names
        try:
            geo_names = self.get_geo_names(self.selected_geo_area)
            self.geo_name_combo.config(values=geo_names, state="readonly")
            self.geo_name_combo.bind("<<ComboboxSelected>>", self.on_geo_name_selected)
            self.status_label.config(text=f"Select a {self.selected_geo_area} name")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to load geographic names: {str(e)}")
    
    def on_geo_name_selected(self, event=None):
        """Handle geographic name selection"""
        self.selected_geo_name = self.geo_name_combo.get()
        self.generate_button.config(state="normal")
        
        # Set default output file name if not already set
        if not self.output_var.get():
            # Sanitize filename - remove parentheses and other special characters
            safe_geo_name = self.selected_geo_name.replace("(", "").replace(")", "").replace(" ", "_")
            default_filename = f"{safe_geo_name}_{self.selected_geo_area}_Rental_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # Ensure desktop path exists
            try:
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop_path):
                    # Fall back to current directory if Desktop doesn't exist
                    desktop_path = os.getcwd()
            except:
                desktop_path = os.getcwd()
                
            self.output_var.set(os.path.join(desktop_path, default_filename))
            self.output_file = os.path.join(desktop_path, default_filename)
        
        self.status_label.config(text=f"Ready to generate analysis for {self.selected_geo_name} ({self.selected_geo_area})")
    
    def get_geo_names(self, geo_area):
        """Get available geographic names for the selected area type"""
        names = set()
        found_files = False
        
        # Try to get names from multiple data sources
        for data_type, subdir in self.SUB_DIRS.items():
            try:
                dir_path = os.path.join(self.BASE_DIR, subdir)
                if not os.path.exists(dir_path):
                    continue
                
                # Skip if pattern doesn't exist for this geo_area
                if geo_area.lower() not in self.FILE_PATTERNS[data_type]:
                    continue
                    
                file_pattern = self.FILE_PATTERNS[data_type][geo_area.lower()]
                
                for file in os.listdir(dir_path):
                    if file_pattern.lower() in file.lower():
                        found_files = True
                        file_path = os.path.join(dir_path, file)
                        print(f"Reading file: {file_path}")
                        df = self.read_data_file(file_path)
                        
                        if df is not None and not df.empty:
                            # Look for the geographic name column
                            geo_col = self.find_geographic_column(df, geo_area)
                            
                            if geo_col:
                                print(f"Found geography column: {geo_col}")
                                # Convert all values to strings and filter out likely non-geographic names
                                df[geo_col] = df[geo_col].astype(str)
                                
                                # Filter out values that look like dates or numbers
                                area_names = []
                                for name in df[geo_col].dropna().unique().tolist():
                                    name_str = str(name)
                                    # Skip if it looks like a date format (2021-01-01, etc.)
                                    if re.match(r'^\d{4}-\d{2}-\d{2}', name_str) or re.match(r'^\d{2}/\d{2}/\d{4}', name_str):
                                        continue
                                    # Skip if it's just a number
                                    if name_str.isdigit():
                                        continue
                                    # Skip very short names (likely codes)
                                    if len(name_str) < 2:
                                        continue
                                    # Skip if it's an LGA/SA code
                                    if re.match(r'^LGA\d+$', name_str) or re.match(r'^SA\d+$', name_str):
                                        continue
                                    
                                    area_names.append(name_str)
                                
                                print(f"Found {len(area_names)} geographic names")
                                names.update(area_names)
            except Exception as e:
                print(f"Error getting names from {data_type}: {str(e)}")
        
        if not found_files:
            raise Exception(f"No data files found for {geo_area}. Please check your data directory.")
        
        if not names:
            raise Exception(f"No geographic names found for {geo_area}. Check that your data files contain the expected columns.")
            
        name_list = sorted(list(names))
        if name_list:
            print(f"Final list of geographic names: {name_list[:10]}... (showing first 10 of {len(name_list)})")
        else:
            print("No geographic names found.")
        return name_list
    
    def read_data_file(self, file_path):
        """Read data from Excel or Parquet file"""
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # Use explicit converter and dtype to avoid datetime conversion issues
                return pd.read_excel(
                    file_path,
                    dtype={0: str},  # Force first column to be string
                    converters={0: str}  # Ensure first column is converted to string
                )
            elif file_path.endswith('.parquet'):
                df = pq.read_table(file_path).to_pandas()
                
                # Ensure the first column is a string to avoid comparison issues
                if len(df.columns) > 0:
                    df[df.columns[0]] = df[df.columns[0]].astype(str)
                    
                return df
            else:
                print(f"Unsupported file format: {file_path}")
                return None
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            return None
    
    def find_geographic_column(self, df, geo_area):
        """Find the column containing geographic area names"""
        # Print the column names for debugging
        print(f"Available columns: {df.columns.tolist()}")
        
        # Direct matches - highest priority columns that definitely contain geographic names
        priority_columns = ['region_name', 'area_name', 'location_name', f'{geo_area.lower()}_name', 'name']
        
        # Check for exact matches in priority columns first
        for col in priority_columns:
            if col in df.columns:
                print(f"Found priority geography column: {col}")
                
                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col
        
        # Common names for geographic columns - next priority
        geo_keywords = [
            geo_area.lower(), 'name', 'area', 'region', 'location', 'geography',
            'district', 'locality', 'suburb', 'lga', 'sa3', 'sa4', 'gccsa', 'ced', 'sed'
        ]
        
        # Check for columns that contain these keywords
        for col in df.columns:
            col_lower = str(col).lower()
            for keyword in geo_keywords:
                if keyword in col_lower and 'code' not in col_lower and 'type' not in col_lower:
                    print(f"Found geography column by keyword: {col}")
                    # Show sample values
                    if not df.empty:
                        sample = df[col].dropna().head(5).astype(str).tolist()
                        print(f"Sample values from '{col}': {sample}")
                    return col
        
        # If we haven't found a match yet, specifically look for 'region_name' which we know exists
        if 'region_name' in df.columns:
            print(f"Falling back to region_name column")
            # Show sample values
            if not df.empty:
                sample = df['region_name'].dropna().head(5).astype(str).tolist()
                print(f"Sample values from 'region_name': {sample}")
            return 'region_name'
            
        # If still no match, check for columns that might contain place names
        for col in df.columns:
            try:
                sample = df[col].dropna().head(5).astype(str).tolist()
                print(f"Checking column '{col}' with sample values: {sample}")
                
                # Skip columns where values appear to be dates or numbers
                if all(not re.match(r'^\d{4}-\d{2}-\d{2}', str(x)) and 
                       not re.match(r'^\d{2}/\d{2}/\d{4}', str(x)) and
                       not str(x).replace('.', '').isdigit() and
                       len(str(x)) > 2
                       for x in sample):
                    
                    # Check if the values look like place names (contain alphabetic characters)
                    if all(any(c.isalpha() for c in str(x)) for x in sample):
                        print(f"Detected possible location names in column: {col}")
                        return col
            except Exception as e:
                print(f"Error checking column {col}: {str(e)}")
        
        # If no suitable column found, look for codes as a last resort
        for col in df.columns:
            col_lower = str(col).lower()
            if ('code' in col_lower and any(kw in col_lower for kw in geo_keywords)) or 'region_code' in col_lower:
                print(f"Falling back to code column: {col}")
                # Show sample values
                if not df.empty:
                    sample = df[col].dropna().head(5).astype(str).tolist()
                    print(f"Sample values from '{col}': {sample}")
                return col
                
        # Absolute last resort - first column
        if len(df.columns) > 0:
            first_col = df.columns[0]
            print(f"No geographic column identified, using first column as fallback: {first_col}")
            # Show sample values
            if not df.empty:
                sample = df[first_col].dropna().head(5).astype(str).tolist()
                print(f"Sample values from first column: {sample}")
            return first_col
        
        return None
    
    def run_analysis(self):
        """Run the analysis and generate Excel output"""
        if not self.selected_geo_area or not self.selected_geo_name:
            messagebox.showerror("Error", "Please select both a geographic area type and name")
            return
        
        if not self.output_file:
            messagebox.showerror("Error", "Please select an output file location")
            return
        
        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    messagebox.showerror("Error", f"Cannot create output directory: {str(e)}")
                    return
                    
            # Make sure we have a valid file path for saving
            if not self.output_file.endswith('.xlsx'):
                self.output_file += '.xlsx'
                
            self.status_label.config(text="Collecting reference data...")
            self.root.update_idletasks()
            
            # First get reference data for Greater Sydney and Rest of NSW
            self.collect_reference_data()
                
            self.status_label.config(text="Collecting data...")
            self.root.update_idletasks()
            
            # Collect data from various sources
            self.collect_data()
            
            self.status_label.config(text="Creating Excel output...")
            self.root.update_idletasks()
            
            # Create Excel output
            self.create_excel_output()
            
            self.status_label.config(text=f"Analysis completed and saved to {self.output_file}")
            messagebox.showinfo("Success", f"Analysis completed and saved to {self.output_file}")
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in run_analysis: {str(e)}")
            print(error_details)
            self.status_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate analysis: {str(e)}\n\nSee console for details.")
    
    def collect_reference_data(self):
        """Collect reference data for Greater Sydney and Rest of NSW"""
        print(f"\n\n{'='*50}")
        print(f"COLLECTING REFERENCE DATA FOR GREATER SYDNEY AND REST OF NSW")
        print(f"{'='*50}")
        
        # Collect reference data for median rents
        try:
            print("\nCollecting median rent reference data...")
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["median_rents"])
            
            # Find files for LGA data as it's most reliable for GS aggregation
            file_pattern = self.FILE_PATTERNS["median_rents"]["lga"]
            
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    file_path = os.path.join(dir_path, file)
                    print(f"Processing rent file for reference: {file_path}")
                    df = self.read_data_file(file_path)
                    
                    if df is not None and not df.empty:
                        # If we have a month column, get the most recent month
                        latest_month = None
                        if 'month' in df.columns:
                            df['month'] = pd.to_datetime(df['month'], errors='coerce')
                            latest_month = df['month'].max()
                            df_latest = df[df['month'] == latest_month]
                            print(f"Filtered to latest month: {latest_month}")
                        else:
                            df_latest = df
                            
                        # Find the geographic column
                        geo_col = self.find_geographic_column(df_latest, "LGA")
                        
                        if geo_col:
                            print(f"Using geographic column: {geo_col}")
                            
                            # If we have property_type, get the "All Dwellings" type
                            if 'property_type' in df_latest.columns:
                                if 'All Dwellings' in df_latest['property_type'].values:
                                    df_latest = df_latest[df_latest['property_type'] == 'All Dwellings']
                                    print("Filtered to 'All Dwellings' property type")
                            
                            # Find columns for annual growth
                            growth_col = None
                            for col_suffix in ['annual_growth', 'annual_increase', 'yearly_growth', 'yearly_increase']:
                                for col in df_latest.columns:
                                    if col.lower().endswith(col_suffix.lower()):
                                        growth_col = col
                                        break
                                if growth_col:
                                    break
                            
                            # Find columns for median rent - prefer 3-month median
                            rent_col = None
                            for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                                for col in df_latest.columns:
                                    if col.startswith(col_prefix) and not any(x in col for x in ['growth', 'increase', 'change']):
                                        rent_col = col
                                        break
                                if rent_col:
                                    break
                            
                            print(f"Median rent column: {rent_col}")
                            print(f"Annual increase column: {growth_col}")
                            
                            if rent_col and growth_col:
                                # Filter for Greater Sydney LGAs
                                df_gs = df_latest[df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get statistics for Greater Sydney
                                if not df_gs.empty:
                                    # Get median rent for Greater Sydney
                                    gs_median_rent = df_gs[rent_col].mean()
                                    print(f"Greater Sydney median rent: ${gs_median_rent:.2f}")
                                    
                                    # Get median annual increase for Greater Sydney
                                    gs_annual_increase = df_gs[growth_col].mean()
                                    if gs_annual_increase < 1:  # If it's a decimal (e.g. 0.121 for 12.1%)
                                        gs_annual_increase = gs_annual_increase * 100
                                    print(f"Greater Sydney annual increase: {gs_annual_increase:.2f}%")
                                    
                                    self.GS_REFERENCE_DATA["median_rent"]["value"] = round(gs_annual_increase, 1)
                                
                                # Filter for Rest of NSW (not in Greater Sydney)
                                df_ron = df_latest[~df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get statistics for Rest of NSW
                                if not df_ron.empty:
                                    # Get median rent for Rest of NSW
                                    ron_median_rent = df_ron[rent_col].mean()
                                    print(f"Rest of NSW median rent: ${ron_median_rent:.2f}")
                                    
                                    # Get median annual increase for Rest of NSW
                                    ron_annual_increase = df_ron[growth_col].mean()
                                    if ron_annual_increase < 1:  # If it's a decimal (e.g. 0.086 for 8.6%)
                                        ron_annual_increase = ron_annual_increase * 100
                                    print(f"Rest of NSW annual increase: {ron_annual_increase:.2f}%")
                                    
                                    self.RON_REFERENCE_DATA["median_rent"]["value"] = round(ron_annual_increase, 1)
                            
                            break
        except Exception as e:
            print(f"Error collecting median rent reference data: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Fallback to default values if error
            if self.GS_REFERENCE_DATA["median_rent"]["value"] is None:
                self.GS_REFERENCE_DATA["median_rent"]["value"] = 12.1
            if self.RON_REFERENCE_DATA["median_rent"]["value"] is None:
                self.RON_REFERENCE_DATA["median_rent"]["value"] = 8.6
        
        # Collect reference data for vacancy rates
        try:
            print("\nCollecting vacancy rate reference data...")
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["vacancy_rates"])
            
            # Find files for LGA data
            file_pattern = self.FILE_PATTERNS["vacancy_rates"]["lga"]
            
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    file_path = os.path.join(dir_path, file)
                    print(f"Processing vacancy rate file for reference: {file_path}")
                    df = self.read_data_file(file_path)
                    
                    if df is not None and not df.empty:
                        # If we have a month column, get the most recent month
                        latest_month = None
                        if 'month' in df.columns:
                            df['month'] = pd.to_datetime(df['month'], errors='coerce')
                            latest_month = df['month'].max()
                            df_latest = df[df['month'] == latest_month]
                            print(f"Filtered to latest month: {latest_month}")
                            
                            # Get data from one year ago
                            one_year_ago = latest_month - pd.DateOffset(months=12)
                            df_year_ago = df[df['month'] == one_year_ago]
                            if df_year_ago.empty:
                                print(f"No data found for vacancy rates one year ago ({one_year_ago})")
                        else:
                            df_latest = df
                            df_year_ago = pd.DataFrame()
                        
                        # Find the geographic column
                        geo_col = self.find_geographic_column(df_latest, "LGA")
                        
                        if geo_col:
                            print(f"Using geographic column: {geo_col}")
                            
                            # Find vacancy rate column - specifically use rental_vacancy_rate_3m_smoothed
                            rate_col = None
                            if 'rental_vacancy_rate_3m_smoothed' in df_latest.columns:
                                rate_col = 'rental_vacancy_rate_3m_smoothed'
                                print(f"Found specific vacancy rate column: {rate_col}")
                            else:
                                # Fallback to other columns if the specific one is not found
                                for col_name in ['rental_vacancy_rate', 'vacancy_rate', 'rate']:
                                    if col_name in df_latest.columns:
                                        rate_col = col_name
                                        print(f"Using fallback vacancy rate column: {rate_col}")
                                        break
                            
                            if rate_col:
                                # Filter for Greater Sydney LGAs
                                df_gs_latest = df_latest[df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get vacancy rate change for Greater Sydney
                                if not df_gs_latest.empty:
                                    # Calculate current average vacancy rate for Greater Sydney
                                    gs_current_rate = df_gs_latest[rate_col].mean()
                                    
                                    # Ensure it's a percentage (not decimal)
                                    if gs_current_rate > 0 and gs_current_rate < 1:
                                        gs_current_rate = gs_current_rate * 100
                                    
                                    # If we have historical data, calculate the annual change
                                    if not df_year_ago.empty:
                                        df_gs_year_ago = df_year_ago[df_year_ago[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                        if not df_gs_year_ago.empty and rate_col in df_gs_year_ago.columns:
                                            gs_prev_rate = df_gs_year_ago[rate_col].mean()
                                            
                                            # Ensure it's a percentage (not decimal)
                                            if gs_prev_rate > 0 and gs_prev_rate < 1:
                                                gs_prev_rate = gs_prev_rate * 100
                                                
                                            gs_change = gs_current_rate - gs_prev_rate
                                            print(f"Greater Sydney vacancy rate change: {gs_change:.2f}%")
                                            
                                            self.GS_REFERENCE_DATA["vacancy_rates"]["value"] = round(gs_change, 1)
                                
                                # Filter for Rest of NSW (not in Greater Sydney)
                                df_ron_latest = df_latest[~df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get vacancy rate change for Rest of NSW
                                if not df_ron_latest.empty:
                                    # Calculate current average vacancy rate for Rest of NSW
                                    ron_current_rate = df_ron_latest[rate_col].mean()
                                    
                                    # Ensure it's a percentage (not decimal)
                                    if ron_current_rate > 0 and ron_current_rate < 1:
                                        ron_current_rate = ron_current_rate * 100
                                    
                                    # If we have historical data, calculate the annual change
                                    if not df_year_ago.empty:
                                        df_ron_year_ago = df_year_ago[~df_year_ago[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                        if not df_ron_year_ago.empty and rate_col in df_ron_year_ago.columns:
                                            ron_prev_rate = df_ron_year_ago[rate_col].mean()
                                            
                                            # Ensure it's a percentage (not decimal)
                                            if ron_prev_rate > 0 and ron_prev_rate < 1:
                                                ron_prev_rate = ron_prev_rate * 100
                                                
                                            ron_change = ron_current_rate - ron_prev_rate
                                            print(f"Rest of NSW vacancy rate change: {ron_change:.2f}%")
                                            
                                            self.RON_REFERENCE_DATA["vacancy_rates"]["value"] = round(ron_change, 1)
                            
                            break
        except Exception as e:
            print(f"Error collecting vacancy rate reference data: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Fallback to default values if error
            if self.GS_REFERENCE_DATA["vacancy_rates"]["value"] is None:
                self.GS_REFERENCE_DATA["vacancy_rates"]["value"] = -0.3
            if self.RON_REFERENCE_DATA["vacancy_rates"]["value"] is None:
                self.RON_REFERENCE_DATA["vacancy_rates"]["value"] = -0.1
        
        # Collect reference data for affordability
        try:
            print("\nCollecting affordability reference data...")
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["affordability"])
            
            # Find files for LGA data
            file_pattern = self.FILE_PATTERNS["affordability"]["lga"]
            
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    file_path = os.path.join(dir_path, file)
                    print(f"Processing affordability file for reference: {file_path}")
                    df = self.read_data_file(file_path)
                    
                    if df is not None and not df.empty:
                        # If we have a month column, get the most recent month
                        latest_month = None
                        if 'month' in df.columns:
                            df['month'] = pd.to_datetime(df['month'], errors='coerce')
                            latest_month = df['month'].max()
                            df_latest = df[df['month'] == latest_month]
                            print(f"Filtered to latest month: {latest_month}")
                        else:
                            df_latest = df
                        
                        # Find the geographic column
                        geo_col = self.find_geographic_column(df_latest, "LGA")
                        
                        if geo_col:
                            print(f"Using geographic column: {geo_col}")
                            
                            # Find affordability column - look for keywords
                            pct_col = None
                            
                            # First priority: direct affordability columns
                            affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                            if affordability_columns:
                                # Prefer 3-month affordability for stability
                                if 'rental_affordability_3mo' in affordability_columns:
                                    pct_col = 'rental_affordability_3mo'
                                elif 'rental_affordability_1mo' in affordability_columns:
                                    pct_col = 'rental_affordability_1mo'
                                else:
                                    pct_col = affordability_columns[0]  # Take the first one if no preferred column
                            
                            # If no direct affordability column, try to find rent-to-income ratio
                            if not pct_col:
                                for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                                    for col in df_latest.columns:
                                        if all(keyword.lower() in col.lower() for keyword in keywords):
                                            pct_col = col
                                            break
                                    if pct_col:
                                        break
                            
                            # Find improvement column
                            improvement_col = None
                            for col_suffix in ['improvement', 'change', 'growth']:
                                for col in df_latest.columns:
                                    if col_suffix in col.lower() and any(x in col.lower() for x in ['annual', 'yearly']):
                                        improvement_col = col
                                        break
                                if improvement_col:
                                    break
                            
                            print(f"Affordability column: {pct_col}")
                            print(f"Annual improvement column: {improvement_col}")
                            
                            if pct_col:
                                # Filter for Greater Sydney LGAs
                                df_gs = df_latest[df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get statistics for Greater Sydney
                                if not df_gs.empty:
                                    # Get average affordability for Greater Sydney
                                    gs_affordability = df_gs[pct_col].mean()
                                    
                                    # Ensure it's a percentage (not decimal)
                                    if gs_affordability > 0 and gs_affordability < 1:
                                        gs_affordability = gs_affordability * 100
                                        
                                    print(f"Greater Sydney affordability: {gs_affordability:.2f}%")
                                    
                                    self.GS_REFERENCE_DATA["affordability"]["value"] = round(gs_affordability, 1)
                                    
                                    # Get average annual change for Greater Sydney
                                    if improvement_col and improvement_col in df_gs.columns:
                                        gs_change = df_gs[improvement_col].mean()
                                        
                                        # Ensure it's a percentage (not decimal)
                                        if gs_change < 1 and gs_change > -1:
                                            gs_change = gs_change * 100
                                            
                                        print(f"Greater Sydney affordability change: {gs_change:.2f}%")
                                        
                                        self.GS_REFERENCE_DATA["affordability"]["annual_change"] = round(gs_change, 1)
                                
                                # Filter for Rest of NSW (not in Greater Sydney)
                                df_ron = df_latest[~df_latest[geo_col].isin(self.GREATER_SYDNEY_LGAS)]
                                
                                # Get statistics for Rest of NSW
                                if not df_ron.empty:
                                    # Get average affordability for Rest of NSW
                                    ron_affordability = df_ron[pct_col].mean()
                                    
                                    # Ensure it's a percentage (not decimal)
                                    if ron_affordability > 0 and ron_affordability < 1:
                                        ron_affordability = ron_affordability * 100
                                        
                                    print(f"Rest of NSW affordability: {ron_affordability:.2f}%")
                                    
                                    self.RON_REFERENCE_DATA["affordability"]["value"] = round(ron_affordability, 1)
                                    
                                    # Get average annual change for Rest of NSW
                                    if improvement_col and improvement_col in df_ron.columns:
                                        ron_change = df_ron[improvement_col].mean()
                                        
                                        # Ensure it's a percentage (not decimal)
                                        if ron_change < 1 and ron_change > -1:
                                            ron_change = ron_change * 100
                                            
                                        print(f"Rest of NSW affordability change: {ron_change:.2f}%")
                                        
                                        self.RON_REFERENCE_DATA["affordability"]["annual_change"] = round(ron_change, 1)
                            
                            break
        except Exception as e:
            print(f"Error collecting affordability reference data: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Fallback to default values if error
            if self.GS_REFERENCE_DATA["affordability"]["value"] is None:
                self.GS_REFERENCE_DATA["affordability"]["value"] = 45.2
            if self.GS_REFERENCE_DATA["affordability"]["annual_change"] is None:
                self.GS_REFERENCE_DATA["affordability"]["annual_change"] = 5.1
            if self.RON_REFERENCE_DATA["affordability"]["value"] is None:
                self.RON_REFERENCE_DATA["affordability"]["value"] = 41.7
            if self.RON_REFERENCE_DATA["affordability"]["annual_change"] is None:
                self.RON_REFERENCE_DATA["affordability"]["annual_change"] = 3.5
        
        # Census data (renter percentage and social housing) are more static, so use defaults if dynamic lookup fails
        # Renters data
        try:
            print("\nCollecting census renters reference data...")
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["census_dwelling"])
            
            # Find census dwelling files
            file_pattern = self.FILE_PATTERNS["census_dwelling"]["gccsa"]  # GCCSA level for Greater Sydney
            
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    file_path = os.path.join(dir_path, file)
                    print(f"Processing census file for reference: {file_path}")
                    df = self.read_data_file(file_path)
                    
                    if df is not None and not df.empty:
                        # Find the geographic column
                        geo_col = self.find_geographic_column(df, "GCCSA")
                        
                        if geo_col:
                            print(f"Using geographic column: {geo_col}")
                            
                            # Look for Greater Sydney
                            gs_rows = df[df[geo_col].str.contains("Greater Sydney", case=False, na=False)]
                            if not gs_rows.empty:
                                # Calculate rental percentage
                                total_dwellings = None
                                if "dwellings" in gs_rows.columns:
                                    total_dwellings = float(gs_rows["dwellings"].iloc[0]) if not pd.isna(gs_rows["dwellings"].iloc[0]) else 0
                                
                                total_rented = None
                                if "dwellings_rented" in gs_rows.columns:
                                    total_rented = float(gs_rows["dwellings_rented"].iloc[0]) if not pd.isna(gs_rows["dwellings_rented"].iloc[0]) else 0
                                
                                if total_dwellings is not None and total_rented is not None and total_dwellings > 0:
                                    rental_pct = (total_rented / total_dwellings) * 100
                                    print(f"Greater Sydney renter percentage: {rental_pct:.1f}%")
                                    
                                    self.GS_REFERENCE_DATA["renters"]["value"] = round(rental_pct, 1)
                            
                            # Look for Rest of NSW (State minus Greater Sydney)
                            nsw_rows = df[df[geo_col].str.contains("New South Wales", case=False, na=False)]
                            if not nsw_rows.empty and not gs_rows.empty:
                                # Calculate total NSW dwellings
                                nsw_total_dwellings = 0
                                if "dwellings" in nsw_rows.columns:
                                    nsw_total_dwellings = float(nsw_rows["dwellings"].iloc[0]) if not pd.isna(nsw_rows["dwellings"].iloc[0]) else 0
                                
                                nsw_total_rented = 0
                                if "dwellings_rented" in nsw_rows.columns:
                                    nsw_total_rented = float(nsw_rows["dwellings_rented"].iloc[0]) if not pd.isna(nsw_rows["dwellings_rented"].iloc[0]) else 0
                                
                                # Get GS values
                                gs_total_dwellings = 0
                                if "dwellings" in gs_rows.columns:
                                    gs_total_dwellings = float(gs_rows["dwellings"].iloc[0]) if not pd.isna(gs_rows["dwellings"].iloc[0]) else 0
                                
                                gs_total_rented = 0
                                if "dwellings_rented" in gs_rows.columns:
                                    gs_total_rented = float(gs_rows["dwellings_rented"].iloc[0]) if not pd.isna(gs_rows["dwellings_rented"].iloc[0]) else 0
                                
                                # Calculate Rest of NSW values
                                ron_total_dwellings = nsw_total_dwellings - gs_total_dwellings
                                ron_total_rented = nsw_total_rented - gs_total_rented
                                
                                if ron_total_dwellings > 0:
                                    ron_rental_pct = (ron_total_rented / ron_total_dwellings) * 100
                                    print(f"Rest of NSW renter percentage: {ron_rental_pct:.1f}%")
                                    
                                    self.RON_REFERENCE_DATA["renters"]["value"] = round(ron_rental_pct, 1)
                            
                            break
        except Exception as e:
            print(f"Error collecting census renters reference data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # Social housing data
        try:
            print("\nCollecting social housing reference data...")
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["census_dwelling"])
            
            # Find census dwelling files
            file_pattern = self.FILE_PATTERNS["census_dwelling"]["gccsa"]  # GCCSA level for Greater Sydney
            
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    file_path = os.path.join(dir_path, file)
                    print(f"Processing census file for social housing reference: {file_path}")
                    df = self.read_data_file(file_path)
                    
                    if df is not None and not df.empty:
                        # Find the geographic column
                        geo_col = self.find_geographic_column(df, "GCCSA")
                        
                        if geo_col:
                            print(f"Using geographic column: {geo_col}")
                            
                            # Look for Greater Sydney
                            gs_rows = df[df[geo_col].str.contains("Greater Sydney", case=False, na=False)]
                            if not gs_rows.empty:
                                # Calculate social housing percentage
                                total_dwellings = None
                                if "dwellings" in gs_rows.columns:
                                    total_dwellings = float(gs_rows["dwellings"].iloc[0]) if not pd.isna(gs_rows["dwellings"].iloc[0]) else 0
                                
                                # Find social housing data - specifically add dwellings_rented_sha + dwellings_rented_chp
                                gs_social_housing_sha = 0
                                gs_social_housing_chp = 0
                                
                                # Get SHA data
                                if "dwellings_rented_sha" in gs_rows.columns:
                                    sha_value = gs_rows["dwellings_rented_sha"].iloc[0]
                                    gs_social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                                
                                # Get CHP data
                                if "dwellings_rented_chp" in gs_rows.columns:
                                    chp_value = gs_rows["dwellings_rented_chp"].iloc[0]
                                    gs_social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                                
                                # Calculate total social housing
                                gs_total_social = gs_social_housing_sha + gs_social_housing_chp
                                
                                if total_dwellings is not None and total_dwellings > 0:
                                    gs_social_pct = (gs_total_social / total_dwellings) * 100
                                    print(f"Greater Sydney social housing percentage: {gs_social_pct:.1f}%")
                                    
                                    self.GS_REFERENCE_DATA["social_housing"]["value"] = round(gs_social_pct, 1)
                            
                            # Look for Rest of NSW (State minus Greater Sydney)
                            nsw_rows = df[df[geo_col].str.contains("New South Wales", case=False, na=False)]
                            if not nsw_rows.empty and not gs_rows.empty:
                                # Calculate total NSW dwellings
                                nsw_total_dwellings = 0
                                if "dwellings" in nsw_rows.columns:
                                    nsw_total_dwellings = float(nsw_rows["dwellings"].iloc[0]) if not pd.isna(nsw_rows["dwellings"].iloc[0]) else 0
                                
                                # Get NSW social housing
                                nsw_social_housing_sha = 0
                                nsw_social_housing_chp = 0
                                
                                if "dwellings_rented_sha" in nsw_rows.columns:
                                    sha_value = nsw_rows["dwellings_rented_sha"].iloc[0]
                                    nsw_social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                                
                                if "dwellings_rented_chp" in nsw_rows.columns:
                                    chp_value = nsw_rows["dwellings_rented_chp"].iloc[0]
                                    nsw_social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                                
                                nsw_total_social = nsw_social_housing_sha + nsw_social_housing_chp
                                
                                # Get GS values
                                gs_total_dwellings = 0
                                if "dwellings" in gs_rows.columns:
                                    gs_total_dwellings = float(gs_rows["dwellings"].iloc[0]) if not pd.isna(gs_rows["dwellings"].iloc[0]) else 0
                                
                                gs_total_social = 0
                                if "dwellings_rented_sha" in gs_rows.columns:
                                    sha_value = gs_rows["dwellings_rented_sha"].iloc[0]
                                    gs_social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                                    gs_total_social += gs_social_housing_sha
                                
                                if "dwellings_rented_chp" in gs_rows.columns:
                                    chp_value = gs_rows["dwellings_rented_chp"].iloc[0]
                                    gs_social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                                    gs_total_social += gs_social_housing_chp
                                
                                # Calculate Rest of NSW values
                                ron_total_dwellings = nsw_total_dwellings - gs_total_dwellings
                                ron_total_social = nsw_total_social - gs_total_social
                                
                                if ron_total_dwellings > 0:
                                    ron_social_pct = (ron_total_social / ron_total_dwellings) * 100
                                    print(f"Rest of NSW social housing percentage: {ron_social_pct:.1f}%")
                                    
                                    self.RON_REFERENCE_DATA["social_housing"]["value"] = round(ron_social_pct, 1)
                            
                            break
        except Exception as e:
            print(f"Error collecting social housing reference data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # Use default values for any missing reference data
        if self.GS_REFERENCE_DATA["renters"]["value"] is None:
            self.GS_REFERENCE_DATA["renters"]["value"] = 32.6
        if self.RON_REFERENCE_DATA["renters"]["value"] is None:
            self.RON_REFERENCE_DATA["renters"]["value"] = 26.8
            
        if self.GS_REFERENCE_DATA["social_housing"]["value"] is None:
            self.GS_REFERENCE_DATA["social_housing"]["value"] = 4.5
        if self.RON_REFERENCE_DATA["social_housing"]["value"] is None:
            self.RON_REFERENCE_DATA["social_housing"]["value"] = 4.0
            
        if self.GS_REFERENCE_DATA["median_rent"]["value"] is None:
            self.GS_REFERENCE_DATA["median_rent"]["value"] = 12.1
        if self.RON_REFERENCE_DATA["median_rent"]["value"] is None:
            self.RON_REFERENCE_DATA["median_rent"]["value"] = 8.6
            
        if self.GS_REFERENCE_DATA["vacancy_rates"]["value"] is None:
            self.GS_REFERENCE_DATA["vacancy_rates"]["value"] = -0.3
        if self.RON_REFERENCE_DATA["vacancy_rates"]["value"] is None:
            self.RON_REFERENCE_DATA["vacancy_rates"]["value"] = -0.1
            
        if self.GS_REFERENCE_DATA["affordability"]["value"] is None:
            self.GS_REFERENCE_DATA["affordability"]["value"] = 45.2
        if self.GS_REFERENCE_DATA["affordability"]["annual_change"] is None:
            self.GS_REFERENCE_DATA["affordability"]["annual_change"] = 5.1
        if self.RON_REFERENCE_DATA["affordability"]["value"] is None:
            self.RON_REFERENCE_DATA["affordability"]["value"] = 41.7
        if self.RON_REFERENCE_DATA["affordability"]["annual_change"] is None:
            self.RON_REFERENCE_DATA["affordability"]["annual_change"] = 3.5
        
        print("\nReference data collection complete")
        print(f"Greater Sydney reference data: {self.GS_REFERENCE_DATA}")
        print(f"Rest of NSW reference data: {self.RON_REFERENCE_DATA}")
        print(f"{'='*50}\n")
    
    def collect_data(self):
        """Collect data from various sources"""
        self.data = {}
        print(f"\n\n{'='*50}")
        print(f"STARTING DATA COLLECTION FOR: {self.selected_geo_name} ({self.selected_geo_area})")
        print(f"{'='*50}")
        
        # We'll use both Greater Sydney and Rest of NSW reference data for comparisons
        print(f"Will compare metrics to both Greater Sydney and Rest of NSW reference data")
        
        # Collect Census dwelling data
        print("\nCollecting Census dwelling data...")
        try:
            # Find census dwelling file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["census_dwelling"])
            print(f"Looking for census files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["census_dwelling"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            census_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    census_files_found.append(file)
            
            print(f"Census files found: {census_files_found}")
            
            for file in census_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing census file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Print a few sample values from the column
                        sample_values = df[geo_col].dropna().unique()[:10].tolist()
                        print(f"Sample values in column: {sample_values}")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # Calculate rental percentage by finding total rental and total dwellings
                            # Method 1: Look for direct percentage columns
                            pct_col = None
                            for col in df_filtered.columns:
                                if "rented" in col.lower() and "percent" in col.lower():
                                    pct_col = col
                                    break
                            
                            # Method 2: Calculate from raw counts
                            total_dwellings = None
                            if "dwellings" in df_filtered.columns:
                                total_dwellings = float(df_filtered["dwellings"].iloc[0]) if not pd.isna(df_filtered["dwellings"].iloc[0]) else 0
                            
                            total_rented = None
                            if "dwellings_rented" in df_filtered.columns:
                                total_rented = float(df_filtered["dwellings_rented"].iloc[0]) if not pd.isna(df_filtered["dwellings_rented"].iloc[0]) else 0
                            
                            # Calculate percentage if we have the raw counts
                            if total_dwellings is not None and total_rented is not None and total_dwellings > 0:
                                rental_pct = (total_rented / total_dwellings) * 100
                                rental_count = int(total_rented)
                                print(f"Calculated rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")
                                
                                self.data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["renters"]
                                }
                            elif pct_col:
                                rental_pct = float(df_filtered[pct_col].iloc[0]) if not pd.isna(df_filtered[pct_col].iloc[0]) else 0
                                renter_count_col = pct_col.replace("percent", "count")
                                rental_count = int(df_filtered[renter_count_col].iloc[0]) if renter_count_col in df_filtered.columns and not pd.isna(df_filtered[renter_count_col].iloc[0]) else 0
                                print(f"Found rental percentage: {rental_pct:.1f}% ({rental_count} dwellings)")
                                
                                self.data["renters"] = {
                                    "percentage": round(rental_pct, 1),
                                    "count": rental_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["renters"]
                                }
                            
                            # Find social housing data - specifically add dwellings_rented_sha + dwellings_rented_chp
                            social_housing_sha = 0
                            social_housing_chp = 0
                            
                            # Get SHA data
                            if "dwellings_rented_sha" in df_filtered.columns:
                                sha_value = df_filtered["dwellings_rented_sha"].iloc[0]
                                social_housing_sha = float(sha_value) if not pd.isna(sha_value) else 0
                                print(f"SHA dwellings: {social_housing_sha}")
                            
                            # Get CHP data
                            if "dwellings_rented_chp" in df_filtered.columns:
                                chp_value = df_filtered["dwellings_rented_chp"].iloc[0]
                                social_housing_chp = float(chp_value) if not pd.isna(chp_value) else 0
                                print(f"CHP dwellings: {social_housing_chp}")
                            
                            # Calculate total social housing
                            total_social = social_housing_sha + social_housing_chp
                            print(f"Total social housing dwellings: {total_social}")
                            
                            # Calculate social housing percentage
                            if total_dwellings is not None and total_dwellings > 0:
                                social_pct = (total_social / total_dwellings) * 100
                                social_count = int(total_social)
                                print(f"Calculated social housing percentage: {social_pct:.1f}% ({social_count} dwellings)")
                                
                                self.data["social_housing"] = {
                                    "percentage": round(social_pct, 1),
                                    "count": social_count,
                                    "period": "2021",
                                    "source": "ABS Census",
                                    "comparison_gs": self.GS_REFERENCE_DATA["social_housing"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["social_housing"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting census data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "renters" not in self.data:
            print("Using DEFAULT renter data")
            self.data["renters"] = {
                "percentage": 25.5,
                "count": 8402,
                "period": "2021",
                "source": "ABS Census",
                "comparison_gs": self.GS_REFERENCE_DATA["renters"],
                "comparison_ron": self.RON_REFERENCE_DATA["renters"]
            }
            
        if "social_housing" not in self.data:
            print("Using DEFAULT social housing data")
            self.data["social_housing"] = {
                "percentage": 2.8,
                "count": 938,
                "period": "2021",
                "source": "ABS Census",
                "comparison_gs": self.GS_REFERENCE_DATA["social_housing"],
                "comparison_ron": self.RON_REFERENCE_DATA["social_housing"]
            }
        
        # Collect Median Rent data
        print("\nCollecting Median Rent data...")
        try:
            # Find median rent file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["median_rents"])
            print(f"Looking for median rent files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["median_rents"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            rent_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    rent_files_found.append(file)
            
            print(f"Rent files found: {rent_files_found}")
            
            for file in rent_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing rent file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Print a few sample values from the column
                        sample_values = df[geo_col].dropna().unique()[:10].tolist()
                        print(f"Sample values in column: {sample_values}")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            latest_month = None
                            if 'month' in df_filtered.columns:
                                try:
                                    df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                    latest_month = df_filtered['month'].max()
                                    
                                    # Format latest_month for output
                                    latest_month_str = latest_month.strftime("%b-%Y")  # e.g., "Apr-2025"
                                    
                                    df_latest = df_filtered[df_filtered['month'] == latest_month]
                                    print(f"Filtered to latest month: {latest_month} ({latest_month_str})")
                                    
                                    # Find data from 12 months ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    df_year_ago = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    if df_year_ago.empty:
                                        print(f"No data found for exactly 12 months ago. Looking for closest month before that date...")
                                        # Try to find the closest month before one year ago
                                        prior_months = df_filtered[df_filtered['month'] < one_year_ago]['month']
                                        if not prior_months.empty:
                                            closest_prior_month = prior_months.max()
                                            df_year_ago = df_filtered[df_filtered['month'] == closest_prior_month]
                                            print(f"Using data from {closest_prior_month} as previous year comparison")
                                except Exception as e:
                                    print(f"Error processing dates: {str(e)}")
                                    df_latest = df_filtered
                                    df_year_ago = pd.DataFrame()  # Empty DataFrame if we couldn't process dates
                            else:
                                df_latest = df_filtered
                                df_year_ago = pd.DataFrame()  # Empty DataFrame if no month column
                                
                            # If we have property_type, get the "All Dwellings" type
                            if 'property_type' in df_latest.columns:
                                if 'All Dwellings' in df_latest['property_type'].values:
                                    df_latest = df_latest[df_latest['property_type'] == 'All Dwellings']
                                    print("Filtered to 'All Dwellings' property type")
                                    # Also filter year ago data if we have it
                                    if not df_year_ago.empty and 'property_type' in df_year_ago.columns:
                                        if 'All Dwellings' in df_year_ago['property_type'].values:
                                            df_year_ago = df_year_ago[df_year_ago['property_type'] == 'All Dwellings']
                            
                            # Find columns for median rent data - prefer 3-month median
                            rent_col = None
                            for col_prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent', 'rent_median']:
                                for col in df_latest.columns:
                                    if col.startswith(col_prefix) and not any(x in col for x in ['growth', 'increase', 'change']):
                                        rent_col = col
                                        break
                                if rent_col:
                                    break
                                    
                            # Find annual growth column
                            growth_col = None
                            for col_suffix in ['annual_growth', 'annual_increase', 'yearly_growth', 'yearly_increase']:
                                for col in df_latest.columns:
                                    if col.endswith(col_suffix):
                                        growth_col = col
                                        break
                                if growth_col:
                                    break
                            
                            print(f"Median rent column: {rent_col}")
                            print(f"Annual increase column: {growth_col}")
                            
                            # Extract data
                            if rent_col:
                                # Get the median rent value
                                if len(df_latest) > 0:
                                    rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                                    print(f"Current median rent: ${rent_value}")
                                    
                                    # Get annual increase - prefer to calculate from year ago data
                                    annual_increase = None
                                    prev_year_rent = None
                                    
                                    # Method 1: Calculate from year ago data (most accurate)
                                    if not df_year_ago.empty and rent_col in df_year_ago.columns:
                                        prev_year_rent = float(df_year_ago[rent_col].iloc[0]) if not pd.isna(df_year_ago[rent_col].iloc[0]) else 0
                                        
                                        if prev_year_rent > 0:
                                            annual_increase = ((rent_value - prev_year_rent) / prev_year_rent) * 100
                                            print(f"Calculated annual increase: {annual_increase:.1f}% (from ${prev_year_rent} to ${rent_value})")
                                    
                                    # Method 2: Use provided annual increase column
                                    if annual_increase is None and growth_col and len(df_latest) > 0:
                                        annual_increase_value = df_latest[growth_col].iloc[0]
                                        if not pd.isna(annual_increase_value):
                                            annual_increase = float(annual_increase_value) * 100 if float(annual_increase_value) < 1 else float(annual_increase_value)
                                            print(f"Using provided annual increase: {annual_increase:.1f}%")
                                            
                                            # If we have current rent and annual increase but not previous rent, calculate it
                                            if prev_year_rent is None:
                                                prev_year_rent = rent_value / (1 + (annual_increase / 100))
                                                print(f"Calculated previous year rent: ${prev_year_rent:.2f}")
                                    
                                    self.data["median_rent"] = {
                                        "value": int(round(rent_value, 0)),
                                        "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                        "source": "NSW Fair Trading Corelogic Data",
                                        "annual_increase": round(annual_increase, 1) if annual_increase is not None else 0,
                                        "previous_year_rent": int(round(prev_year_rent, 0)) if prev_year_rent is not None else None,
                                        "comparison_gs": self.GS_REFERENCE_DATA["median_rent"],
                                        "comparison_ron": self.RON_REFERENCE_DATA["median_rent"]
                                    }
                    break
        except Exception as e:
            print(f"Error collecting median rent data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "median_rent" not in self.data:
            print("Using DEFAULT median rent data")
            self.data["median_rent"] = {
                "value": 595,
                "period": "Apr-25",
                "source": "NSW Fair Trading Corelogic Data",
                "annual_increase": 10.2,
                "previous_year_rent": 540,
                "comparison_gs": self.GS_REFERENCE_DATA["median_rent"],
                "comparison_ron": self.RON_REFERENCE_DATA["median_rent"]
            }
        
        # Collect Vacancy Rate data
        print("\nCollecting Vacancy Rate data...")
        try:
            # Find vacancy rate file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["vacancy_rates"])
            print(f"Looking for vacancy rate files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["vacancy_rates"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            vacancy_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    vacancy_files_found.append(file)
            
            print(f"Vacancy rate files found: {vacancy_files_found}")
            
            for file in vacancy_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing vacancy rate file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            if 'month' in df_filtered.columns:
                                df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                latest_month = df_filtered['month'].max()
                                df_latest = df_filtered[df_filtered['month'] == latest_month]
                                print(f"Filtered to latest month: {latest_month}")
                            else:
                                df_latest = df_filtered
                                
                            # Find vacancy rate column - specifically use rental_vacancy_rate_3m_smoothed
                            rate_col = None
                            if 'rental_vacancy_rate_3m_smoothed' in df_latest.columns:
                                rate_col = 'rental_vacancy_rate_3m_smoothed'
                                print(f"Found specific vacancy rate column: {rate_col}")
                            else:
                                # Fallback to other columns if the specific one is not found
                                for col_name in ['rental_vacancy_rate', 'vacancy_rate', 'rate']:
                                    if col_name in df_latest.columns:
                                        rate_col = col_name
                                        print(f"Using fallback vacancy rate column: {rate_col}")
                                        break
                                    
                            # Find annual change column
                            change_col = None
                            for col_suffix in ['annual_change', 'yearly_change', 'annual_growth']:
                                for col in df_latest.columns:
                                    if col_suffix in col.lower():
                                        change_col = col
                                        break
                                if change_col:
                                    break
                            
                            # Instead of calculating the change, get the rate from 12 months ago
                            previous_year_rate = None
                            if 'month' in df_filtered.columns and rate_col:
                                try:
                                    # Get current month's value
                                    current_value = float(df_latest[rate_col].iloc[0]) if not pd.isna(df_latest[rate_col].iloc[0]) else 0
                                    
                                    # Try to find data from a year ago
                                    one_year_ago = latest_month - pd.DateOffset(months=12)
                                    one_year_ago_str = one_year_ago.strftime("%b-%Y")
                                    year_ago_data = df_filtered[df_filtered['month'] == one_year_ago]
                                    
                                    if not year_ago_data.empty and rate_col in year_ago_data.columns:
                                        year_ago_value = float(year_ago_data[rate_col].iloc[0]) if not pd.isna(year_ago_data[rate_col].iloc[0]) else 0
                                        
                                        # Ensure it's formatted as a percentage (but not multiplied)
                                        previous_year_rate = year_ago_value
                                        if previous_year_rate > 0 and previous_year_rate < 1:
                                            # Keep as decimal - we'll format correctly later
                                            pass
                                            
                                        print(f"Found vacancy rate from {one_year_ago_str}: {previous_year_rate}")
                                except Exception as e:
                                    print(f"Error getting previous year vacancy rate: {str(e)}")
                            
                            print(f"Vacancy rate column: {rate_col}")
                            
                            # Extract data
                            if rate_col and len(df_latest) > 0:
                                rate_value = float(df_latest[rate_col].iloc[0]) if not pd.isna(df_latest[rate_col].iloc[0]) else 0
                                
                                # Check if the value is already in percentage format (>1) or decimal format (<1)
                                # For vacancy rates, we want to store the actual percentage (e.g., 0.75% not 75%)
                                # But keep as decimal internally for calculations
                                is_decimal_format = (rate_value > 0 and rate_value < 1)
                                
                                # Print debug info to check the value
                                print(f"Raw vacancy rate value: {rate_value}")
                                print(f"Is decimal format: {is_decimal_format}")
                                
                                # For previous year rate, also check format
                                if previous_year_rate is not None:
                                    is_prev_decimal = (previous_year_rate > 0 and previous_year_rate < 1)
                                    print(f"Raw previous year rate: {previous_year_rate}")
                                    print(f"Previous year is decimal format: {is_prev_decimal}")
                                
                                self.data["vacancy_rates"] = {
                                    "value": rate_value,  # Store as decimal
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "previous_year_rate": previous_year_rate,  # Store as decimal
                                    "comparison_gs": self.GS_REFERENCE_DATA["vacancy_rates"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["vacancy_rates"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting vacancy rate data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "vacancy_rates" not in self.data:
            print("Using DEFAULT vacancy rate data")
            self.data["vacancy_rates"] = {
                "value": 0.72,  # Stored as decimal
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "previous_year_rate": 1.0,  # Previous year also as decimal
                "comparison_gs": self.GS_REFERENCE_DATA["vacancy_rates"],
                "comparison_ron": self.RON_REFERENCE_DATA["vacancy_rates"]
            }
        
        # Collect Affordability data
        print("\nCollecting Affordability data...")
        try:
            # Find affordability file
            dir_path = os.path.join(self.BASE_DIR, self.SUB_DIRS["affordability"])
            print(f"Looking for affordability files in: {dir_path}")
            file_pattern = self.FILE_PATTERNS["affordability"][self.selected_geo_area.lower()]
            print(f"Using file pattern: {file_pattern}")
            
            affordability_files_found = []
            for file in os.listdir(dir_path):
                if file_pattern.lower() in file.lower():
                    affordability_files_found.append(file)
            
            print(f"Affordability files found: {affordability_files_found}")
            
            for file in affordability_files_found:
                file_path = os.path.join(dir_path, file)
                print(f"Processing affordability file: {file_path}")
                df = self.read_data_file(file_path)
                
                if df is not None and not df.empty:
                    # Find the geographic column
                    geo_col = self.find_geographic_column(df, self.selected_geo_area)
                    
                    if geo_col:
                        print(f"Using geographic column: {geo_col}")
                        # Ensure both values are strings for comparison
                        df[geo_col] = df[geo_col].astype(str)
                        selected_name_str = str(self.selected_geo_name)
                        print(f"Looking for exact match: '{selected_name_str}'")
                        
                        # Check for exact match
                        df_filtered = df[df[geo_col] == selected_name_str]
                        if df_filtered.empty:
                            print(f"No exact match found. Trying to find partial matches...")
                            
                            # Try partial match
                            matches = []
                            for value in df[geo_col].dropna().unique():
                                if selected_name_str.lower() in value.lower() or value.lower() in selected_name_str.lower():
                                    matches.append(value)
                            
                            if matches:
                                print(f"Potential partial matches: {matches}")
                                best_match = matches[0]  # Use the first match for simplicity
                                print(f"Using best match: '{best_match}'")
                                df_filtered = df[df[geo_col] == best_match]
                            else:
                                print("No partial matches found either")
                        else:
                            print(f"Found exact match with {len(df_filtered)} rows")
                        
                        if not df_filtered.empty:
                            print(f"Filtered data columns: {df_filtered.columns.tolist()}")
                            
                            # If we have a month column, get the most recent month
                            if 'month' in df_filtered.columns:
                                df_filtered['month'] = pd.to_datetime(df_filtered['month'], errors='coerce')
                                latest_month = df_filtered['month'].max()
                                df_latest = df_filtered[df_filtered['month'] == latest_month]
                                print(f"Filtered to latest month: {latest_month}")
                            else:
                                df_latest = df_filtered
                                
                            # Find affordability column - look for keywords
                            pct_col = None
                            
                            # First priority: direct affordability columns
                            affordability_columns = [col for col in df_latest.columns if 'affordability' in col.lower()]
                            if affordability_columns:
                                # Prefer 3-month affordability for stability
                                if 'rental_affordability_3mo' in affordability_columns:
                                    pct_col = 'rental_affordability_3mo'
                                elif 'rental_affordability_1mo' in affordability_columns:
                                    pct_col = 'rental_affordability_1mo'
                                else:
                                    pct_col = affordability_columns[0]  # Take the first one if no preferred column
                            
                            # If no direct affordability column, try to find rent-to-income ratio
                            if not pct_col:
                                for keywords in [['rent', 'income'], ['rental', 'affordability'], ['income', 'rent']]:
                                    for col in df_latest.columns:
                                        if all(keyword.lower() in col.lower() for keyword in keywords):
                                            pct_col = col
                                            break
                                    if pct_col:
                                        break
                            
                            # If still no column, try to calculate it ourselves if we have rent and income columns
                            if not pct_col:
                                median_rent_cols = [col for col in df_latest.columns if 'median_rent' in col.lower() and not any(x in col.lower() for x in ['growth', 'increase', 'change'])]
                                income_cols = [col for col in df_latest.columns if 'income' in col.lower() and 'index' not in col.lower()]
                                
                                if median_rent_cols and income_cols and len(df_latest) > 0:
                                    # Prefer 3-month median rent for stability
                                    rent_col = None
                                    for prefix in ['median_rent_3mo', 'median_rent_1mo', 'median_rent']:
                                        matching_cols = [col for col in median_rent_cols if col.startswith(prefix)]
                                        if matching_cols:
                                            rent_col = matching_cols[0]
                                            break
                                    
                                    if not rent_col and median_rent_cols:
                                        rent_col = median_rent_cols[0]
                                    
                                    # Use the first income column we find
                                    income_col = income_cols[0]
                                    
                                    if rent_col and income_col:
                                        rent_value = float(df_latest[rent_col].iloc[0]) if not pd.isna(df_latest[rent_col].iloc[0]) else 0
                                        income_value = float(df_latest[income_col].iloc[0]) if not pd.isna(df_latest[income_col].iloc[0]) else 0
                                        
                                        if income_value > 0:
                                            # Calculate weekly rental affordability (rent/income)
                                            # For weekly rent and annual income: (weekly_rent * 52) / annual_income * 100
                                            affordability = (rent_value * 52) / income_value * 100
                                            
                                            # Create a placeholder dataframe with our calculated value
                                            df_latest['calculated_affordability'] = affordability
                                            pct_col = 'calculated_affordability'
                                            print(f"Calculated affordability manually: {affordability:.1f}%")
                            
                            # Find improvement column
                            improvement_col = None
                            for col_suffix in ['improvement', 'change', 'growth']:
                                for col in df_latest.columns:
                                    if col_suffix in col.lower() and any(x in col.lower() for x in ['annual', 'yearly']):
                                        improvement_col = col
                                        break
                                if improvement_col:
                                    break
                            
                            print(f"Affordability column: {pct_col}")
                            print(f"Annual improvement column: {improvement_col}")
                            
                            # Extract data
                            if pct_col and len(df_latest) > 0:
                                pct_value = float(df_latest[pct_col].iloc[0]) if not pd.isna(df_latest[pct_col].iloc[0]) else 0
                                
                                # Ensure the value is properly formatted as a percentage
                                if pct_value > 0 and pct_value < 1:
                                    pct_value = pct_value * 100  # Convert decimal to percentage
                                
                                # Get annual increase if available
                                annual_improvement = None
                                if improvement_col and len(df_latest) > 0:
                                    annual_improvement_value = df_latest[improvement_col].iloc[0]
                                    if not pd.isna(annual_improvement_value):
                                        annual_improvement = float(annual_improvement_value) * 100 if float(annual_improvement_value) < 1 else float(annual_improvement_value)
                                
                                # If we couldn't find an annual improvement column but have multiple months of data,
                                # try to calculate it manually by comparing with previous year
                                if annual_improvement is None and 'month' in df_filtered.columns:
                                    try:
                                        # Get current month's value
                                        current_value = pct_value
                                        
                                        # Try to find data from a year ago
                                        one_year_ago = latest_month - pd.DateOffset(months=12)
                                        one_year_ago_str = one_year_ago.strftime("%b-%Y")  # e.g., "Apr-2024"
                                        
                                        year_ago_data = df_filtered[df_filtered['month'] == one_year_ago]
                                        
                                        if not year_ago_data.empty and pct_col in year_ago_data.columns:
                                            year_ago_value = float(year_ago_data[pct_col].iloc[0]) if not pd.isna(year_ago_data[pct_col].iloc[0]) else 0
                                            if year_ago_value > 0 and year_ago_value < 1:
                                                year_ago_value = year_ago_value * 100  # Convert decimal to percentage
                                            
                                            # Calculate change - note that for affordability, a decrease is an improvement
                                            # (lower percentage of income spent on rent is better)
                                            change = current_value - year_ago_value
                                            annual_improvement = -change  # Negative change = improvement for affordability
                                            
                                            print(f"Calculated annual affordability change: {change:.2f}% (improvement: {annual_improvement:.2f}%)")
                                            print(f"Comparing {latest_month.strftime('%b-%Y')} ({current_value:.1f}%) to {one_year_ago_str} ({year_ago_value:.1f}%)")
                                        else:
                                            print(f"No data found for {one_year_ago_str} to calculate annual improvement")
                                    except Exception as e:
                                        print(f"Error calculating annual improvement: {str(e)}")
                                
                                # Use default annual change if we couldn't calculate it
                                if annual_improvement is None:
                                    annual_improvement = 0
                                
                                self.data["affordability"] = {
                                    "percentage": round(pct_value, 1),
                                    "period": latest_month.strftime("%b-%Y") if latest_month is not None else "Apr-2025",
                                    "source": "NSW Fair Trading Prop Track Data",
                                    "annual_improvement": round(annual_improvement, 2) if annual_improvement is not None else 0,
                                    "comparison_gs": self.GS_REFERENCE_DATA["affordability"],
                                    "comparison_ron": self.RON_REFERENCE_DATA["affordability"]
                                }
                    break
        except Exception as e:
            print(f"Error collecting affordability data: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If no data collected, use defaults
        if "affordability" not in self.data:
            print("Using DEFAULT affordability data")
            self.data["affordability"] = {
                "percentage": 43.6,
                "period": "Apr-25",
                "source": "NSW Fair Trading Prop Track Data",
                "annual_improvement": 0,
                "comparison_gs": self.GS_REFERENCE_DATA["affordability"],
                "comparison_ron": self.RON_REFERENCE_DATA["affordability"]
            }
            
        # Removed rental contacts data as requested
        
        print("\nData collection complete")
        print(f"{'='*50}\n")
    
    def find_column(self, df, keywords):
        """Find column that contains all the keywords (case-insensitive)"""
        print(f"Looking for column with keywords: {keywords}")
        
        columns_found = []
        for col in df.columns:
            if all(keyword.lower() in str(col).lower() for keyword in keywords):
                columns_found.append(col)
                print(f"Found matching column: {col}")
        
        if columns_found:
            # Return the first match
            return columns_found[0]
            
        # Try with partial matching if no exact match
        partial_matches = []
        for col in df.columns:
            col_lower = str(col).lower()
            match_score = sum(1 for kw in keywords if kw.lower() in col_lower)
            if match_score > 0:
                partial_matches.append((col, match_score))
        
        # Sort by match score
        partial_matches.sort(key=lambda x: x[1], reverse=True)
        
        if partial_matches:
            best_match = partial_matches[0][0]
            print(f"No exact match found. Using best partial match: {best_match}")
            return best_match
            
        print(f"No matching column found for keywords: {keywords}")
        return None
    
    def generate_comparison_comment(self, metric, value, comparison_gs, comparison_ron):
        """Generate a comparison comment for a metric that shows both Greater Sydney and Rest of NSW references"""
        
        if metric == "renters":
            gs_text = ""
            if value < comparison_gs["value"] - 1:  # 1% buffer to avoid "slightly lower" for small differences
                gs_text = f"lower than the Greater Sydney average of {comparison_gs['value']}%"
            elif value > comparison_gs["value"] + 1:
                gs_text = f"higher than the Greater Sydney average of {comparison_gs['value']}%"
            else:
                gs_text = f"similar to the Greater Sydney average of {comparison_gs['value']}%"
                
            ron_text = ""
            if value < comparison_ron["value"] - 1:
                ron_text = f"and lower than the Rest of NSW average of {comparison_ron['value']}%"
            elif value > comparison_ron["value"] + 1:
                ron_text = f"and higher than the Rest of NSW average of {comparison_ron['value']}%"
            else:
                ron_text = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
                
            return f"{self.selected_geo_name} ({self.selected_geo_area}) has a concentration of renters that is {gs_text} {ron_text}."
        
        elif metric == "social_housing":
            gs_text = ""
            if value < comparison_gs["value"] - 0.5:  # 0.5% buffer
                gs_text = f"lower than the Greater Sydney average of {comparison_gs['value']}%"
            elif value > comparison_gs["value"] + 0.5:
                gs_text = f"higher than the Greater Sydney average of {comparison_gs['value']}%"
            else:
                gs_text = f"similar to the Greater Sydney average of {comparison_gs['value']}%"
                
            ron_text = ""
            if value < comparison_ron["value"] - 0.5:
                ron_text = f"and lower than the Rest of NSW average of {comparison_ron['value']}%"
            elif value > comparison_ron["value"] + 0.5:
                ron_text = f"and higher than the Rest of NSW average of {comparison_ron['value']}%"
            else:
                ron_text = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
                
            return f"{self.selected_geo_name} ({self.selected_geo_area}) has a concentration of social housing that is {gs_text} {ron_text}."
        
        elif metric == "median_rent":
            local_increase = self.data["median_rent"]["annual_increase"]
            if pd.isna(local_increase):
                local_increase = 0
                
            gs_text = ""
            if local_increase < comparison_gs["value"] - 1:  # 1% buffer
                gs_text = f"lower than Greater Sydney's annual increase of {comparison_gs['value']}%"
            elif local_increase > comparison_gs["value"] + 1:
                gs_text = f"higher than Greater Sydney's annual increase of {comparison_gs['value']}%"
            else:
                gs_text = f"similar to Greater Sydney's annual increase of {comparison_gs['value']}%"
                
            ron_text = ""
            if local_increase < comparison_ron["value"] - 1:
                ron_text = f"and lower than Rest of NSW's annual increase of {comparison_ron['value']}%"
            elif local_increase > comparison_ron["value"] + 1:
                ron_text = f"and higher than Rest of NSW's annual increase of {comparison_ron['value']}%"
            else:
                ron_text = f"and similar to Rest of NSW's annual increase of {comparison_ron['value']}%"
                
            return f"{self.selected_geo_name} ({self.selected_geo_area})'s median annual rental increase of {local_increase}% is {gs_text} {ron_text}."
        
        elif metric == "vacancy_rates":
            current_rate = self.data["vacancy_rates"]["value"]
            previous_rate = self.data["vacancy_rates"]["previous_year_rate"]
            
            # Format rates as percentages for display (if they're in decimal format)
            if current_rate < 1 and current_rate > 0:
                current_rate_display = current_rate  # Already as a percentage
            else:
                current_rate_display = current_rate
                
            if previous_rate is not None:
                if previous_rate < 1 and previous_rate > 0:
                    previous_rate_display = previous_rate  # Already as a percentage
                else:
                    previous_rate_display = previous_rate
            
            # Generate text about market tightening/loosening if previous year data available
            trend_text = ""
            if previous_rate is not None:
                if current_rate < previous_rate - 0.1:
                    trend_text = f"The vacancy rate has tightened from {previous_rate_display:.2f}% a year ago to {current_rate_display:.2f}% now. "
                elif current_rate > previous_rate + 0.1:
                    trend_text = f"The vacancy rate has loosened from {previous_rate_display:.2f}% a year ago to {current_rate_display:.2f}% now. "
                else:
                    trend_text = f"The vacancy rate has remained stable at around {current_rate_display:.2f}% compared to {previous_rate_display:.2f}% a year ago. "
            
            # Add comparisons to Greater Sydney and Rest of NSW
            comparison_text = f"For reference, Greater Sydney has experienced a change of {comparison_gs['value']}% and Rest of NSW {comparison_ron['value']}% over the past year."
            
            return trend_text + comparison_text
        
        elif metric == "affordability":
            local_improvement = self.data["affordability"]["annual_improvement"]
            if pd.isna(local_improvement):
                local_improvement = 0
                
            local_pct = self.data["affordability"]["percentage"]
            
            # Compare with Greater Sydney
            gs_comparison = ""
            if local_pct > comparison_gs["value"] + 2:  # 2% buffer
                gs_comparison = f"less affordable than the Greater Sydney average of {comparison_gs['value']}%"
            elif local_pct < comparison_gs["value"] - 2:
                gs_comparison = f"more affordable than the Greater Sydney average of {comparison_gs['value']}%"
            else:
                gs_comparison = f"similar to the Greater Sydney average of {comparison_gs['value']}%"
            
            # Compare with Rest of NSW
            ron_comparison = ""
            if local_pct > comparison_ron["value"] + 2:
                ron_comparison = f"and less affordable than the Rest of NSW average of {comparison_ron['value']}%"
            elif local_pct < comparison_ron["value"] - 2:
                ron_comparison = f"and more affordable than the Rest of NSW average of {comparison_ron['value']}%"
            else:
                ron_comparison = f"and similar to the Rest of NSW average of {comparison_ron['value']}%"
            
            # Evaluate the trend
            if local_improvement > 0.1:  # Improvement
                change_text = f"an improvement of {abs(local_improvement)}%"
            elif local_improvement < -0.1:  # Deterioration
                change_text = f"a deterioration of {abs(local_improvement)}%"
            else:
                change_text = "relatively stable affordability"
            
            return (f"{self.selected_geo_name} ({self.selected_geo_area}) rental affordability is {gs_comparison} {ron_comparison}, "
                   f"with {change_text} over the past year. Greater Sydney had a change of {abs(comparison_gs['annual_change'])}% "
                   f"while Rest of NSW had a change of {abs(comparison_ron['annual_change'])}%.")
        
        return ""
    
    def create_excel_output(self):
        """Create a nicely formatted Excel output with the analysis"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.selected_geo_name} Analysis"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        metric_font = Font(bold=True, size=11)
        metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        metric_alignment = Alignment(vertical="center", wrap_text=True)
        
        value_font = Font(size=11)
        value_alignment = Alignment(vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Rental Market Analysis for {self.selected_geo_name} ({self.selected_geo_area})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers - Row 3
        headers = ["Metric", "Value", "Period", "Source", "Comment"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        
        # Add metrics data
        row = 4
        
        # Renters
        ws.cell(row=row, column=1).value = "# and % of renters"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['renters']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['renters']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['renters']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("renters", self.data['renters']['percentage'], 
                                              self.data['renters']['comparison_gs'], self.data['renters']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['renters']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Social Housing
        row += 1
        ws.cell(row=row, column=1).value = "# and % of Social Housing"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['percentage']}% of all residential dwellings"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['social_housing']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['social_housing']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("social_housing", self.data['social_housing']['percentage'], 
                                                self.data['social_housing']['comparison_gs'], self.data['social_housing']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        ws.cell(row=row, column=2).value = f"{self.data['social_housing']['count']:,}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Median Weekly Rent
        row += 1
        ws.cell(row=row, column=1).value = "Median Weekly Rent"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"${self.data['median_rent']['value']}"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['median_rent']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['median_rent']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("median_rent", self.data['median_rent']['value'], 
                                            self.data['median_rent']['comparison_gs'], self.data['median_rent']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        # Show both annual increase and the previous year's rent
        annual_increase = self.data['median_rent']['annual_increase']
        prev_year_rent = self.data['median_rent']['previous_year_rent']
        
        display_text = f"Annual increase {annual_increase}%"
        if prev_year_rent is not None:
            display_text += f" (from ${prev_year_rent} last year)"
            
        ws.cell(row=row, column=2).value = display_text
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Vacancy Rates
        row += 1
        ws.cell(row=row, column=1).value = "Vacancy Rates"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        # Format vacancy rate value - if it's between 0 and 1, it's likely already a percentage (e.g., 0.75%)
        vacancy_value = self.data['vacancy_rates']['value']
        if vacancy_value < 1 and vacancy_value > 0:
            formatted_vacancy = f"{vacancy_value:.2f}%"
        else:
            formatted_vacancy = f"{vacancy_value:.2f}%"
            
        ws.cell(row=row, column=2).value = formatted_vacancy
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['vacancy_rates']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['vacancy_rates']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("vacancy_rates", self.data['vacancy_rates']['value'], 
                                             self.data['vacancy_rates']['comparison_gs'], self.data['vacancy_rates']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        previous_year_rate = self.data['vacancy_rates']['previous_year_rate']
        
        # Format previous year rate - check if it's already a percentage
        if previous_year_rate is not None:
            if previous_year_rate < 1 and previous_year_rate > 0:
                previous_year_text = f"Previous year: {previous_year_rate:.2f}%"
            else:
                previous_year_text = f"Previous year: {previous_year_rate:.2f}%"
        else:
            previous_year_text = "Previous year data not available"
            
        ws.cell(row=row, column=2).value = previous_year_text
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Rental Affordability
        row += 1
        ws.cell(row=row, column=1).value = "Rental affordability*"
        ws.cell(row=row, column=1).font = metric_font
        ws.cell(row=row, column=1).fill = metric_fill
        ws.cell(row=row, column=1).alignment = metric_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2).value = f"{self.data['affordability']['percentage']}% of income on rent"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3).value = self.data['affordability']['period']
        ws.cell(row=row, column=3).font = value_font
        ws.cell(row=row, column=3).alignment = value_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4).value = self.data['affordability']['source']
        ws.cell(row=row, column=4).font = value_font
        ws.cell(row=row, column=4).alignment = value_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        comment = self.generate_comparison_comment("affordability", self.data['affordability']['percentage'], 
                                              self.data['affordability']['comparison_gs'], self.data['affordability']['comparison_ron'])
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = value_font
        ws.cell(row=row, column=5).alignment = value_alignment
        ws.cell(row=row, column=5).border = thin_border
        
        row += 1
        annual_improvement = self.data['affordability']['annual_improvement']
        improvement_text = "improvement" if annual_improvement > 0 else "deterioration"
        ws.cell(row=row, column=2).value = f"Annual {improvement_text} {abs(annual_improvement)}%"
        ws.cell(row=row, column=2).font = value_font
        ws.cell(row=row, column=2).alignment = value_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        for col in [1, 3, 4, 5]:
            ws.cell(row=row, column=col).border = thin_border
        
        # Rental Contacts section has been removed as requested
        
        # Add footnote for affordability
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].value = ("* Methodology: the rental affordability is calculated by taking median weekly rental household incomes for the "
                              "geographic area and comparing that to median weekly rents for the same area. Any number high than 30% of income "
                              "on rent is considered that a household is experiencing rental stress. This metric is calculated by Fair Trading "
                              "using ABS income and indexation data as well as Core Logic rental data.")
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Set row heights
        for r in range(3, row):
            ws.row_dimensions[r].height = 30
        
        ws.row_dimensions[row].height = 45  # Footnote row
        
        # Save the workbook
        wb.save(self.output_file)


def main():
    # Set pandas to not display warnings
    pd.options.mode.chained_assignment = None
    
    try:
        analyzer = RentalDataAnalyzer()
        analyzer.create_gui()
    except Exception as e:
        import traceback
        print(f"Critical error: {str(e)}")
        print(traceback.format_exc())
        
        # Show error in GUI if possible
        try:
            import tkinter.messagebox as msgbox
            msgbox.showerror("Critical Error", f"An unexpected error occurred:\n\n{str(e)}\n\nPlease check the console for details.")
        except:
            pass

if __name__ == "__main__":
    main()